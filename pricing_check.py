import logging

import pandas as pd
from datetime import datetime, timedelta
from services.excel import OpenPyXLFileHandler
from services.google_sheets_service import GoogleSheetsService

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def load_inventory_items(file_handler):
    """Load inventory items into a dictionary from all sheets in the workbook.

    Args:
        file_handler (OpenPyXLFileHandler): An object handling the Excel file.

    Returns:
        dict: A dictionary mapping codes (column B) to supplier codes (column AB).
    """
    workbook = file_handler.workbook  # Access the workbook from the handler
    _inventory_dict = {}

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Check if 'PkId' is in cell A2
        a2_value = worksheet['A2'].value
        if a2_value != "PkId":
            logging.info(f"Skipping sheet: {sheet} - 'PkId' not found in A2.")
            continue

        for row in worksheet.iter_rows(min_row=3, values_only=True):  # Start from row 3 (data rows)
            code = row[1]  # Column B (0-indexed, so it's at index 1)
            supplier_code = row[27]  # Column AB (0-indexed, so it's at index 27)
            if code and supplier_code:  # Ensure both fields are not None
                _inventory_dict[code] = supplier_code

    return _inventory_dict


def load_google_sheet_data(google_sheets_service, spreadsheet_id, range_name):
    """Load Google Sheet data into a dictionary with adjusted pricing logic.

    Args:
        google_sheets_service (GoogleSheetsService): An instance of the GoogleSheetsService class.
        spreadsheet_id (str): The google sheet id.
        range_name (str): The range from the google sheet.

    Returns:
        dict: A dictionary mapping codes (column A) to adjusted prices.
    """
    sheet_data = google_sheets_service.fetch_sheet_data(spreadsheet_id, range_name)  # Get data from Google Sheets
    _google_sheet_dict = {}

    for row in sheet_data:
        code = row[0]  # Column A
        raw_price = row[28] if row[28] else row[15]  # Column AC (index 28), fallback to column P (index 15)
        unit = row[4]  # Column E
        divisor = row[32]  # Column AG (index 32)

        if code and raw_price:
            # Adjust the price based on the unit type
            if unit == "SQM":
                divisor = 1
            else:
                if not divisor or divisor == 0:
                    logging.error(f"Width required but missing for Unleashed code {code}")
                    continue

            try:
                raw_price = float(raw_price)
                divisor = float(divisor)
            except ValueError:
                logging.error(f"Invalid price value for {code}: {raw_price} with UOM {unit} and width {divisor}")
                continue

            if divisor == 0:
                logging.error(f"Width still zero for {code}: {raw_price} with UOM {unit}. Why?!")
                continue

            _google_sheet_dict[code] = raw_price / divisor

    return _google_sheet_dict


def process_pricing_file(pricing_file_path, inventory_dict, google_sheet_dict, tolerance=0.005):
    """Process the pricing Excel file."""
    # Use openpyxl to handle the cell-based logic
    from openpyxl import load_workbook
    wb = load_workbook(pricing_file_path)
    modified_sheets = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Check if 'PkId' is in cell A1
        a1_value = sheet['A1'].value
        if a1_value != "PkId":
            logging.info(f"Deleting sheet: {sheet_name} - 'PkId' not found in A1.")
            wb.remove(sheet)  # Remove the sheet
            continue

        # Load the sheet into pandas for processing
        sheet_df = pd.DataFrame(sheet.values)
        rows_to_keep = []

        for _, row in sheet_df.iterrows():
            code = row[1]
            supplier_code = inventory_dict.get(code)

            if not supplier_code:
                # Skip rows without a matching supplier code
                logging.warning(f"No matching supplier code ({supplier_code}) found for {code}")
                continue

            # Perform pricing lookup
            google_price = google_sheet_dict.get(supplier_code)
            if google_price is None:
                # Skip rows without a matching price
                logging.warning(f"Supplier code ({supplier_code}) for {code} not found in Unleashed")
                continue

            # Find first non-zero price in columns M-Q
            price_columns = [12, 13, 14, 15, 16]
            prices = [row[col] for col in price_columns if row[col] > 0]

            if len(prices) == 0:
                logging.warning(f"No prices found in {price_columns} for row {code}.")
                continue
            elif len(prices) > 1:
                logging.warning(f"Multiple non-zero prices found in {price_columns} for row {code}. Taking first value.")

            excel_price = prices[0]

            # Compare prices
            if abs(google_price - excel_price) / google_price > tolerance:
                # Modify the row
                row['PkId'] = None
                row['Operation'] = 'A'
                row['Date From'] = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
                rows_to_keep.append(row)

        # Save the modified sheet
        if rows_to_keep:
            modified_sheets[sheet_name] = pd.DataFrame(rows_to_keep)

    # Save updated Excel file
    with pd.ExcelWriter(pricing_file_path, mode='w') as writer:
        for sheet_name, modified_df in modified_sheets.items():
            modified_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Save the workbook after removing sheets
    wb.save(pricing_file_path)


# Example usage
g_inventory_file = OpenPyXLFileHandler(file_path='uploads/items.xlsm')
g_inventory_file.load_workbook()
g_google_sheet_file = GoogleSheetsService(f'./static/buz-app-439103-b6ae046c4723.json')
g_pricing_file = 'uploads/pricing.xlsm'

g_inventory_dict = load_inventory_items(g_inventory_file)
g_google_sheet_dict = load_google_sheet_data(
            g_google_sheet_file,
            "1OE7CnACQjFFGgIm_yFjN8Qhf0l5xEQkJqXwFV4plNlI",
            "Data!A:AG"
)
process_pricing_file(g_pricing_file, g_inventory_dict, g_google_sheet_dict)
