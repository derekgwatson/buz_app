from datetime import datetime
from openpyxl import Workbook
from excel import OpenPyXLFileHandler
from google_sheets_service import GoogleSheetsService


def process_inventory_backorder_with_services(
    _file_handler: OpenPyXLFileHandler,
    _sheets_service: GoogleSheetsService,
    spreadsheet_id: str,
    range_name: str,
    original_filename: str,
    upload_filename: str,
    header_row: int = 1,
):
    """
    Process inventory items for backorder messages using custom service classes and generate upload and original files.

    Args:
        _file_handler (OpenPyXLFileHandler): Instance of OpenPyXLFileHandler for handling Excel files.
        _sheets_service (GoogleSheetsService): Instance of GoogleSheetsService for interacting with Google Sheets.
        spreadsheet_id (str): Google Sheet ID containing supplier codes and backorder dates.
        range_name (str): Range in the Google Sheet to fetch data from.
        original_filename (str): Path to save the original rows file.
        upload_filename (str): Path to save the resulting upload file.
        header_row (int): Row that header is on.
    """
    # Step 1: Load the inventory workbook
    _file_handler.load_workbook()

    # Step 2: Read supplier codes and backorder dates from Google Sheets
    google_sheet_data = _sheets_service.fetch_sheet_data(spreadsheet_id, range_name)
    print(f"Google Sheet Data: {google_sheet_data}")

    # Extract headers and map them to indices
    headers = google_sheet_data[0]
    header_map = {header: index for index, header in enumerate(headers)}

    # Ensure required columns exist
    required_columns = ['Unleashed Code', 'On backorder until']
    missing_columns = [col for col in required_columns if col not in header_map]
    if missing_columns:
        raise ValueError(f"Missing required columns in Google Sheet: {', '.join(missing_columns)}")

    # Convert Google Sheet data into a lookup dictionary
    backorder_lookup = {
        row[header_map['Unleashed Code']]: row[header_map['On backorder until']]
        for row in google_sheet_data[1:]  # Skip the header row
        if row[header_map['Unleashed Code']] and row[header_map['On backorder until']]  # Ensure non-empty values
    }

    # Step 3: Process each sheet in the inventory workbook
    upload_workbook = Workbook()
    original_workbook = Workbook()
    upload_workbook.remove(upload_workbook.active)  # Remove the default sheet
    original_workbook.remove(original_workbook.active)

    # Read all sheets at once
    inventory_data = _file_handler.read_sheet_to_dict(header_row=header_row)
    processed_sheets = set()  # Track sheets that are included in the upload file

    for sheet_name, rows in inventory_data.items():
        # Check if the sheet has "PkId" in cell A2
        first_row = rows[0] if rows else {}  # Get the first row if it exists
        if first_row.get("PkId", None) is None:
            print(f"Skipping sheet {sheet_name} as it does not have 'PkId' in first column.")
            continue

        updated_rows = []
        original_rows = []

        for row in rows:
            original_row = row.copy()  # Preserve the original row before any changes
            supplier_code = row.get('Supplier Product Code')
            warning_message = row.get('Warning', '')

            # Concatenate columns D, E, and F with spaces
            # Safely concatenate columns D, E, and F with spaces
            description_parts = [
                (row.get("DescnPart1 (Material)", "") or "").strip(),
                (row.get("DescnPart2 (Material Types)", "") or "").strip(),
                (row.get("DescnPart3 (Colour)", "") or "").strip(),
            ]
            description = " ".join(part for part in description_parts if part)

            # Check for existing backorder message
            if "on backorder until" in warning_message:
                existing_date_str = ""
                try:
                    # Parse the date from the existing message (new format: "d mmm yyyy")
                    existing_date_str = warning_message.split("until")[1].strip().rstrip(".")
                    existing_date = datetime.strptime(existing_date_str, "%d %b %Y")  # New format
                    if existing_date < datetime.now():
                        # Remove the outdated warning
                        row['Warning'] = ''
                        row['Operation'] = 'E'
                        updated_rows.append(row)
                        original_rows.append(original_row)  # Keep the unmodified row in the original file
                        continue
                except ValueError:
                    # If parsing fails, skip date comparison
                    print(f"Failed to parse date: {existing_date_str}")
                    pass

            if supplier_code in backorder_lookup:
                # Format backorder date
                raw_date = backorder_lookup[supplier_code]
                try:
                    formatted_date = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%d %b %Y")
                except ValueError:
                    formatted_date = raw_date  # Fallback to raw value if formatting fails

                backorder_message = f"{description} on backorder until {formatted_date}."
                row['Warning'] = backorder_message  # Update the warning field
                row['Operation'] = 'E'
                updated_rows.append(row)
                original_rows.append(original_row)  # Keep the unmodified row in the original file

        # Only add updated rows to the upload workbook
        if updated_rows:
            processed_sheets.add(sheet_name)
            upload_sheet = upload_workbook.create_sheet(title=sheet_name)
            upload_sheet.append(list(updated_rows[0].keys()))  # Write header
            for updated_row in updated_rows:
                upload_sheet.append(list(updated_row.values()))

        # Add original rows to the original workbook
        if original_rows:
            original_sheet = original_workbook.create_sheet(title=sheet_name)
            original_sheet.append(list(original_rows[0].keys()))  # Write header
            for original_row in original_rows:
                original_sheet.append(list(original_row.values()))

    # Step 4: Save the workbooks
    upload_workbook.save(upload_filename)
    original_workbook.save(original_filename)
    print(f"Upload file created: {upload_filename}")
    print(f"Original values file created: {original_filename}")
