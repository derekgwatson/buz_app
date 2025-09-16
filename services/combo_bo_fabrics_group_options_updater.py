from typing import Dict, List, Tuple, Set
from openpyxl import load_workbook, Workbook
from .database import DatabaseManager
import re

SHEETS = ["ROLLCB", "WSROLLCB", "ROLLFLEX", "WSROLLFLEX"]
FLEX_SHEETS = {"ROLLFLEX", "WSROLLFLEX"}
START_ROW = 17


def _norm_text(s: str | None) -> str:
    """Trim, collapse whitespace, convert NBSP to space."""
    if s is None:
        return ""
    s = str(s).replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


class ComboBOFabricsGroupOptionsUpdater:
    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager

    # === DB ===
    def _fetch_db_triples(self, group_code: str) -> List[Tuple[str, str, str]]:
        """
        Return list of (fabric, colour, code) for active Blockout items in a group.
        """
        sql = """
            SELECT DescnPart1, DescnPart3, Code
            FROM inventory_items
            WHERE inventory_group_code = ?
              AND lower(DescnPart2) = 'blockout'
              AND CAST(Active AS INTEGER) = 1
        """
        cur = self.db_manager.execute_query(sql, (group_code,))
        return [
            (
                _norm_text(row["DescnPart1"] or ""),
                _norm_text(row["DescnPart3"] or ""),
                (row["Code"] or "").strip(),
            )
            for row in cur.fetchall()
        ]

    # === Workbook read ===
    def _load_wb_lists(
        self, ws, sheet_name: str
    ) -> Tuple[Set[str], Set[Tuple[str, str, str]]]:
        """
        Read BLOCKOUTFABRIC (plain names) and BLOCKOUTFABRICCOLOUR (fabric|colour|code)
        as two independent sets. Robust to blank rows by scanning until a long empty streak.
        """
        fabric_col = colour_col = None
        for cell in ws[1]:
            v = str(cell.value).strip().upper() if cell.value else ""
            if v == "BLOCKOUTFABRIC":
                fabric_col = cell.column
            elif v == "BLOCKOUTFABRICCOLOUR":
                colour_col = cell.column
        if fabric_col is None or colour_col is None:
            raise ValueError(f"{ws.title}: missing BLOCKOUTFABRIC / BLOCKOUTFABRICCOLOUR headers")

        def strip_yes(s: str) -> str:
            s = s.strip()
            return s[4:].strip() if (sheet_name in FLEX_SHEETS and s.upper().startswith("YES|")) else s

        # --- Fabrics column ---
        wb_fabrics: Set[str] = set()
        r = START_ROW
        while True:
            val = ws.cell(row=r, column=fabric_col).value
            if val is None or str(val).strip() == "":
                break

            name = strip_yes(str(val).strip())
            if name:
                wb_fabrics.add(_norm_text(name).lower())

            r += 1

        # --- Colour triples column ---
        wb_triples: Set[Tuple[str, str, str]] = set()
        r = START_ROW
        while True:
            val = ws.cell(row=r, column=colour_col).value
            if val is None or str(val).strip() == "":
                break
            txt = str(val).strip()
            if "|" in txt:
                parts = [p.strip() for p in txt.split("|")]
                if len(parts) >= 2:
                    f = _norm_text(parts[0]).lower()
                    c = _norm_text(parts[1]).lower()
                    code = parts[2].strip() if len(parts) >= 3 else ""
                    wb_triples.add((f, c, code))
            r += 1

        return wb_fabrics, wb_triples

    # === Overwrite ===
    def _overwrite_columns(
        self, ws, sheet_name: str, fabrics: List[str], triples: List[Tuple[str, str, str]]
    ):
        """
        Overwrite BLOCKOUTFABRIC and BLOCKOUTFABRICCOLOUR independently.
        """
        fabric_col = colour_col = None
        for cell in ws[1]:
            v = str(cell.value).strip().upper() if cell.value else ""
            if v == "BLOCKOUTFABRIC":
                fabric_col = cell.column
            elif v == "BLOCKOUTFABRICCOLOUR":
                colour_col = cell.column
        if fabric_col is None or colour_col is None:
            raise ValueError(f"{ws.title}: missing BLOCKOUTFABRIC / BLOCKOUTFABRICCOLOUR headers")

        # Clear both columns
        for r in range(START_ROW, ws.max_row + 1):
            ws.cell(row=r, column=fabric_col).value = None
            ws.cell(row=r, column=colour_col).value = None

        # Write fabrics (with YES| for Flex)
        for i, f in enumerate(fabrics, start=START_ROW):
            out = f"YES|{f}" if (sheet_name in FLEX_SHEETS and f) else f
            ws.cell(row=i, column=fabric_col).value = out

        # Write triples
        for j, (f, c, code) in enumerate(triples, start=START_ROW):
            ws.cell(row=j, column=colour_col).value = f"{f}|{c}|{code}"

    # === Main ===
    def update_options_file(self, in_path: str, out_path: str) -> Dict[str, Dict]:
        wb_in = load_workbook(in_path, data_only=True)
        wb_out = Workbook()
        for s in wb_out.worksheets:
            wb_out.remove(s)

        summary: Dict[str, Dict] = {}

        for sheet in SHEETS:
            if sheet not in wb_in.sheetnames:
                summary[sheet] = {"status": "missing_in_input"}
                continue

            ws = wb_in[sheet]
            base_group = "ROLL" if sheet == "ROLLCB" else "WSROLL" if sheet == "WSROLLCB" else sheet

            # DB -> normalized sets
            db_triples = self._fetch_db_triples(base_group)
            db_triples_norm = {
                (_norm_text(f).lower(), _norm_text(c).lower(), code.strip())
                for (f, c, code) in db_triples
            }

            # WB -> normalized sets
            wb_fabrics, wb_triples = self._load_wb_lists(ws, sheet)

            # fabric sets
            db_fabrics: Set[str] = {_norm_text(f).lower() for f, _, _ in db_triples}
            wb_fabrics_from_col: Set[str] = set(wb_fabrics)  # from BLOCKOUTFABRIC (left col)
            wb_fabrics_from_triples: Set[str] = {f for (f, _, _) in wb_triples}  # inferred from triples (right col)

            print("=" * 60)
            print(f"DEBUG for sheet {sheet}")
            print(f"Base group: {base_group}")
            print(f"DB triple count: {len(db_triples)}")
            print(f"WB triple count: {len(wb_triples)}")
            print(f"DB fabrics count: {len(db_fabrics)}")
            print(f"WB fabrics (left col) count: {len(wb_fabrics_from_col)}")
            print(f"WB fabrics (from triples) count: {len(wb_fabrics_from_triples)}")

            print("DB triples sample:", list(db_triples)[:5])
            print("WB triples sample:", list(wb_triples)[:5])

            # Focus test (ABC case)
            target = "abc"
            print(f"Has '{target.upper()}'?  "
                  f"DB_fabrics={target in db_fabrics}  "
                  f"WB_fabrics_col={target in wb_fabrics_from_col}  "
                  f"WB_fabrics_from_triples={target in wb_fabrics_from_triples}")

            print(f"DB '{target.upper()}' triples sample:",
                  [t for t in db_triples if _norm_text(t[0]).lower() == target][:3])
            print(f"WB '{target.upper()}' triples sample:",
                  [t for t in wb_triples if t[0] == target][:3])

            # Diffs (CORRECT: compare DB vs left column; and DB vs triples)
            fabrics_added = sorted(db_fabrics - wb_fabrics_from_col)  # should include 'abc' if missing in left col
            fabrics_removed = sorted(wb_fabrics_from_col - db_fabrics)
            triples_added = sorted(db_triples_norm - wb_triples)
            triples_removed = sorted(wb_triples - db_triples_norm)

            print("Missing fabrics (in DB but not WB left col) sample:", fabrics_added[:10])
            print("Extra fabrics (in WB left col but not DB) sample:", fabrics_removed[:10])
            print("Missing triples sample:", list(triples_added)[:5])
            print("Extra triples sample:", list(triples_removed)[:5])

            if not fabrics_added and not fabrics_removed and not triples_added and not triples_removed:
                summary[sheet] = {
                    "status": "unchanged",
                    "fabrics": len(db_fabrics),
                    "triples": len(db_triples_norm),
                }
                continue

            # Copy original and overwrite columns independently
            ws_copy = wb_out.create_sheet(sheet)
            for row in ws.iter_rows(values_only=False):
                ws_copy.append([c.value for c in row])

            fabrics_sorted = sorted({f for (f, _, _) in db_triples}, key=lambda s: _norm_text(s).lower())
            triples_sorted = sorted(
                db_triples,
                key=lambda t: (_norm_text(t[0]).lower(), _norm_text(t[1]).lower(), t[2]),
            )
            self._overwrite_columns(ws_copy, sheet, fabrics_sorted, triples_sorted)

            summary[sheet] = {
                "status": "changed",
                "fabrics_total": len(db_fabrics),
                "triples_total": len(db_triples_norm),
                "fabrics_added": fabrics_added,
                "fabrics_removed": fabrics_removed,
                "triples_added": triples_added,
                "triples_removed": triples_removed,
            }

        if not wb_out.sheetnames:
            ws_dummy = wb_out.create_sheet("NO_CHANGES")
            ws_dummy["A1"] = "No changes detected"

        wb_out.save(out_path)
        return summary
