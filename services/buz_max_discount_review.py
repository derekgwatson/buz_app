# services/buz_max_discount_review.py
from __future__ import annotations

import asyncio
import logging
from pathlib import Path
from typing import Optional, Dict, Any, List
from dataclasses import dataclass
from playwright.async_api import async_playwright, Page, Browser, BrowserContext, Download
import openpyxl
import tempfile
import os

logger = logging.getLogger(__name__)


@dataclass
class InventoryGroupDiscount:
    """Inventory group with max discount"""
    code: str
    description: str
    max_discount_pct: Optional[float]

    def __hash__(self):
        return hash(self.code)


@dataclass
class OrgDiscounts:
    """Discounts for one org"""
    org_name: str
    inventory_groups: List[InventoryGroupDiscount]
    file_path: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            'org_name': self.org_name,
            'file_path': self.file_path,
            'inventory_groups': [
                {
                    'code': ig.code,
                    'description': ig.description,
                    'max_discount_pct': ig.max_discount_pct
                }
                for ig in self.inventory_groups
            ]
        }


class MaxDiscountReviewResult:
    """Result of max discount review"""

    def __init__(self):
        self.orgs: List[OrgDiscounts] = []
        self.steps: list[str] = []

    def add_step(self, message: str):
        """Add a step to the result log"""
        self.steps.append(message)
        logger.info(message)

    def to_dict(self) -> Dict[str, Any]:
        """Convert result to dictionary"""
        return {
            'orgs': [org.to_dict() for org in self.orgs],
            'steps': self.steps
        }


class BuzMaxDiscountReview:
    """Download and compare max discount percentages across Buz orgs"""

    EXPORT_IMPORT_URL = "https://go.buzmanager.com/Settings/ExportImportAllInventorySettings/Create"

    # Org configurations
    ORGS = {
        'canberra': {
            'display_name': 'Watson Blinds (Canberra)',
            'storage_state': '.secrets/buz_storage_state_canberra.json'
        },
        'tweed': {
            'display_name': 'Tweed',
            'storage_state': '.secrets/buz_storage_state_tweed.json'
        },
        'bay': {
            'display_name': 'Batemans Bay',
            'storage_state': '.secrets/buz_storage_state_bay.json'
        },
        'shoalhaven': {
            'display_name': 'Shoalhaven',
            'storage_state': '.secrets/buz_storage_state_shoalhaven.json'
        },
        'wagga': {
            'display_name': 'Wagga Wagga',
            'storage_state': '.secrets/buz_storage_state_wagga.json'
        }
    }

    def __init__(self, output_dir: Path, headless: bool = True):
        """
        Initialize max discount review.

        Args:
            output_dir: Directory to save downloaded Excel files
            headless: Run browser in headless mode
        """
        self.output_dir = output_dir
        self.headless = headless
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.playwright = None
        self.result = MaxDiscountReviewResult()

    async def __aenter__(self):
        """Context manager entry - launch browser"""
        self.playwright = await async_playwright().__aenter__()
        self.browser = await self.playwright.chromium.launch(headless=self.headless)
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - close browser"""
        if self.context:
            await self.context.close()
        if self.browser:
            await self.browser.close()

    async def switch_to_org(self, org_key: str):
        """
        Switch to a different Buz org by creating a new browser context.

        Args:
            org_key: Key from ORGS dict (e.g., 'canberra', 'tweed')
        """
        org_config = self.ORGS[org_key]
        storage_state_path = Path(org_config['storage_state'])

        if not storage_state_path.exists():
            raise FileNotFoundError(
                f"Auth storage state not found at {storage_state_path}. "
                f"Run tools/buz_auth_bootstrap.py {org_key} first."
            )

        self.result.add_step(f"Switching to: {org_config['display_name']}")

        # Close current context if exists
        if self.context:
            await self.context.close()

        # Create new context with org's authentication
        self.context = await self.browser.new_context(
            storage_state=str(storage_state_path)
        )

        self.result.add_step(f"✓ Switched to {org_config['display_name']}")

    async def handle_org_selector_if_present(self, page: Page, intended_url: str):
        """
        Check if we're on the org selector page and automatically click through.

        Args:
            page: The page object
            intended_url: The URL we were trying to reach
        """
        if "mybuz/organizations" not in page.url:
            return

        self.result.add_step("⚠️  Landed on org selector, clicking through...")

        org_link = page.locator('td a').first
        if await org_link.count() > 0:
            await org_link.click()
            await page.wait_for_load_state('networkidle')
            self.result.add_step("✓ Clicked through org selector")

            # Re-navigate to intended destination
            self.result.add_step(f"Re-navigating to intended page...")
            await page.goto(intended_url, wait_until='networkidle')
        else:
            raise Exception("On org selector page but couldn't find org link to click")

    async def download_inventory_groups_excel(self, org_key: str) -> Path:
        """
        Download the Inventory Groups Excel file for an org.

        Args:
            org_key: Key from ORGS dict

        Returns:
            Path to downloaded Excel file
        """
        org_config = self.ORGS[org_key]
        org_name = org_config['display_name']

        self.result.add_step(f"Downloading inventory groups for {org_name}")

        page = await self.context.new_page()
        try:
            # Navigate to export/import page if not already there
            current_url = page.url
            if self.EXPORT_IMPORT_URL not in current_url:
                self.result.add_step(f"Navigating to export page...")
                try:
                    # Use 'domcontentloaded' instead of 'networkidle' - it's much faster and more reliable
                    await page.goto(self.EXPORT_IMPORT_URL, wait_until='domcontentloaded', timeout=30000)
                except Exception as e:
                    # If navigation fails, check if we ended up on the right domain anyway
                    self.result.add_step(f"Navigation timeout (this is usually fine): {str(e)[:80]}")
                    if "buzmanager.com" not in page.url:
                        raise

                # Handle org selector if present
                await self.handle_org_selector_if_present(page, self.EXPORT_IMPORT_URL)
            else:
                self.result.add_step("Already on export page")

            # Wait for the export link to be available
            await page.wait_for_selector('a[href="#exportModal"]', state='visible', timeout=10000)

            # Click the download link to open modal
            download_link = page.locator('a[href="#exportModal"]')
            await download_link.click()
            await page.wait_for_timeout(1000)
            self.result.add_step("Opened export modal")

            # Select "Inventory Groups" from dropdown (it's a multi-select with id="SheetList")
            # Wait for modal to be visible
            await page.wait_for_selector('select#SheetList', state='visible', timeout=5000)
            select = page.locator('select#SheetList')
            await select.select_option(value='Inventory Groups')
            self.result.add_step("Selected 'Inventory Groups' option")

            # Set up download handler before clicking export
            async with page.expect_download() as download_info:
                # Click Export button
                export_btn = page.locator('button#btnExport')
                await export_btn.click()

            download = await download_info.value

            # Save to output directory with org name
            filename = f"inventory_groups_{org_key}.xlsx"
            save_path = self.output_dir / filename
            await download.save_as(save_path)

            self.result.add_step(f"✓ Downloaded: {filename}")
            return save_path

        finally:
            await page.close()

    def parse_inventory_groups_excel(self, excel_path: Path) -> List[InventoryGroupDiscount]:
        """
        Parse the Inventory Groups Excel file.

        Args:
            excel_path: Path to Excel file

        Returns:
            List of InventoryGroupDiscount objects
        """
        self.result.add_step(f"Parsing: {excel_path.name}")

        wb = openpyxl.load_workbook(excel_path, read_only=True)

        # Find the "Inventory Groups" sheet
        if "Inventory Groups" not in wb.sheetnames:
            raise ValueError(f"Sheet 'Inventory Groups' not found in {excel_path.name}")

        ws = wb["Inventory Groups"]

        inventory_groups = []

        # Skip header row (row 1), start from row 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Column B = Description (index 1)
            # Column C = Code (index 2)
            # Column G = Max Discount Percentage (index 6)
            # Column N = Can be ordered (index 13)

            description = row[1] if len(row) > 1 else None
            code = row[2] if len(row) > 2 else None
            max_discount = row[6] if len(row) > 6 else None
            can_be_ordered = row[13] if len(row) > 13 else None

            # Skip empty rows
            if not code and not description:
                continue

            # Skip rows where "Can be ordered" is not YES
            if can_be_ordered != "YES":
                continue

            # Parse max discount percentage
            # Buz stores percentages as the actual number (0.5 = 0.5%, 50 = 50%)
            max_discount_pct = None
            if max_discount is not None:
                try:
                    max_discount_pct = float(max_discount)
                except (ValueError, TypeError):
                    pass

            inventory_groups.append(InventoryGroupDiscount(
                code=code or "",
                description=description or "",
                max_discount_pct=max_discount_pct
            ))

        wb.close()

        self.result.add_step(f"✓ Parsed {len(inventory_groups)} orderable inventory groups")
        return inventory_groups

    async def upload_inventory_groups_excel(self, org_key: str, file_path: Path) -> Dict[str, Any]:
        """
        Upload an Inventory Groups Excel file to Buz.

        Args:
            org_key: Key from ORGS dict
            file_path: Path to the Excel file to upload

        Returns:
            Dict with upload results (added, edited counts)
        """
        org_config = self.ORGS[org_key]
        org_name = org_config['display_name']

        self.result.add_step(f"Uploading inventory groups to {org_name}")

        page = await self.context.new_page()
        try:
            # Navigate to export/import page
            self.result.add_step(f"Navigating to import page...")
            await page.goto(self.EXPORT_IMPORT_URL, wait_until='domcontentloaded', timeout=30000)

            # Check for org selector
            await self.handle_org_selector_if_present(page, self.EXPORT_IMPORT_URL)

            # Select the file
            self.result.add_step(f"Selecting file: {file_path.name}")
            file_input = page.locator('input#ImportFile[type="file"]')
            await file_input.set_input_files(str(file_path))

            # Click upload button
            self.result.add_step(f"Uploading file...")
            upload_button = page.locator('input#btnUpload[type="submit"]')
            await upload_button.click()

            # Wait for the select dropdown to populate
            self.result.add_step(f"Waiting for sheet selection...")
            await page.wait_for_selector('select#SelectSheets option[value="Inventory Groups"]', timeout=30000)

            # Select "Inventory Groups" option
            select_sheets = page.locator('select#SelectSheets')
            await select_sheets.select_option('Inventory Groups')
            self.result.add_step(f"Selected 'Inventory Groups' sheet")

            # Click import button
            self.result.add_step(f"Importing...")
            import_button = page.locator('input#btnImport[type="submit"]')
            await import_button.click()

            # Wait for success message
            await page.wait_for_selector('p[style*="white-space"]', timeout=60000)

            # Parse the result messages
            result_paragraphs = await page.locator('p[style*="white-space"]').all_text_contents()

            results = {
                'success': False,
                'message': '',
                'added': 0,
                'edited': 0
            }

            for text in result_paragraphs:
                self.result.add_step(f"Buz says: {text.strip()}")

                if 'Save Successful' in text:
                    results['success'] = True

                # Parse "Inventory Groups - Added(X), Edited(Y)"
                if 'Inventory Groups' in text and 'Added' in text:
                    results['message'] = text.strip()
                    # Extract numbers
                    import re
                    added_match = re.search(r'Added\s*\((\d+)\)', text)
                    edited_match = re.search(r'Edited\s*\((\d+)\)', text)

                    if added_match:
                        results['added'] = int(added_match.group(1))
                    if edited_match:
                        results['edited'] = int(edited_match.group(1))

            if results['success']:
                self.result.add_step(f"✓ Upload successful: {results['added']} added, {results['edited']} edited")
            else:
                self.result.add_step(f"⚠️  Upload completed but no success message found")

            return results

        except Exception as e:
            self.result.add_step(f"❌ Upload failed: {str(e)}")
            raise
        finally:
            await page.close()

    async def review_max_discounts(self, selected_orgs: Optional[List[str]] = None) -> MaxDiscountReviewResult:
        """
        Download and parse max discounts from selected orgs.

        Args:
            selected_orgs: List of org keys to process (e.g., ['canberra', 'tweed']).
                          If None, processes all orgs.

        Returns:
            MaxDiscountReviewResult with data from selected orgs
        """
        self.result.add_step("=== Starting Max Discount Review ===")

        # Ensure output directory exists
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Filter orgs if selection provided
        orgs_to_process = self.ORGS.items()
        if selected_orgs:
            orgs_to_process = [(k, v) for k, v in self.ORGS.items() if k in selected_orgs]

        # Process each org
        for idx, (org_key, org_config) in enumerate(orgs_to_process):
            try:
                # Switch to org
                await self.switch_to_org(org_key)

                # Download Excel file
                excel_path = await self.download_inventory_groups_excel(org_key)

                # Parse Excel file
                inventory_groups = self.parse_inventory_groups_excel(excel_path)

                # Store in result
                org_discounts = OrgDiscounts(
                    org_name=org_config['display_name'],
                    inventory_groups=inventory_groups,
                    file_path=str(excel_path)
                )
                self.result.orgs.append(org_discounts)

            except Exception as e:
                self.result.add_step(f"❌ Error processing {org_config['display_name']}: {str(e)}")
                logger.exception(f"Error processing {org_key}")
                # Continue with other orgs even if one fails

        self.result.add_step("=== Review Complete ===")
        return self.result


async def review_max_discounts_all_orgs(
    output_dir: Path,
    headless: bool = True,
    selected_orgs: Optional[List[str]] = None,
    job_update_callback=None
) -> MaxDiscountReviewResult:
    """
    High-level function to review max discounts across selected orgs.

    Args:
        output_dir: Directory to save downloaded Excel files
        headless: Run browser in headless mode
        selected_orgs: List of org keys to process (e.g., ['canberra', 'tweed']).
                      If None, processes all orgs.
        job_update_callback: Optional callback(pct, message) for job progress

    Returns:
        MaxDiscountReviewResult
    """
    def update(pct: int, msg: str):
        if job_update_callback:
            job_update_callback(pct, msg)
        logger.info(f"[{pct}%] {msg}")

    update(0, "Starting max discount review")

    async with BuzMaxDiscountReview(output_dir=output_dir, headless=headless) as review:
        # Wrap add_step to provide progress updates
        original_add_step = review.result.add_step

        def wrapped_add_step(message: str):
            original_add_step(message)
            # Estimate progress
            step_count = len(review.result.steps)
            pct = min(5 + (step_count * 3), 95)
            update(pct, message)

        review.result.add_step = wrapped_add_step

        # Run the review
        result = await review.review_max_discounts(selected_orgs=selected_orgs)

    update(100, "Complete")
    return result


async def upload_max_discount_files(
    upload_files: Dict[str, Path],
    headless: bool = True,
    job_update_callback=None
) -> Dict[str, Any]:
    """
    Upload inventory group Excel files to multiple Buz orgs.

    Args:
        upload_files: Dict mapping org_key -> file_path
        headless: Run browser in headless mode
        job_update_callback: Optional callback(pct, message) for job progress

    Returns:
        Dict with upload results for each org
    """
    def update(pct: int, msg: str):
        if job_update_callback:
            job_update_callback(pct, msg)
        logger.info(f"[{pct}%] {msg}")

    update(0, "Starting upload to Buz")

    # Create a temp output dir (not actually used for this operation, but required by class)
    import tempfile
    with tempfile.TemporaryDirectory() as temp_dir:
        output_dir = Path(temp_dir)

        async with BuzMaxDiscountReview(output_dir=output_dir, headless=headless) as review:
            # Wrap add_step to provide progress updates
            original_add_step = review.result.add_step

            def wrapped_add_step(message: str):
                original_add_step(message)
                # Estimate progress
                step_count = len(review.result.steps)
                pct = min(5 + (step_count * 5), 95)
                update(pct, message)

            review.result.add_step = wrapped_add_step

            # Upload to each org
            upload_results = {}
            total_orgs = len(upload_files)

            for idx, (org_key, file_path) in enumerate(upload_files.items()):
                try:
                    # Switch to org
                    await review.switch_to_org(org_key)

                    # Upload file
                    result = await review.upload_inventory_groups_excel(org_key, file_path)
                    upload_results[org_key] = result

                    # Update progress
                    progress = int(10 + ((idx + 1) / total_orgs) * 85)
                    update(progress, f"Completed upload to {review.ORGS[org_key]['display_name']}")

                except Exception as e:
                    org_name = review.ORGS[org_key]['display_name']
                    error_msg = f"Failed to upload to {org_name}: {str(e)}"
                    review.result.add_step(f"❌ {error_msg}")
                    upload_results[org_key] = {
                        'success': False,
                        'error': str(e)
                    }

        update(100, "Upload complete")

        return {
            'results': upload_results,
            'steps': review.result.steps
        }
