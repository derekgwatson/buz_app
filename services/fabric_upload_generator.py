import json
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
import logging

logger = logging.getLogger(__name__)


def _resolve_output_dir(output_dir: str) -> str:
    # Make relative paths resolve to the project root (…/services/ -> project root)
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    if not os.path.isabs(output_dir):
        output_dir = os.path.join(project_root, output_dir)
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def update_fabric_mappings_from_report(db_manager, report: list, config_path="config.json", output_dir="uploads"):
    """
    Apply changes from a fabric mapping validation report and generate Buz upload Excel file.

    Args:
        db_manager: DatabaseManager instance
        report: list of dicts from check_inventory_groups_against_unleashed()
        config_path: path to config.json containing 'buz_inventory_item_file'
        output_dir: directory to save generated upload files

    Returns:
        str: Path to generated Excel upload file
    """
    logger.info("🛠 Applying fabric mapping updates from report...")

    wrote_any_rows = False

    # Load config for column order
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)
    item_file_config = config.get("buz_inventory_item_file", {})

    # 1) Start with a safe workbook that ALWAYS has a visible sheet
    wb = Workbook()
    placeholder = wb.active
    placeholder.title = "README"
    placeholder.append(["No changes yet. This sheet will be removed if any group tabs are created."])

    tomorrow_str = (datetime.now() + timedelta(days=1)).strftime("%d/%m/%Y")

    for item in report:
        status = item.get("status")

        # --- Add missing fabrics ---
        if status == "missing_fabric":
            logger.info(f"➕ Adding new fabric {item['product_code']}")
            db_manager.insert_item("fabrics", {
                "supplier_product_code": item["product_code"],
                "supplier_id": None,  # TODO: resolve actual supplier_id
                "description_1": item["product_description"],
                "description_2": None,
                "description_3": None,
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            })

        # --- Add missing mappings ---
        elif status == "missing_mapping":
            fabric_row = db_manager.execute_query(
                "SELECT id, supplier_product_code FROM fabrics WHERE supplier_product_code = ?",
                (item["product_code"],)
            ).fetchone()
            if not fabric_row:
                continue
            fabric_id = fabric_row["id"]

            for group in item["missing_groups"]:
                logger.info(f"➕ Adding mapping: {item['product_code']} → {group}")
                db_manager.insert_item("fabric_group_mappings", {
                    "fabric_id": fabric_id,
                    "inventory_group_code": group
                })

                # Add to upload sheet for this inventory group
                if group not in wb.sheetnames:
                    ws = wb.create_sheet(title=group)
                    ws.append([""] + list(item_file_config.keys()))  # Row 1 blank, Row 2 headers
                ws = wb[group]

                # Build upload row
                row_data = [""] * len(item_file_config)
                for col_idx, field_name in enumerate(item_file_config.values()):
                    if field_name.lower() == "inventory code":
                        row_data[col_idx] = item["product_code"]
                    elif field_name.lower() == "date from":
                        row_data[col_idx] = tomorrow_str
                    elif field_name.lower() == "operation":
                        row_data[col_idx] = "A"
                ws.append([""] + row_data)  # prepend blank for row 1 alignment
                wrote_any_rows = True

        # --- Remove invalid mappings ---
        elif status == "invalid_mapping":
            fabric_row = db_manager.execute_query(
                "SELECT id FROM fabrics WHERE supplier_product_code = ?",
                (item["product_code"],)
            ).fetchone()
            if not fabric_row:
                continue
            fabric_id = fabric_row["id"]
            for bad in item["invalid_groups"]:
                group = bad["group"]
                logger.info(f"🗑 Removing invalid mapping: {item['product_code']} → {group}")
                db_manager.execute_query(
                    "DELETE FROM fabric_group_mappings WHERE fabric_id = ? AND inventory_group_code = ?",
                    (fabric_id, group)
                )
                # Optionally: add 'D' operation rows for Buz deletions

    # 4) Remove placeholder only if we actually created at least one real tab
    if wrote_any_rows:
        for name in wb.sheetnames:
            if name != "README":
                wb.active = wb.sheetnames.index(name)
                break
        if "README" in wb.sheetnames and len(wb.sheetnames) > 1:
            try:
                wb.remove(wb["README"])
            except Exception as e:
                logger.warning(f"Could not remove README sheet: {e}")
    else:
        logger.info("No changes found — skipping file generation.")
        return None

    # 5) Resolve output dir robustly and save
    try:
        output_dir_abs = _resolve_output_dir(output_dir)
        file_path = os.path.join(output_dir_abs, f"buz_fabric_upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(file_path)
        logger.info(f"📦 Upload file saved to {file_path}")
        return file_path
    except Exception as e:
        logger.error(f"Failed to save upload workbook: {e}")
        # Optional: return None so your route can 204/no-content or show a message
        return None
