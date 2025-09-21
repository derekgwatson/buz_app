# services/discount_groups_sync.py
# PEP 8 compliant

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional
from datetime import datetime
import re

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from services.google_sheets_service import GoogleSheetsService  # adjust import path if needed
from services.config_service import ConfigManager

import logging
from dataclasses import fields, MISSING
from collections import defaultdict

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class GridConfig:
    sheet_name: str
    header_row: int
    code_col: int
    desc_col: int
    first_group_col: int


@dataclass(frozen=True)
class MappingConfig:
    sheet_name: str
    tab_product_header: str
    grid_code_header: str


@dataclass(frozen=True)
class CustomerTabConfig:
    header_row: int
    product_col_header: str
    discount_col_header: str


def load_cfg_strict(cfg_cls, config_mgr, section_path: str):
    """
    Load a dataclass from ConfigManager, requiring all fields present.
    Raises ValueError on missing/invalid config. Warns on unknown keys.
    """
    raw = config_mgr.get(section_path, default=None)
    if not isinstance(raw, dict) or not raw:
        raise ValueError(f"Missing or empty config section '{section_path}'")

    all_fields = {f.name: f for f in fields(cfg_cls)}
    required = [name for name, f in all_fields.items()
                if f.default is MISSING and f.default_factory is MISSING]

    missing = [k for k in required if k not in raw]
    if missing:
        raise ValueError(
            f"Config '{section_path}' missing required keys: {', '.join(missing)}"
        )

    unknown = [k for k in raw.keys() if k not in all_fields]
    if unknown:
        logger.warning("Config '%s' has unknown keys (ignored): %s",
                       section_path, ", ".join(unknown))

    try:
        return cfg_cls(**{k: raw[k] for k in all_fields.keys() if k in raw})
    except TypeError as e:
        # catches wrong types / extra surprises
        raise ValueError(f"Invalid values in config '{section_path}': {e}")


def _norm(s: str) -> str:
    return re.sub(r"[\s_]+", "", str(s or "").strip().lower())


def _find_header_row_in_col(col_values: list[str], wanted_header: str, default_header_row: int) -> int:
    """Return 1-based header row; fall back to default if not found."""
    wanted = _norm(wanted_header)
    for i, v in enumerate(col_values, start=1):  # 1-based
        if _norm(v) == wanted:
            return i
    return default_header_row


def _as_percent(val: str | float | int) -> float | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.endswith("%"):
        s = s[:-1].strip()
    s = s.replace(",", "")
    try:
        n = float(s)
    except ValueError:
        return None

    # Only scale true fractions; 1 should be 1%, not 100%
    if 0 < n < 1:
        n *= 100.0
    return round(n, 4)


class DiscountGroupsSync:
    """
    Orchestrates:
      1) Manual step hints (create group, attach to customer).
      2) Read Google Sheet tabs and mapping.
      3) Write a new column (per customer) into Discount Grid .xlsm, preserving macros.
    """

    def __init__(self, config: ConfigManager) -> None:
        self.config = config
        self.grid_cfg = load_cfg_strict(GridConfig, self.config, "discount_grid.grid")
        self.map_cfg = load_cfg_strict(MappingConfig, self.config, "discount_grid.mapping_tab")
        self.tab_cfg = load_cfg_strict(CustomerTabConfig, self.config, "discount_grid.customer_tabs")

        self.sheet_id = (config.get("discount_grid.google_sheet_id", default="") or "").strip()
        if not self.sheet_id:
            raise ValueError("Missing discount_grid.google_sheet_id in config.json")

        # Buz URLs for manual steps
        self.url_create_group = "https://go.buzmanager.com/Settings/CustomerDiscountGroups"
        self.url_find_customer = "https://go.buzmanager.com/Contacts/Customers"
        self.url_discount_group_grid = "https://go.buzmanager.com/Settings/InventoryDiscountGroups/Create"

        # Name pattern for the newly added discount group columns in the grid
        # By default use the tab name (customer name). Override via config if needed.
        self.new_group_name_template = self.config.get(
            "discount_grid.new_group_name_template",
            default="{customer_tab}"
        )
        self.ignore_tabs_raw = config.get("discount_grid", "ignore_tabs", default=[]) \
                               or config.get("discount_grid.ignore_tabs", default=[]) \
                               or []
        self.ignore_tabs = {_norm(x) for x in self.ignore_tabs_raw}

    def _should_ignore_tab(self, tab: str) -> bool:
        """True if tab is mapping/utility/ignored by config."""
        if tab == self.map_cfg.sheet_name:
            return True
        if tab.startswith("_"):
            return True
        return _norm(tab) in self.ignore_tabs

    def list_customer_tabs(self, gs: GoogleSheetsService) -> list[str]:
        """All usable customer tabs after filtering."""
        tabs = gs.get_sheet_names(self.sheet_id, include_hidden=False)
        return [t for t in tabs if not self._should_ignore_tab(t)]

    @staticmethod
    def _extract_sheet_id(sheet_url: str) -> str:
        # Supports full URLs like:
        # https://docs.google.com/spreadsheets/d/<ID>/edit
        m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", sheet_url)
        return m.group(1) if m else sheet_url

    # -----------------------------
    # Public API
    # -----------------------------

    def get_manual_steps(self, customer_tabs: List[str]) -> List[Dict[str, str]]:
        gs = GoogleSheetsService()
        blocks = gs.fetch_many(self.sheet_id, customer_tabs, "A1:B20")
        steps: List[Dict[str, str]] = []
        for tab in customer_tabs:
            rows = blocks.get(tab, [])
            customers = []
            for row in rows:
                a = (row[0] if len(row) > 0 else "").strip().lower()
                b = (row[1] if len(row) > 1 else "").strip()
                if a.startswith("customer"):
                    seen = set()
                    for n in (x.strip() for x in re.split(r",", b) if x.strip()):
                        k = n.lower()
                        if k not in seen:
                            seen.add(k)
                            customers.append(n)
                    break
            group_name = self.new_group_name_template.format(customer_tab=tab).strip()
            steps.append({
                "customer_tab": tab,
                "discount_group_to_create": group_name,
                "create_group_url": self.url_create_group,
                "find_customer_url": self.url_find_customer,
                "customers": customers,
                "note": "Create the discount group in Buz, then edit the customer card and attach the new group."
            })
        return steps

    @staticmethod
    def _resolve_header(actual_headers: List[str], wanted: str) -> Optional[str]:
        """Return the actual header text in the sheet that matches 'wanted' (case/space insensitive)."""
        want = _norm(wanted)
        for h in actual_headers:
            if _norm(h) == want:
                return h
        return None

    # top-level helpers in this file
    def build(self, input_xlsm_path: str, output_xlsm_path: str) -> Dict[str, Any]:
        """
        Read Google Sheet, add/update columns in XLSM grid, and save as a new file.
        Returns a summary (for UI).
        """
        gs = GoogleSheetsService()

        # --- 1) Mapping sheet: tab product -> grid product code (headers are on row 1)
        mapping_rows = gs.read_tab_as_records(
            spreadsheet_id=self.sheet_id,
            sheet_name=self.map_cfg.sheet_name,
            header_row=1,  # mapping tab headers are in the first row
        )
        if not mapping_rows:
            raise ValueError(f"Mapping sheet '{self.map_cfg.sheet_name}' is empty or missing.")

        headers = list(mapping_rows[0].keys())
        tab_hdr = self._resolve_header(headers, self.map_cfg.tab_product_header)
        grid_hdr = self._resolve_header(headers, self.map_cfg.grid_code_header)
        if not tab_hdr or not grid_hdr:
            raise ValueError(
                f"Mapping sheet must have headers '{self.map_cfg.tab_product_header}' and "
                f"'{self.map_cfg.grid_code_header}'. Found: {headers}"
            )

        tab_to_grid_codes: dict[str, list[str]] = defaultdict(list)
        for row in mapping_rows:
            tab_name = str(row.get(tab_hdr, "")).strip()
            grid_code = str(row.get(grid_hdr, "")).strip()
            if not tab_name or not grid_code:
                continue
            key = _norm(tab_name)  # case/space-insensitive match
            if grid_code not in tab_to_grid_codes[key]:
                tab_to_grid_codes[key].append(grid_code)

        # --- 2) Discover customer tabs (all tabs except mapping + those not prefixed with "_")
        customer_tabs = self.list_customer_tabs(gs)  # gs.get_sheet_names(self.sheet_id)

        # --- 3) Open .xlsm grid (preserve VBA)
        wb = load_workbook(input_xlsm_path, keep_vba=True)
        ws: Worksheet = wb[self.grid_cfg.sheet_name]

        # --- 4) Build row index in grid: grid product code -> row index
        code_to_row: Dict[str, int] = {}
        for r in range(self.grid_cfg.header_row + 1, ws.max_row + 1):
            code = str(ws.cell(row=r, column=self.grid_cfg.code_col).value or "").strip()
            if code:
                code_to_row[code] = r

        # --- 5) Existing group columns (header row)
        existing_headers: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            name = str(ws.cell(row=self.grid_cfg.header_row, column=c).value or "").strip()
            if name:
                existing_headers[name] = c

        # --- 6) For each customer tab, add/update a column
        customers_summary: List[Dict[str, Any]] = []
        total_changes = 0
        changes_sample: List[Dict[str, Any]] = []

        # NEW: batch fetch columns A and C for all tabs in 2 requests total
        max_rows = 2000
        colA_by_tab = gs.fetch_many(self.sheet_id, customer_tabs, f"A1:A{max_rows}")  # one values:batchGet
        colC_by_tab = gs.fetch_many(self.sheet_id, customer_tabs, f"C1:C{max_rows}")  # one values:batchGet

        def _flatten_col(block: list[list[str]]) -> list[str]:
            # values API returns rows like [["hdr"], ["val"], ...] for a single column
            return [row[0] if row else "" for row in (block or [])]

        for tab in customer_tabs:
            col_a = _flatten_col(colA_by_tab.get(tab, []))
            col_c = _flatten_col(colC_by_tab.get(tab, []))

            hdr_row = _find_header_row_in_col(col_a, self.tab_cfg.product_col_header, self.tab_cfg.header_row)
            start_row = hdr_row + 1

            # Build “rows” dicts from the two columns we already have in memory
            rows: list[dict[str, str]] = []
            end = max(len(col_a), len(col_c))
            for r in range(start_row, end + 1):
                i = r - 1
                prod = col_a[i] if i < len(col_a) else ""
                disc = col_c[i] if i < len(col_c) else ""
                if not (prod or disc):
                    continue
                rows.append({
                    self.tab_cfg.product_col_header: prod,
                    self.tab_cfg.discount_col_header: disc,
                })

            if not rows:
                customers_summary.append({"tab": tab, "added": False, "reason": "empty tab"})
                continue

            # Resolve product/discount keys on this tab (case/space insensitive)
            hdr_keys = list(rows[0].keys())
            prod_key = self._resolve_header(hdr_keys, self.tab_cfg.product_col_header)
            disc_key = self._resolve_header(hdr_keys, self.tab_cfg.discount_col_header)
            if not prod_key or not disc_key:
                customers_summary.append({
                    "tab": tab,
                    "group_column": "",
                    "cells_changed": 0,
                    "reason": f"missing headers '{self.tab_cfg.product_col_header}' and/or '{self.tab_cfg.discount_col_header}'"
                })
                continue

            # Determine/ensure the target group column in the grid
            group_col_name = self.new_group_name_template.format(customer_tab=tab).strip()
            target_col = existing_headers.get(group_col_name)
            if not target_col:
                target_col = ws.max_column + 1
                ws.cell(row=self.grid_cfg.header_row, column=target_col).value = group_col_name
                existing_headers[group_col_name] = target_col

            # Push discounts
            changed = 0
            for row in rows:
                tab_product = str(row.get(prod_key, "")).strip()
                pct = _as_percent(row.get(disc_key))
                if not tab_product or pct is None:
                    continue

                codes = tab_to_grid_codes.get(_norm(tab_product), [])
                if not codes:
                    continue  # nothing mapped for this tab product label

                for grid_code in codes:
                    grid_row = code_to_row.get(grid_code)
                    if not grid_row:
                        continue  # product code not present in the Excel grid

                    cell = ws.cell(row=grid_row, column=target_col)
                    before = cell.value
                    after = float(pct)
                    if before is None or float(before) != after:
                        cell.value = after
                        changed += 1
                        total_changes += 1
                        if len(changes_sample) < 50:
                            changes_sample.append({
                                "product_code": grid_code,
                                "group_column": group_col_name,
                                "before": before,
                                "after": after,
                            })

            customers_summary.append({
                "tab": tab,
                "group_column": group_col_name,
                "cells_changed": changed,
            })

        # --- 7) Save new copy (preserving macros)
        wb.save(output_xlsm_path)

        return {
            "output_file": output_xlsm_path,
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "totals": {"cells_changed": total_changes},
            "customers": customers_summary,
            "changes_sample": changes_sample,
        }

    def get_sheet_url(self) -> str:
        # plain sheet open; no extra API calls or gid lookups
        return f"https://docs.google.com/spreadsheets/d/{self.sheet_id}/edit"

    def read_customer_names_for_tab(self, gs: GoogleSheetsService, sheet_name: str) -> list[str]:
        """
        Look for a header that starts with 'customer' in col A and take the value in col B.
        Split by commas; strip blanks. Returns [] if not found.
        """
        # Read a small window (A1:B20) to keep it cheap
        vals = gs.read_range(self.sheet_id, sheet_name, "A1:B20")  # returns list[list[str]]
        for row in vals:
            a = (row[0] if len(row) > 0 else "").strip().lower()
            b = (row[1] if len(row) > 1 else "").strip()
            if a.startswith("customer"):
                names = [n.strip() for n in re.split(r",", b) if n.strip()]
                # dedupe while preserving order
                seen, out = set(), []
                for n in names:
                    if n.lower() not in seen:
                        seen.add(n.lower())
                        out.append(n)
                return out
        return []
