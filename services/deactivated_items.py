from openpyxl import Workbook
from datetime import datetime
from flask import current_app
from services.database import DatabaseManager
from services.helper import parse_headers
import os


def generate_deactivation_upload(db_manager: DatabaseManager):
    """
    Generate an upload file for deactivating obsolete/unsellable items in Buz.

    Args:
        db_manager (DatabaseManager): Instance of your DatabaseManager class.
    """
    # Load headers from app config
    headers_config = current_app.config["headers"]
    expected_headers, db_fields = parse_headers(headers_config, "buz_inventory_item_file")

    # Query database for obsolete/unsellable items
    query = f"""
    SELECT id, inventory_group_code, {', '.join(db_fields)} FROM inventory_items
    WHERE LOWER(Supplier) = 'unleashed'
    AND SupplierProductCode != ''
    AND LOWER(SupplierProductCode) NOT IN (
        SELECT LOWER(ProductCode) FROM unleashed_products
        WHERE IsObsoleted = 'No' AND IsSellable = 'Yes'
    )
    """
    cursor = db_manager.execute_query(query)

    # Convert cursor results to list of dictionaries
    columns = [col[0] for col in cursor.description]  # Extract column names
    items = [dict(zip(columns, row)) for row in cursor.fetchall()]  # Convert rows to dictionaries

    # Organize items by InventoryGroupCode
    grouped_items = {}
    for item in items:
        group = item['inventory_group_code']
        if group not in grouped_items:
            grouped_items[group] = []
        grouped_items[group].append(item)

    # Create Excel workbook
    workbook = Workbook()
    for group, items in grouped_items.items():
        sheet = workbook.create_sheet(title=group)
        # Add headers
        sheet.append([])    # starts with a blank row
        sheet.append(expected_headers)
        for item in items:
            # Populate row with database fields
            row = [item.get(field) for field in db_fields]

            # Insert 'D' in the second-to-last column position
            if len(row) > 0:
                row[-1] = 'D'  # Replace the value in the last column with 'D'

            sheet.append(row)

    # Remove default empty sheet
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Save workbook
    filename = f"deactivate_items_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename = os.path.join(current_app.config['upload_folder'], filename)
    workbook.save(filename)
    return filename
