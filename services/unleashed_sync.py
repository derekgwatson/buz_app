import json
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
import logging
from typing import Any
import re
from services.excel_safety import save_workbook_gracefully

logger = logging.getLogger(__name__)

UNLEASHED_SUPPLIER_CODE = "UNLEASHED"
UNLEASHED_SUPPLIER_NAME = "Unleashed ERP"


# ---------- small utils ----------


# allow \t, \n, \r; strip the rest of C0 controls
_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_FORMULA_LEADERS = ("=", "+", "-", "@")  # anything Excel may treat as a formula


def _clean_cell(v):
    """Return a version of v that Excel accepts, never as an accidental formula."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    # everything else -> string sanitize
    s = str(v)
    s = _CONTROL_CHARS.sub("", s)[:32767]
    # escape potential formula injection / misinterpretation
    if s.startswith(_FORMULA_LEADERS):
        s = "'" + s
    return s


def _safe_ws_append(ws, values):
    """Append a row with all cell values sanitized."""
    ws.append([_clean_cell(x) for x in values])


def _save_workbook_atomically(wb, final_path: str):
    """
    Save to a temp file, flush to disk, then atomically replace the target.
    On Windows, fsync can be finicky; we treat it as best-effort.
    """
    tmp_path = f"{final_path}.tmp"
    has_real_data = save_workbook_gracefully(wb, tmp_path)
    if not has_real_data:
        logger.warning("No data matched your filters — exported a placeholder workbook.")
    try:
        wb.close()
    except Exception:
        pass

    # Best-effort durability
    try:
        # open read/write so flush+fsync work on Windows too
        with open(tmp_path, "rb+") as f:
            f.flush()
            os.fsync(f.fileno())
    except (OSError, AttributeError, ValueError):
        # Ignore if the platform/filesystem doesn't support it
        pass

    os.replace(tmp_path, final_path)


def _norm_group_name(s: str | None) -> str:
    return (s or "").strip()


def _norm_code(s: str | None) -> str:
    """Strip any leading asterisks and normalise case/whitespace."""
    if not s:
        return ""
    return s.lstrip("*").strip().upper()


def _project_root() -> str:
    return os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))


def _resolve_output_dir(output_dir: str) -> str:
    if not os.path.isabs(output_dir):
        output_dir = os.path.join(_project_root(), output_dir)
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def _safe_wb():
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["This file contains Buz upload data. This sheet is removed if any data sheets are created."])
    return wb


def _norm_group_key(s: str | None) -> str:
    return (s or "").strip()


def _get_friendly_descs(db, supplier_code_norm: str) -> tuple[str | None, str | None, str | None]:
    # 1) fabrics table
    row = db.execute_query(
        """
        SELECT description_1, description_2, description_3
        FROM fabrics
        WHERE UPPER(LTRIM(TRIM(supplier_product_code), '*')) = ?
        LIMIT 1
        """,
        (supplier_code_norm,),
    ).fetchone()

    if row:
        d = dict(row)
        d1 = (d.get("description_1") or "").strip() or None
        d2 = (d.get("description_2") or "").strip() or None
        d3 = (d.get("description_3") or "").strip() or None
        if d1 or d2 or d3:
            return d1, d2, d3

    # 2) unleashed_products “friendly” fields if present; fall back to ProductDescription
    row = db.execute_query(
        """
        SELECT *
        FROM unleashed_products_fabric 
        WHERE UPPER(LTRIM(TRIM(ProductCode), '*')) = ?
        LIMIT 1
        """,
        (supplier_code_norm,),
    ).fetchone()

    if row:
        d = dict(row)
        d1 = (d.get("FriendlyDescription1") or "").strip() or None
        d2 = (d.get("FriendlyDescription2") or "").strip() or None
        d3 = (d.get("FriendlyDescription3") or "").strip() or None
        if d1 or d2 or d3:
            return d1, d2, d3

        pd = (d.get("ProductDescription") or "").strip()
        if pd:
            parts = [p.strip() for p in pd.split(" - ")]
            if len(parts) < 2:
                words = pd.split()
                if len(words) >= 3:
                    parts = [" ".join(words[:-2]), words[-2], words[-1]]
            d1 = parts[0] if len(parts) > 0 else None
            d2 = parts[1] if len(parts) > 1 else None
            d3 = parts[2] if len(parts) > 2 else None
            return d1, d2, d3

    return None, None, None


def _fetch_buz_item_by_supplier_norm(db, supplier_code_norm: str) -> dict | None:
    row = db.execute_query(
        """
        SELECT *
        FROM inventory_items
        WHERE UPPER(LTRIM(TRIM(SupplierProductCode), '*')) = ?
        ORDER BY (LastEditDate IS NULL), LastEditDate DESC
        LIMIT 1
        """,
        (supplier_code_norm,),
    ).fetchone()
    return dict(row) if row else None


def _append_delete_row(ws, items_headers: list[str], items_dbfield_by_col: dict[str, str], rec: dict):
    row = [""] * len(items_headers)

    # Fill each column from the corresponding db field if present
    for i, col in enumerate(items_headers):
        dbf = items_dbfield_by_col.get(col)
        if dbf:
            row[i] = rec.get(dbf, "")

    # Force Operation = 'D'
    for i, col in enumerate(items_headers):
        if col.lower() == "operation":
            row[i] = "D"
            break

    _safe_ws_append(ws, row)


def _get_group_display_name(db, group_code: str) -> str:
    """
    Best-effort lookup of a friendly product name for an inventory group code.
    Falls back to the code itself if no description is stored.
    """
    if not group_code:
        return "Product"

    # try common schema
    row = None
    try:
        row = db.execute_query(
            "SELECT group_description AS descr FROM inventory_groups WHERE group_code = ? LIMIT 1",
            (group_code,),
        ).fetchone()
    except Exception:
        row = None

    if not row:
        try:
            row = db.execute_query(
                "SELECT description AS descr FROM inventory_groups WHERE group_code = ? LIMIT 1",
                (group_code,),
            ).fetchone()
        except Exception:
            row = None

    if row and row["descr"]:
        return str(row["descr"]).strip()

    # last resort: use the code (e.g., ROLL → Roller Blind if you later add a mapping)
    return group_code


def _build_description(product_name: str, d1: str | None, d2: str | None, d3: str | None) -> str:
    colour = (d3 or "").strip()
    if colour and colour.lower() == "to be confirmed":
        colour = "Colour To Be Confirmed"

    parts = [p for p in [d1, d2, colour] if p]
    base = " ".join(parts).strip()
    return (f"{product_name} {base}").strip() if base else product_name


def build_sequential_code_provider(db, config: dict, default_start: int = 10000):
    """
    Returns a callable(db, supplier_code_norm) -> new Buz 'Code'.

    Prefix choice:
      1) Try to infer from Unleashed ProductGroup via config['unleashed_group_to_inventory_groups'] (first entry)
      2) Fallback to 'FAB' if no mapping is found

    Uniqueness:
      - Looks up the current MAX numeric suffix for that prefix in inventory_items.Code
      - Caches next number per prefix to avoid multiple queries when adding many items
    """
    # Normalize mapping keys once
    group_map_raw = config.get("unleashed_group_to_inventory_groups", {})
    group_map = { _norm_group_key(k): v for k, v in group_map_raw.items() }

    next_by_prefix: dict[str, int] = {}

    def _infer_prefix(db_inner, supplier_code_norm: str) -> str:
        # Find the Unleashed ProductGroup for this product code
        row = db_inner.execute_query(
            """
            SELECT ProductGroup
            FROM unleashed_products_fabric 
            WHERE UPPER(LTRIM(TRIM(ProductCode), '*')) = ?
            LIMIT 1
            """,
            (supplier_code_norm,),
        ).fetchone()

        if row:
            pg = _norm_group_key(row["ProductGroup"])
            inv_groups = group_map.get(pg) or []
            if inv_groups:
                return inv_groups[0]  # take the first configured inventory group code

        return "FAB"  # safe fallback

    def _init_prefix_counter(db_inner, prefix: str) -> int:
        # SUBSTR is 1-based in SQLite; strip the prefix so we can cast the remainder to INTEGER
        res = db_inner.execute_query(
            """
            SELECT MAX(CAST(SUBSTR(Code, ?) AS INTEGER)) AS max_num
            FROM inventory_items
            WHERE Code LIKE ?
            """,
            (len(prefix) + 1, f"{prefix}%")
        ).fetchone()

        max_num = res["max_num"] if res and res["max_num"] is not None else (default_start - 1)
        return int(max_num) + 1

    def provider(db_inner, supplier_code_norm: str) -> str:
        prefix = _infer_prefix(db_inner, supplier_code_norm)

        if prefix not in next_by_prefix:
            next_by_prefix[prefix] = _init_prefix_counter(db_inner, prefix)

        num = next_by_prefix[prefix]
        code = f"{prefix}{num:05d}"
        next_by_prefix[prefix] = num + 1
        return code

    return provider

# ---------- supplier setup ----------


def get_or_create_unleashed_supplier(db) -> int:
    row = db.execute_query(
        "SELECT id FROM suppliers WHERE supplier_code = ?",
        (UNLEASHED_SUPPLIER_CODE,)
    ).fetchone()
    if row:
        return row["id"]

    now = datetime.utcnow()
    db.insert_item("suppliers", {
        "supplier_code": UNLEASHED_SUPPLIER_CODE,
        "supplier": UNLEASHED_SUPPLIER_NAME,   # NOTE: your column name is 'supplier'
        "is_active": 1,
        "created_at": now,
        "updated_at": now,
    })
    row = db.execute_query(
        "SELECT id FROM suppliers WHERE supplier_code = ?",
        (UNLEASHED_SUPPLIER_CODE,)
    ).fetchone()
    logger.info(f"Created supplier {UNLEASHED_SUPPLIER_NAME} with id={row['id']}")
    return row["id"]


def tag_local_fabrics_as_unleashed(db, unleashed_supplier_id: int):
    """Mark any fabric whose code appears in Unleashed (after stripping leading *) as Unleashed-sourced."""
    u_rows = db.execute_query(
        "SELECT DISTINCT ProductCode FROM unleashed_products_fabric"
    ).fetchall()
    u_codes = {_norm_code(r["ProductCode"]) for r in u_rows}

    f_rows = db.execute_query(
        "SELECT id, supplier_product_code, supplier_id FROM fabrics"
    ).fetchall()

    updated = 0
    now = datetime.utcnow()
    for f in f_rows:
        if _norm_code(f["supplier_product_code"]) in u_codes and f["supplier_id"] != unleashed_supplier_id:
            db.execute_query(
                "UPDATE fabrics SET supplier_id = ?, updated_at = ? WHERE id = ?",
                (unleashed_supplier_id, now, f["id"])
            )
            updated += 1
    if updated:
        logger.info(f"Tagged {updated} local fabrics as Unleashed.")


# ---------- Buz code helpers ----------

def _get_existing_buz_code(db, supplier_product_code_norm: str) -> str | None:
    """
    Find the current Buz 'Code' for a fabric by supplier product code.
    Assumes inventory_items table stores SupplierProductCode and Code.
    """
    row = db.execute_query(
        """
        SELECT Code
        FROM inventory_items
        WHERE UPPER(LTRIM(TRIM(SupplierProductCode), '*')) = ?
        ORDER BY (LastEditDate IS NULL), LastEditDate DESC
        LIMIT 1
        """,
        (supplier_product_code_norm,)
    ).fetchone()
    return row["Code"] if row else None


def _append_item_row(
    ws,
    items_headers: list[str],
    buz_code: str,
    op: str,
    desc1: str | None = None,
    desc2: str | None = None,
    desc3: str | None = None,
    supplier: str | None = None,
    supplier_prod_code: str | None = None,
    description: str | None = None,   # NEW
):
    row = [""] * len(items_headers)

    def idx_eq(name):
        lname = name.lower()
        for i, col in enumerate(items_headers):
            if col.lower() == lname:
                return i
        return None

    def idx_starts(prefix):
        lp = prefix.lower()
        for i, col in enumerate(items_headers):
            if col.lower().startswith(lp):
                return i
        return None

    i_code = idx_eq("Code")
    i_op   = idx_eq("Operation")
    i_desc = idx_eq("Description")            # column C
    i_d1   = idx_starts("DescnPart1")
    i_d2   = idx_starts("DescnPart2")
    i_d3   = idx_starts("DescnPart3")
    i_sup  = idx_eq("Supplier")
    i_spc  = idx_eq("Supplier Product Code")

    if i_code is not None: row[i_code] = buz_code
    if i_op   is not None: row[i_op]   = op
    if i_desc is not None and description is not None: row[i_desc] = description
    if i_d1   is not None and desc1 is not None: row[i_d1] = desc1
    if i_d2   is not None and desc2 is not None: row[i_d2] = desc2
    if i_d3   is not None and desc3 is not None: row[i_d3] = desc3
    if i_sup  is not None and supplier is not None: row[i_sup] = supplier
    if i_spc  is not None and supplier_prod_code is not None: row[i_spc] = supplier_prod_code

    _safe_ws_append(ws, row)


def _append_pricing_row(ws, pricing_headers: list[str], buz_code: str, date_str: str, sell: float | None, cost: float | None):
    row = [""] * len(pricing_headers)

    def idx(name):
        lname = name.lower()
        for i, col in enumerate(pricing_headers):
            if col.lower() == lname:
                return i
        return None

    i_code = idx("Inventory Code")
    i_date = idx("Date From")
    i_op   = idx("Operation")

    if i_code is not None:
        row[i_code] = buz_code
    if i_date is not None:
        row[i_date] = date_str
    if i_op is not None:
        row[i_op] = "A"

    i_sell = idx("SellSQM")
    i_cost = idx("CostSQM")
    if i_sell is not None and sell is not None:
        row[i_sell] = sell
    if i_cost is not None and cost is not None:
        row[i_cost] = cost

    _safe_ws_append(ws, row)


# ---------- main orchestration ----------

def sync_unleashed_fabrics(
    db,
    config_path: str = "config.json",
    output_dir: str = "uploads",
    code_provider=None,            # optional: callable(db, normalized_supplier_code) -> new Buz Code string
    price_provider=None,           # optional: callable(db, normalized_supplier_code) -> {"sell": x, "cost": y}
    pricing_tolerance: float = 0.005,
    progress=None,
):
    """
    Sync Unleashed ↔ Buz (scoped to Unleashed-sourced fabrics only).

    - Ensures UNLEASHED supplier exists and is active.
    - Tags local fabrics as Unleashed where code appears in Unleashed export (handles any leading '*').
    - Computes adds (inventory A) and deletes (inventory D).
    - Optionally emits pricing adds (A) for codes where price delta exceeds tolerance.
    - Uses Buz 'Code' for upload rows (NOT the Unleashed ProductCode).

    Returns:
        dict {
          "items_file": <path|None>,
          "pricing_file": <path|None>,
          "adds": [supplier_product_code_norm...],
          "deletes": [supplier_product_code_norm...],
          "pricing_count": int
        }
    """

    def _p(msg, pct=None):
        if callable(progress):
            progress(msg, pct)

    _p("Loading config…", 5)
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    items_headers = [h["spreadsheet_column"] for h in config["headers"]["buz_inventory_item_file"]]
    pricing_headers = [h["spreadsheet_column"] for h in config["headers"]["buz_pricing_file"]]

    # map spreadsheet column -> database field (e.g. "Code" -> "Code")
    items_dbfield_by_col = {
        h["spreadsheet_column"]: h.get("database_field")
        for h in config["headers"]["buz_inventory_item_file"]
    }

    # NEW: use config directly (no current_app)
    group_map = {_norm_group_name(k): (v or [])
                 for k, v in config.get("unleashed_group_to_inventory_groups", {}).items()}

    _p("Ensuring UNLEASHED supplier exists…", 8)
    unleashed_supplier_id = get_or_create_unleashed_supplier(db)

    _p("Tagging local fabrics as Unleashed…", 12)
    tag_local_fabrics_as_unleashed(db, unleashed_supplier_id)

    # --- build Unleashed set U (as you already do) ---
    _p("Building Unleashed and Buz sets…", 20)
    u_rows = db.execute_query(
        """
        SELECT DISTINCT ProductCode
        FROM unleashed_products_fabric
        """
    ).fetchall()
    U = {_norm_code(r["ProductCode"]) for r in u_rows}

    # --- build Buz set I from inventory_items restricted to Unleashed-sourced items ---
    # If you store supplier as text on inventory_items, match it case-insensitively.
    # Adjust the WHERE clause if you instead store a supplier_code column.
    i_rows = db.execute_query(
        """
        SELECT
            Code,
            UPPER(LTRIM(TRIM(SupplierProductCode), '*')) AS code_norm,
            inventory_group_code
        FROM inventory_items
        WHERE Supplier IS NOT NULL
          AND TRIM(Supplier) != ''
          AND UPPER(TRIM(Supplier)) IN ('UNLEASHED', 'UNLEASHED ERP')
          AND SupplierProductCode IS NOT NULL
          AND TRIM(SupplierProductCode) != ''
        """
    ).fetchall()

    # Map Buz by supplier product code (normalized)
    I_map = {r["code_norm"]: {"code": r["Code"], "group": r["inventory_group_code"]} for r in i_rows}
    I = set(I_map.keys())

    # --- final diffs (what we actually want to upload) ---
    adds = sorted(U - I)  # in Unleashed, not yet in Buz
    deletes = sorted(I - U)  # in Buz (Unleashed-sourced), no longer in Unleashed
    _p(f"Diff computed: {len(adds)} adds, {len(deletes)} deletes.", 35)

    logger.info(f"Unleashed sync — adds: {len(adds)}, deletes: {len(deletes)} (U={len(U)}, I={len(I)})")
    logger.info(f"ADD sample (first 20): {adds[:20]}")
    logger.info(f"DEL sample (first 20): {deletes[:20]}")

    tomorrow_str = (datetime.now() + timedelta(days=1)).strftime("%d/%m/%Y")
    outdir = _resolve_output_dir(output_dir)

    # ----- INVENTORY UPLOAD (adds + deletes) -----
    # --- inside sync_unleashed_fabrics(), replace the "Items" sheet block ---

    items_file_path = None
    wrote_items = False

    items_wb = _safe_wb()

    ws_by_group: dict[str, Any] = {}

    def _infer_primary_group(db_inner, supplier_code_norm: str) -> str:
        # Look up the product’s ProductGroup in Unleashed
        row = db_inner.execute_query(
            """
            SELECT ProductGroup
            FROM unleashed_products_fabric 
            WHERE UPPER(LTRIM(TRIM(ProductCode), '*')) = ?              
            LIMIT 1
            """,
            (supplier_code_norm,),
        ).fetchone()

        if not row:
            return "Uncategorized"

        pg = _norm_group_name(row["ProductGroup"])
        groups = group_map.get(pg, [])
        return groups[0] if groups else "Uncategorized"

    def _get_existing_item_group(db_inner, buz_code: str | None) -> str | None:
        if not buz_code:
            return None
        row = db_inner.execute_query(
            """
            SELECT inventory_group_code 
            FROM inventory_items 
            WHERE Code=? 
            ORDER BY (LastEditDate IS NULL), LastEditDate DESC 
            LIMIT 1
            """,
            (buz_code,),
        ).fetchone()
        return row["inventory_group_code"] if row else None

    def _ws_for_group(group: str):
        if group not in ws_by_group:
            ws = items_wb.create_sheet(title=group[:31] or "Sheet")  # Excel tab max 31 chars
            # Row 1: blank row; Row 2: headers
            _safe_ws_append(ws, [""] * len(items_headers))
            _safe_ws_append(ws, items_headers)
            ws_by_group[group] = ws
        return ws_by_group[group]

    # Items workbook
    if adds or deletes:
        _p("Preparing items workbook…", 45)

    # ---- Adds (A)
    for supplier_code_norm in adds:
        group = _infer_primary_group(db, supplier_code_norm)
        ws = _ws_for_group(group)

        buz_code = code_provider(db, supplier_code_norm) if callable(code_provider) else supplier_code_norm
        if not buz_code:
            logger.warning(f"No Buz code generated for add {supplier_code_norm}; skipping.")
            continue

        d1, d2, d3 = _get_friendly_descs(db, supplier_code_norm)
        product_name = _get_group_display_name(db, group)
        full_desc = _build_description(product_name, d1, d2, d3)

        _append_item_row(
            ws,
            items_headers,
            buz_code,
            op="A",
            desc1=d1,
            desc2=d2,
            desc3=d3,
            supplier="UNLEASHED",
            supplier_prod_code=supplier_code_norm,
            description=full_desc,  # <- Column C
        )
        wrote_items = True

    # ---- Deletes (D)
    for supplier_code_norm in deletes:
        rec = _fetch_buz_item_by_supplier_norm(db, supplier_code_norm)
        if not rec:
            logger.warning(f"No existing Buz item found for delete {supplier_code_norm}; skipping.")
            continue

        # Prefer the item’s real group for the tab
        group = (rec.get("inventory_group_code") or "").strip() or _infer_primary_group(db, supplier_code_norm)
        ws = _ws_for_group(group)

        _append_delete_row(ws, items_headers, items_dbfield_by_col, rec)
        wrote_items = True

    # finalize workbook
    if wrote_items:
        # make first real tab active, drop README safely
        first_tab = next((n for n in items_wb.sheetnames if n != "README"), None)
        if first_tab:
            items_wb.active = items_wb.sheetnames.index(first_tab)
        if "README" in items_wb.sheetnames and len(items_wb.sheetnames) > 1:
            try:
                items_wb.remove(items_wb["README"])
            except Exception as e:
                logger.warning(f"Could not remove README from items workbook: {e}")

        items_file_path = os.path.join(
            outdir, f"buz_items_unleashed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        _p("Saving items workbook…", 75)
        _save_workbook_atomically(items_wb, items_file_path)
    else:
        items_file_path = None

    # ----- PRICING UPLOAD (optional) -----
    pricing_file_path = None
    pricing_count = 0
    if callable(price_provider):
        _p("Building pricing workbook…", 82)
        pricing_wb = _safe_wb()
        ws_price = pricing_wb.create_sheet(title="Pricing")
        _safe_ws_append(ws_price, pricing_headers)
        wrote_pricing = False

        # Only price codes that exist both sides (U ∩ L)
        for supplier_code_norm in sorted(U & set(I_map.keys())):
            prices = price_provider(db, supplier_code_norm) or {}
            sell = prices.get("sell") or prices.get("SellSQM")
            cost = prices.get("cost") or prices.get("CostSQM")

            # We need the Buz Code to put in the pricing file
            buz_code = _get_existing_buz_code(db, supplier_code_norm)
            if not buz_code:
                logger.warning(f"No existing Buz Code found for pricing {supplier_code_norm}; skipping.")
                continue

            _append_pricing_row(ws_price, pricing_headers, buz_code, tomorrow_str, sell, cost)
            wrote_pricing = True
            pricing_count += 1

        if wrote_pricing:
            _p("Saving pricing workbook…", 88)
            pricing_wb.active = pricing_wb.sheetnames.index("Pricing")
            if "README" in pricing_wb.sheetnames and len(pricing_wb.sheetnames) > 1:
                try:
                    pricing_wb.remove(pricing_wb["README"])
                except Exception as e:
                    logger.warning(f"Could not remove README from pricing workbook: {e}")
            pricing_file_path = os.path.join(outdir, f"buz_pricing_unleashed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            _save_workbook_atomically(pricing_wb, pricing_file_path)

    _p("Finalizing summary…", 95)
    summary = {
        "items_file": items_file_path,
        "pricing_file": pricing_file_path,
        "adds": adds,
        "deletes": deletes,
        "pricing_count": pricing_count,
    }
    logger.info(f"Unleashed sync summary: {summary}")
    _p("Done.", 100)
    return summary

