from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Dict, Optional, Tuple, Union

from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string
from openpyxl.utils.datetime import from_excel as excel_serial_to_dt

# Columns per your spec
WARNING_COL = column_index_from_string("AK")  # warning
LAST_EDIT_COL = column_index_from_string("AN")  # last edit date
OPERATION_COL = column_index_from_string("AO")  # operation

HEADER_ROW = 2
DATA_START_ROW = HEADER_ROW + 1


def _cell_text_lower(cell: Cell) -> str:
    v = cell.value
    return "" if v is None else str(v).strip().lower()


def _coerce_excel_datetime(value: Union[None, int, float, str, datetime], epoch) -> Optional[datetime]:
    from datetime import datetime, timezone
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, (int, float)):
        try:
            dt = excel_serial_to_dt(value, epoch=epoch)
            return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        except Exception:
            return None
    if isinstance(value, str):
        txt = value.strip()
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S",
                    "%d/%m/%Y", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(txt, fmt).replace(tzinfo=timezone.utc)
            except Exception:
                pass
        return None
    return None


def _matches(warn_cell: Cell, last_edit_cell: Cell, *, cutoff_days: int, epoch, now_utc: datetime) -> bool:
    if "deprecated" not in _cell_text_lower(warn_cell):
        return False
    last_dt = _coerce_excel_datetime(last_edit_cell.value, epoch=epoch)
    if last_dt is None:
        return False
    return (now_utc - last_dt).days > cutoff_days


def _copy_row_values(src_ws, dst_ws, src_r: int, dst_r: int) -> None:
    for c in range(1, src_ws.max_column + 1):
        dst_ws.cell(row=dst_r, column=c, value=src_ws.cell(src_r, c).value)


def generate_deactivation_workbook(
    *,
    input_bytes: bytes,
    cutoff_days: int,
    now_utc: Optional[datetime] = None,
) -> Tuple[BytesIO, Dict[str, int]]:
    """
    For each sheet:
      • keep row 1 and 2 as-is
      • keep data rows (>=3) where AK contains 'deprecated' (case-insensitive)
        AND AN is older than cutoff_days
      • set AO='D' on each kept row
    Sheets with zero matches are omitted.

    Notes:
      • Works for .xlsx and .xlsm; macros are ignored (we copy values only).
    """
    now_utc = now_utc or datetime.now(timezone.utc)

    # keep_vba=False ensures macros aren’t preserved even for .xlsm inputs
    wb_in = load_workbook(filename=BytesIO(input_bytes), data_only=True, keep_vba=False)

    out = Workbook()
    tmp = out.active
    tmp.title = "_tmp"

    stats: Dict[str, int] = {}

    for src in wb_in.worksheets:
        if src.max_row < HEADER_ROW:
            continue

        matched = []
        for r in range(DATA_START_ROW, src.max_row + 1):
            if _matches(
                src.cell(r, WARNING_COL),
                src.cell(r, LAST_EDIT_COL),
                cutoff_days=cutoff_days,
                epoch=wb_in.epoch,
                now_utc=now_utc,
            ):
                matched.append(r)

        if not matched:
            continue

        dst = out.create_sheet(title=src.title)
        _copy_row_values(src, dst, 1, 1)
        _copy_row_values(src, dst, 2, 2)

        out_r = 3
        for src_r in matched:
            _copy_row_values(src, dst, src_r, out_r)
            dst.cell(row=out_r, column=OPERATION_COL, value="D")
            dst.cell(row=out_r, column=WARNING_COL, value=None)
            out_r += 1

        stats[src.title] = len(matched)

    if len(out.worksheets) > 1 and out.worksheets[0].title == "_tmp":
        out.remove(out.worksheets[0])

    buf = BytesIO()
    out.save(buf)
    buf.seek(0)
    return buf, stats
