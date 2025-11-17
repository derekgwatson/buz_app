from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging


logger = logging.getLogger(__name__)


def get_current_buz_fabrics(db_manager, curtain_groups=None):
    """
    Fetch current fabrics from the inventory_items table, filtering by curtain fabric groups.

    Args:
        db_manager: DatabaseManager instance.
        curtain_groups: List/tuple of curtain group codes. Defaults to ['CRTWT', 'CRTNT', 'ROMNBQCS'].

    Returns:
        list: List of row dictionaries (sqlite3.Row objects).
    """
    if curtain_groups is None:
        curtain_groups = ['CRTWT', 'CRTNT', 'ROMNBQCS']

    placeholders = ','.join('?' * len(curtain_groups))
    query = f"SELECT * FROM inventory_items WHERE inventory_group_code IN ({placeholders})"
    cursor = db_manager.execute_query(query, params=list(curtain_groups))
    rows = cursor.fetchall()
    return rows


def build_buz_dict(db_rows):
    """
    Build a dictionary mapping SupplierProductCode → list of matching row dicts (one for each group).

    Args:
        db_rows (list): List of sqlite3.Row items from the database.

    Returns:
        dict: Mapping SupplierProductCode → list of row dicts.
    """
    buz_dict = {}
    for row in db_rows:
        code = row['SupplierProductCode'].strip()
        if code not in buz_dict:
            buz_dict[code] = []
        buz_dict[code].append(row)
    return buz_dict


class InventoryWorkbookCreator:
    def __init__(self, headers_config, parse_headers_func):
        self.headers_config = headers_config
        self.parse_headers_func = parse_headers_func
        self.inventory_file_excel_headers, self.inventory_file_db_fields = self.parse_headers_func(
            self.headers_config, "buz_inventory_item_file"
        )
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.sheets = {}

    def _get_or_create_sheet(self, group_name):
        if group_name not in self.sheets:
            ws = self.workbook.create_sheet(title=group_name)
            ws.append([])  # Row 1 blank
            ws.append(self.inventory_file_excel_headers)  # Row 2: Column headings
            self.sheets[group_name] = ws
        return self.sheets[group_name]

    def _add_items_to_sheet(self, group_name, items):
        ws = self._get_or_create_sheet(group_name)
        for item in items:
            item_dict = dict(item)
            row = [item_dict.get(db_field, "") for db_field in self.inventory_file_db_fields]
            ws.append(row)

    def populate_workbook(self, changes=None):
        for group, items in changes.items():
            self._add_items_to_sheet(group, items)

    def save_workbook(self, output_path):
        if not self.sheets:
            logger.info("No pricing updates found. No workbook created.")
            return None  # Optional: return None to indicate no file was saved
        self.workbook.save(output_path)
        return self.workbook

    def save_to_buffer(self):
        """
        Save the workbook to a BytesIO buffer for in-memory file generation.

        Returns:
            BytesIO: Buffer containing the Excel file, or None if no sheets exist.
        """
        if not self.sheets:
            logger.info("No pricing updates found. No workbook created.")
            return None

        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def auto_fit_columns(self):
        for sheet_name, sheet in self.sheets.items():
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 2


def create_inventory_workbook_creator(app):
    from services.helper import parse_headers
    return InventoryWorkbookCreator(
        headers_config=app.config["headers"],
        parse_headers_func=parse_headers
    )
