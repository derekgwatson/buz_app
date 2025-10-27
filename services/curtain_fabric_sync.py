# services/curtain_sync.py
# DB-driven Curtain sync: reads inventory & pricing from DB, Google from XLSX (for now)
# Writes: items_upload.xlsx, pricing_upload.xlsx, change_log.csv

import sys
import re
import os
import time
import pandas as pd
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from collections import Counter
from openpyxl import Workbook
from typing import Any, Mapping
from services.excel_safety import save_workbook_gracefully

# Title case known uppers
KNOWN_ACRONYMS = {
    "II", "FR", "UV", "PVC", "GSM", "LED", "AC", "DC", "LM", "SQM", "GST", "DD", "POA"
}

# =========================
# Config: curtain tabs + product names (loaded from config)
# =========================
# These will be initialized from config in generate_uploads_from_db
# For backwards compatibility with CLI usage, we provide defaults
CURTAIN_TABS = ["CRTWT", "CRTNT", "ROMNDC"]  # default fallback
PRODUCT_NAME_BY_TAB = {
    "CRTWT": "Curtain and Track",
    "CRTNT": "Curtain Only",
    "ROMNDC": "Decorative Romans",
}

# Google sheet settings
GS_TAB = "ML new"
COL_BRAND   = "Brand Name"
COL_FABRIC  = "Fabric Name"
COL_COLOUR  = "Colour"
COL_WIDTH   = "Width (cm)"
COL_DIR     = "Direction"
COL_REPEAT  = "Vertical Pattern Repeat Size (cm)"
COL_COST    = "Cost to DD per metre ROLL (ex GST)"
COL_SELL    = "Proposed NEW Price"

# Inventory (items) headers (Excel-style names we output)
INV_COL_PKID = "PkId"
INV_COL_CODE    = "Code*"
INV_COL_TAXRATE = "Tax Rate"
INV_COL_DESCN1  = "DescnPart1 (Material)"
INV_COL_DESCN2  = "DescnPart2 (Material Types)"
INV_COL_DESCN3  = "DescnPart3 (Colour)"
INV_COL_DESC    = "Description*"
INV_COL_OP      = "Operation"
INV_COL_ACTIVE  = "Active"
INV_COL_PACKSZ  = "Custom Var 1 (PackSize)"    # repeat (mm)
INV_COL_PACKOPT = "Custom Var 2 (PackOpt)"     # width (mm)
INV_COL_PACKTYPE= "Custom Var 3 (PackType)"    # direction (C/B/D)

# Pricing headers (logical)
PRICE_COL_CODE      = "Inventory Code"
PRICE_COL_DESC      = "Description"
PRICE_COL_OP        = "Operation"
PRICE_COL_SELL_W    = "SellLMWide"
PRICE_COL_SELL_H    = "SellLMHeight"
PRICE_COL_COST_W    = "CostLMWide"
PRICE_COL_COST_H    = "CostLMHeight"
PRICE_COL_DATE_FROM = "Date From"
PKID_COL            = "PkId"   # present in pricing template (must be blank)


NEW_FABRIC_EFF_DATE_STR = "01/01/2020"


# Progress logging
VERBOSE = True
PROG_EVERY = 50
t0 = time.perf_counter()


def log(msg="", end="\n"):
    if VERBOSE:
        print(msg, end=end); sys.stdout.flush()


# ---------- Helpers ----------
def _headers_from_config(headers_cfg: dict, key: str) -> list[str]:
    """Return the ordered 'spreadsheet_column' list for a given config block."""
    return [h["spreadsheet_column"] for h in headers_cfg[key]]


def _normalize_items_columns_for_config(df, items_headers: list[str]):
    """
    Our in-memory items DF uses 'Code*' and 'Description*'.
    Config uses 'Code' and 'Description'. Rename to match config headers.
    """
    rename_map = {}
    if "Code*" in df.columns and "Code" in items_headers:
        rename_map["Code*"] = "Code"
    if "Description*" in df.columns and "Description" in items_headers:
        rename_map["Description*"] = "Description"
    if rename_map:
        df = df.rename(columns=rename_map)
    return df


def to_mm_from_cm(val):
    s = str(val).strip() if not pd.isna(val) else ""
    if not s: return ""
    try:
        cm = float(s)
        return str(int(round(cm * 10)))  # cm → mm
    except Exception:
        return ""


def norm_dir_first_letter(val):
    s = str(val or "").strip()
    if not s:
        return ""
    return s[0].upper()  # B / C / D


def build_key(p1, p2, p3):
    return f"{str(p1).strip().lower()}||{str(p2).strip().lower()}||{str(p3).strip().lower()}"


def next_code_for_group(existing_codes, group_prefix, start=10000):
    pat = re.compile(rf"^{re.escape(group_prefix)}(\d+)$")
    max_seen = start - 1
    for c in existing_codes:
        m = pat.match(str(c).strip())
        if m:
            try:
                n = int(m.group(1))
                if n > max_seen:
                    max_seen = n
            except:
                pass
    return f"{group_prefix}{max_seen+1}"[:20]


def to_title_case(s: str) -> str:
    """Title-case words but force known acronyms to UPPER.
    Preserves punctuation: () , . / - etc."""
    s = str(s or "").strip()
    if not s:
        return ""

    def cap_word(tok: str) -> str:
        # split on any non-word chunk, keep delimiters
        parts = re.split(r"([^\w]+)", tok)
        if len(parts) > 1:
            return "".join(cap_word(p) if i % 2 == 0 else p
                           for i, p in enumerate(parts))
        u = tok.upper()
        return u if u in KNOWN_ACRONYMS else tok[:1].upper() + tok[1:].lower()

    # preserve original spacing
    return "".join(cap_word(t) if not t.isspace() else t
                   for t in re.split(r"(\s+)", s))


def colour_for_parts(p3: str) -> str:
    raw = str(p3).strip()
    if not raw:
        return ""
    return to_title_case(raw)


def colour_for_description(p3: str) -> str:
    raw = str(p3).strip()
    if raw.lower() == "specified below":
        return ""
    if raw.lower() == "to be confirmed":
        return "Colour To Be Confirmed"
    return to_title_case(raw)


def rebuild_description(product_name, p1, p2, p3):
    brand  = to_title_case(p1)
    fabric = to_title_case(p2)
    colour = colour_for_description(p3)
    parts = [product_name.strip(), brand, fabric, colour]
    return " ".join([p for p in parts if p])


def _q2(x: str) -> Decimal:
    s = str(x).strip()
    if s == "": return Decimal("0.00")
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def tomorrow_ddmmyyyy():
    return (datetime.today() + timedelta(days=1)).strftime("%d/%m/%Y")


def _df(db: "DatabaseManager", sql: str, params: tuple | list | None = None) -> pd.DataFrame:
    """
    Run SQL via your DatabaseManager and return a pandas DataFrame.
    No fallback to raw sqlite.
    """
    if not hasattr(db, "execute_query"):
        raise TypeError("generate_uploads_from_db expects a DatabaseManager")

    params = params or []
    cur = db.execute_query(sql, params)
    rows = cur.fetchall()
    # When row_factory=sqlite3.Row this yields Row objects; turn into dicts.
    if rows and hasattr(rows[0], "keys"):
        rows = [dict(r) for r in rows]
    return pd.DataFrame(rows)


def _df_from_gsheets(svc, spreadsheet_id: str, worksheet: str) -> pd.DataFrame:
    """
    Use GoogleSheetsService to fetch a worksheet and make a pandas DF with row0 as headers.
    """
    # we don't actually need a range if your helper returns get_all_values()
    rows = svc.fetch_sheet_data(spreadsheet_id, f"{worksheet}!A:ZZZ")  # range is ignored by your helper
    if not rows or not rows[0]:
        raise RuntimeError(f"Google sheet '{worksheet}' is empty or unreadable.")

    header = [h.strip() for h in rows[0]]
    body   = rows[1:]

    # Pad rows to header length so pandas doesn't drop trailing blanks
    width = len(header)
    fixed = [r + [""] * (width - len(r)) if len(r) < width else r[:width] for r in body]

    return pd.DataFrame(fixed, columns=header, dtype=str)


# ---------- DB loaders ----------
# Adapt these mappings if your column names differ.
INV_DB_TO_EXPORT = {
    "pkid": INV_COL_PKID,
    "code": INV_COL_CODE,
    "description": INV_COL_DESC,
    "tax_rate": INV_COL_TAXRATE,
    "descnpart1": INV_COL_DESCN1,
    "descnpart2": INV_COL_DESCN2,
    "descnpart3": INV_COL_DESCN3,
    "active": INV_COL_ACTIVE,
    "custom_var_1": INV_COL_PACKSZ,   # repeat (mm)
    "custom_var_2": INV_COL_PACKOPT,  # width (mm)
    "custom_var_3": INV_COL_PACKTYPE, # direction (C/B/D)
    # computed later:
    # "Operation": INV_COL_OP,
    # "_group", "_key"
}

PR_DB_COLS = {
    "inventory_code": PRICE_COL_CODE,
    "description": PRICE_COL_DESC,
    "sellsqmw": PRICE_COL_SELL_W,
    "sellsqmh": PRICE_COL_SELL_H,
    "costsqmw": PRICE_COL_COST_W,
    "costsqmh": PRICE_COL_COST_H,
    "operation": PRICE_COL_OP,
    # "PkId" blanked on output
}


def _safe_boolish(s):
    return str(s).strip().upper() in ("TRUE","YES","1","Y","T")


def load_inventory_by_group(db, curtain_tabs=None):
    """
    Returns dict[group_code] -> DataFrame mapped to export column names.
    Uses DatabaseManager exclusively.
    """
    if curtain_tabs is None:
        curtain_tabs = CURTAIN_TABS

    # Build SQL IN clause dynamically
    placeholders = ','.join('?' * len(curtain_tabs))
    q = f"""
      SELECT
        Pkid            AS pkid,
        Code            AS code,
        Description     AS description,
        TaxRate         AS tax_rate,
        DescnPart1      AS descnpart1,
        DescnPart2      AS descnpart2,
        DescnPart3      AS descnpart3,
        Active          AS active,
        CustomVar1      AS custom_var_1,   -- repeat (mm)
        CustomVar2      AS custom_var_2,   -- width  (mm)
        CustomVar3      AS custom_var_3,   -- direction
        inventory_group_code
      FROM inventory_items
      WHERE inventory_group_code IN ({placeholders});
    """
    df = _df(db, q, params=list(curtain_tabs)).fillna("")
    if df.empty:
        return {g: pd.DataFrame(columns=list(INV_DB_TO_EXPORT.values()) + ["_group","_key", INV_COL_OP]) for g in curtain_tabs}

    df_ren = df.rename(columns=INV_DB_TO_EXPORT)
    df_ren["_group"] = df["inventory_group_code"].astype(str).str.strip()
    df_ren["_key"] = df.apply(
        lambda r: build_key(r.get("descnpart1",""), r.get("descnpart2",""), r.get("descnpart3","")),
        axis=1
    )

    for c in [INV_COL_OP, INV_COL_ACTIVE, INV_COL_PACKSZ, INV_COL_PACKOPT, INV_COL_PACKTYPE,
              INV_COL_TAXRATE, INV_COL_CODE, INV_COL_DESC, INV_COL_DESCN1, INV_COL_DESCN2, INV_COL_DESCN3,
              INV_COL_PKID]:
        if c not in df_ren.columns:
            df_ren[c] = ""

    out = {}
    for g in curtain_tabs:
        out[g] = df_ren[df_ren["_group"] == g].copy().reset_index(drop=True)
    return out


def load_latest_pricing_by_code(db):
    """
    Return pricing rows with friendly column names + parsed Date From.
    Uses DatabaseManager exclusively.
    """
    q = """
      SELECT
        InventoryCode AS inventory_code,
        Description   AS description,
        DateFrom      AS date_from,
        SellLMWide    AS sellsqmw,
        SellLMHeight  AS sellsqmh,
        CostLMWide    AS costsqmw,
        CostLMHeight  AS costsqmh,
        Operation     AS operation
      FROM pricing_data;
    """
    pr = _df(db, q).fillna("")
    return pr.rename(columns=PR_DB_COLS)


# ---------- Main orchestrator (DB-based) ----------
def generate_uploads_from_db(
    google_source: Any,
    db,
    output_dir: str = ".",
    write_change_log: bool = False,
    headers_cfg: dict | None = None,
    curtain_fabric_groups: dict | None = None,
    progress=None,
):
    if headers_cfg is None or "buz_inventory_item_file" not in headers_cfg or "buz_pricing_file" not in headers_cfg:
        raise RuntimeError("headers config is required (buz_inventory_item_file and buz_pricing_file).")

    # Extract curtain groups config (use module defaults if not provided)
    if curtain_fabric_groups:
        curtain_tabs = list(curtain_fabric_groups.keys())
        product_name_by_tab = curtain_fabric_groups
    else:
        curtain_tabs = CURTAIN_TABS
        product_name_by_tab = PRODUCT_NAME_BY_TAB

    def _ping(msg: str, pct: int | None = None):
        if callable(progress):
            try:
                progress(msg, pct)
            except Exception:
                pass

    t0 = time.perf_counter()
    _ping("Starting Curtain Sync…", 1)

    # --- Load Google (raw, then validate)
    all_cols_needed = [COL_BRAND, COL_FABRIC, COL_COLOUR, COL_WIDTH, COL_DIR,
                       COL_REPEAT, COL_COST, COL_SELL]
    _ping("Loading Google sheet…", 3)
    # NEW: detect which kind of source we received
    if isinstance(google_source, (str, os.PathLike)):
        # old behaviour: XLSX export
        g_raw = pd.read_excel(google_source, sheet_name=GS_TAB, dtype=str)
    elif isinstance(google_source, Mapping) and "svc" in google_source:
        # new behaviour: live Google Sheet
        svc = google_source["svc"]
        spreadsheet_id = google_source["spreadsheet_id"]
        worksheet = google_source.get("worksheet", GS_TAB)
        g_raw = _df_from_gsheets(svc, spreadsheet_id, worksheet)
    else:
        raise RuntimeError("google_source must be a file path or a dict with keys: svc, spreadsheet_id, worksheet")

    _ping("Validating Google sheet rows…", 6)
    dupes = g_raw.columns[g_raw.columns.duplicated()].unique()
    if len(dupes) > 0:
        raise RuntimeError(f"Duplicate columns in Google sheet: {', '.join(dupes)}")

    missing_cols = [c for c in all_cols_needed if c not in g_raw.columns]
    if missing_cols:
        raise RuntimeError(f"Missing required columns in Google sheet: {', '.join(missing_cols)}")

    # validate rows
    issues = []
    for idx, r in g_raw.iterrows():
        excel_row = idx + 2
        b = str(r.get(COL_BRAND, "")).strip()
        f = str(r.get(COL_FABRIC, "")).strip()
        c = str(r.get(COL_COLOUR, "")).strip()
        w = str(r.get(COL_WIDTH, "")).strip()
        d = str(r.get(COL_DIR, "")).strip()
        rep = str(r.get(COL_REPEAT, "")).strip()
        cost = str(r.get(COL_COST, "")).strip()
        sell = str(r.get(COL_SELL, "")).strip()

        row_issues = []
        if not b: row_issues.append("Brand missing")
        if not f: row_issues.append("Fabric missing")
        if not c: row_issues.append("Colour missing")

        try: float(w)
        except Exception: row_issues.append("Width (cm) missing or not numeric")

        try: float(rep)
        except Exception: row_issues.append("Vertical Repeat (cm) missing or not numeric")

        d_letter = norm_dir_first_letter(d)
        if d_letter not in {"B", "C", "D"}:
            row_issues.append("Direction invalid (must begin with B/C/D)")

        if not cost: row_issues.append("Cost missing")
        if not sell: row_issues.append("Sell missing")

        if row_issues:
            issues.append({
                "Row": excel_row, "Brand": b, "Fabric": f, "Colour": c,
                "Problems": "; ".join(row_issues)
            })

    if issues:
        err_df = pd.DataFrame(issues).sort_values(by="Row")
        lines = ["Validation failed: problematic rows in Google sheet:\n"]
        for _, r in err_df.iterrows():
            lines.append(f" Row {r['Row']}: Brand='{r['Brand']}', Fabric='{r['Fabric']}', "
                         f"Colour='{r['Colour']}' -> {r['Problems']}")
        raise RuntimeError("\n".join(lines))

    # Build clean Google df
    g = g_raw[all_cols_needed].copy()
    g = g[g[COL_BRAND].notna()].copy()
    g["_key"]       = g.apply(lambda r: build_key(r.get(COL_BRAND,""), r.get(COL_FABRIC,""), r.get(COL_COLOUR,"")), axis=1)
    g["_width_mm"]  = g[COL_WIDTH].apply(to_mm_from_cm)
    g["_repeat_mm"] = g[COL_REPEAT].apply(to_mm_from_cm)
    g["_dir"]       = g[COL_DIR].apply(norm_dir_first_letter)
    g["_sell"]      = g[COL_SELL].astype(str).str.strip()
    g["_cost"]      = g[COL_COST].astype(str).str.strip()
    g = g.set_index("_key", drop=False)
    keys_g = set(g.index)
    _ping(f"Google sheet OK: {len(g)} fabrics", 10)

    # --- Load Inventory (from DB)
    _ping("Loading inventory from DB…", 12)
    inv_by_group = load_inventory_by_group(db, curtain_tabs=curtain_tabs)
    _ping("Inventory loaded", 18)

    for sheet in curtain_tabs:
        df = inv_by_group.get(sheet, pd.DataFrame())
        if df.empty:
            log(f"[2/5] Inventory tab {sheet}: 0 rows")
        else:
            # HARD EXIT if any blank description parts
            blanks = df[
                (df[INV_COL_DESCN1].astype(str).str.strip() == "") |
                (df[INV_COL_DESCN2].astype(str).str.strip() == "") |
                (df[INV_COL_DESCN3].astype(str).str.strip() == "")
            ]
            if not blanks.empty:
                sample = blanks.head(10)
                msg = [f"Found {len(blanks)} row(s) in inventory group '{sheet}' with blank DescnPart fields."]
                for i, r in sample.iterrows():
                    msg.append(f"  Code={r.get(INV_COL_CODE,'')} "
                               f"DescnPart1='{r.get(INV_COL_DESCN1,'')}', "
                               f"DescnPart2='{r.get(INV_COL_DESCN2,'')}', "
                               f"DescnPart3='{r.get(INV_COL_DESCN3,'')}'")
                if len(blanks) > 10:
                    msg.append(f"  ... and {len(blanks)-10} more rows.")
                raise RuntimeError("\n".join(msg))
            log(f"[2/5] Inventory group {sheet}: {len(df)} rows")

    inv_all = pd.concat(inv_by_group.values(), ignore_index=True) if inv_by_group else pd.DataFrame()
    inv_idx = inv_all.set_index(inv_all.apply(lambda r: build_key(r.get(INV_COL_DESCN1,""), r.get(INV_COL_DESCN2,""), r.get(INV_COL_DESCN3,"")), axis=1))
    keys_inv = set(inv_idx.index)

    # --- Load Pricing (from DB, all rows; we'll filter per-group by joining on inventory)
    _ping("Loading pricing rows…", 20)
    pr_all = load_latest_pricing_by_code(db)

    # --- Normalize and build a key per code, and sort so "latest" is last per code
    pr_all[PRICE_COL_CODE] = pr_all[PRICE_COL_CODE].astype(str).str.strip()

    # Use the plain inventory code as the key
    pr_all["_code_key"] = pr_all[PRICE_COL_CODE]

    # If date_from exists, use it to decide "latest" row; otherwise preserve input order.
    if "date_from" in pr_all.columns:
        pr_all["_date_from_dt"] = pd.to_datetime(pr_all["date_from"], errors="coerce", dayfirst=True)
        # Stable sort so tail(1) reliably gets the newest per code
        pr_all = pr_all.sort_values(["_code_key", "_date_from_dt"], kind="mergesort")
    else:
        pr_all = pr_all.sort_values(["_code_key"], kind="mergesort")

    last_rows = (
        pr_all.groupby("_code_key", as_index=False)
        .tail(1)  # last (most recent) row per code
    )

    last_price_map: dict[str, tuple[Decimal | None, Decimal | None]] = {}
    for _, r in last_rows.iterrows():
        last_price_map[r["_code_key"]] = (
            _q2(r.get(PRICE_COL_SELL_W, "")),
            _q2(r.get(PRICE_COL_COST_W, "")),
        )

    _ping(f"Pricing codes: {len(last_price_map)}", 25)

    pricing_to_append = {s: [] for s in curtain_tabs}
    change_log = []

    # --- Process existing rows (D/E/reactivate/desc) and queue pricing if 2dp changed
    for grp, df in inv_by_group.items():
        _ping(f"Processing existing rows in {grp}…", 30)
        product_name = product_name_by_tab[grp]
        log(f"[4/5] Processing existing rows in {grp} …")
        for i, row in df.iterrows():
            key = row["_key"]
            if key not in keys_g:
                active_val = str(row.get(INV_COL_ACTIVE,"")).strip().upper()
                if active_val in ("TRUE", "YES", "1"):
                    df.at[i, INV_COL_OP] = "D"
                    change_log.append({
                        "Tab": grp, "Operation": "D",
                        "Code": row.get(INV_COL_CODE,""),
                        "Description": row.get(INV_COL_DESC,""),
                        "Reason": "Deleted (not in Google)"
                    })
                continue

            grow = g.loc[key]
            reasons = []

            # width / repeat / direction
            if grow["_width_mm"] and grow["_width_mm"] != str(row.get(INV_COL_PACKOPT,"")).strip():
                df.at[i, INV_COL_PACKOPT] = grow["_width_mm"]; reasons.append("Width changed")
            if grow["_repeat_mm"] and grow["_repeat_mm"] != str(row.get(INV_COL_PACKSZ,"")).strip():
                df.at[i, INV_COL_PACKSZ] = grow["_repeat_mm"]; reasons.append("Repeat changed")
            if grow["_dir"] and grow["_dir"] != str(row.get(INV_COL_PACKTYPE,"")).strip():
                df.at[i, INV_COL_PACKTYPE] = grow["_dir"]; reasons.append("Direction changed")

            # Reactivate if inactive
            if not _safe_boolish(row.get(INV_COL_ACTIVE,"")):
                df.at[i, INV_COL_ACTIVE] = "TRUE"; reasons.append("Reactivated")

            # Rebuild description
            new_desc = rebuild_description(product_name, grow[COL_BRAND], grow[COL_FABRIC], grow[COL_COLOUR])
            old_desc = str(row.get(INV_COL_DESC,"")).rstrip("*").strip()
            if new_desc and new_desc != old_desc:
                df.at[i, INV_COL_DESC] = new_desc; reasons.append("Description corrected")

            if reasons:
                df.at[i, INV_COL_OP] = "E"
                change_log.append({"Tab": grp, "Operation": "E",
                                   "Code": str(row.get(INV_COL_CODE,"")).strip(),
                                   "Description": new_desc or old_desc,
                                   "Reason": "; ".join(reasons)})

            # --- Pricing (append only if price OR cost changed at 2dp, date = tomorrow)
            code = str(row.get(INV_COL_CODE,"")).strip()
            if code:
                # Source values from Google
                sheet_sell_q2 = _q2(grow["_sell"])
                sheet_cost_q2 = _q2(grow["_cost"])

                # Find most recent existing pricing row for this code
                last_sell_q2, last_cost_q2 = last_price_map.get(code, (None, None))

                changed = (
                    last_sell_q2 is None or last_cost_q2 is None or
                    sheet_sell_q2 != last_sell_q2 or sheet_cost_q2 != last_cost_q2
                )
                if changed:
                    pricing_to_append[grp].append({
                        PRICE_COL_CODE: code,
                        PRICE_COL_DESC: new_desc or old_desc,
                        PRICE_COL_DATE_FROM: tomorrow_ddmmyyyy(),
                        PRICE_COL_SELL_W: f"{sheet_sell_q2:.2f}",
                        PRICE_COL_SELL_H: f"{sheet_sell_q2:.2f}",
                        PRICE_COL_COST_W: f"{sheet_cost_q2:.2f}",
                        PRICE_COL_COST_H: f"{sheet_cost_q2:.2f}",
                        PRICE_COL_OP: "A",
                        PKID_COL: ""
                    })

                    # NEW: reflect this in the UI change log
                    def fmt(x): return "—" if x is None else f"{x:.2f}"

                    change_log.append({
                        "Tab": grp,
                        # Keep using E so it fits the existing UI table; summary counts won’t be affected,
                        # because those come from the item DataFrames not this list.
                        "Operation": "P",
                        "Code": code,
                        "Description": new_desc or old_desc,
                        "Reason": (
                            f"Pricing changed "
                            f"(SellW {fmt(last_sell_q2)}→{sheet_sell_q2:.2f}, "
                            f"CostW {fmt(last_cost_q2)}→{sheet_cost_q2:.2f})"
                        )
                    })

            if (i + 1) % PROG_EVERY == 0:
                log(f"  • {grp}: {i + 1}/{len(df)}\n")  # newline so it’s visible next to Werkzeug logs
                pct_here = 30 + int(20 * ((i + 1) / max(1, len(df))))  # 30–50% across this pass
                _ping(f"{grp}: {i + 1}/{len(df)}", pct_here)
        _ping(f"{grp}: done", 50)

    # --- Adds (ensure fabric exists on both tabs)
    codes_by_group = {grp: set(df[INV_COL_CODE].astype(str).str.strip()) for grp, df in inv_by_group.items()}
    new_keys = sorted(keys_g - keys_inv)
    _ping(f"Scanning for new fabrics: {len(new_keys)} fabrics", 56)
    for idx, key in enumerate(new_keys, 1):
        grow = g.loc[key]
        for grp in curtain_tabs:
            df = inv_by_group[grp]
            keys_by_group = {grp: set(df["_key"]) for grp, df in inv_by_group.items()}
            if key in keys_by_group[grp]:
                continue

            code = next_code_for_group(codes_by_group[grp], grp, start=10000)
            codes_by_group[grp].add(code)
            product_name = product_name_by_tab[grp]
            new_desc = rebuild_description(product_name, grow[COL_BRAND], grow[COL_FABRIC], grow[COL_COLOUR])

            new_row = {col: "" for col in df.columns}
            new_row["_group"] = grp
            new_row["_key"] = key
            new_row[INV_COL_PKID] = ""
            new_row[INV_COL_CODE] = code
            new_row[INV_COL_DESCN1] = to_title_case(grow[COL_BRAND])
            new_row[INV_COL_DESCN2] = to_title_case(grow[COL_FABRIC])
            new_row[INV_COL_DESCN3] = colour_for_parts(grow[COL_COLOUR])
            new_row[INV_COL_PACKOPT] = grow["_width_mm"]
            new_row[INV_COL_PACKTYPE] = grow["_dir"]
            new_row[INV_COL_PACKSZ] = grow["_repeat_mm"]
            new_row[INV_COL_TAXRATE] = "GST"
            new_row[INV_COL_ACTIVE] = "TRUE"
            new_row[INV_COL_DESC] = new_desc
            new_row[INV_COL_OP] = "A"

            inv_by_group[grp] = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            keys_by_group[grp].add(key)  # keep the set in sync
            pricing_to_append[grp].append({
                PRICE_COL_CODE: code,
                PRICE_COL_DESC: new_desc,
                PRICE_COL_DATE_FROM: NEW_FABRIC_EFF_DATE_STR,
                PRICE_COL_SELL_W: f"{_q2(grow['_sell']):.2f}",
                PRICE_COL_SELL_H: f"{_q2(grow['_sell']):.2f}",
                PRICE_COL_COST_W: f"{_q2(grow['_cost']):.2f}",
                PRICE_COL_COST_H: f"{_q2(grow['_cost']):.2f}",
                PRICE_COL_OP:   "A",
                PKID_COL: ""
            })
            change_log.append({"Tab": grp, "Operation": "A",
                               "Code": code, "Description": new_desc,
                               "Reason": "New fabric"})
        if VERBOSE and idx % PROG_EVERY == 0:
            log(f"  • adds: {idx}/{len(new_keys)}", end="\r")
    _ping("Adds pass complete", 70)

    # --- Summarise changes
    summary = {"A":0,"E":0,"D":0}
    per_tab = {t:{"A":0,"E":0,"D":0} for t in curtain_tabs}
    for grp, df in inv_by_group.items():
        vc = df[df[INV_COL_OP].isin(["A","E","D"])][INV_COL_OP].value_counts().to_dict()
        for k,v in vc.items(): summary[k]+=v; per_tab[grp][k]=v
    log("Changes summary")
    log(f"  Total  A:{summary['A']:>6}  E:{summary['E']:>6}  D:{summary['D']:>6}")
    for grp in curtain_tabs:
        a,e,d = per_tab[grp].get("A",0), per_tab[grp].get("E",0), per_tab[grp].get("D",0)
        log(f"  {grp:5} A:{a:>6}  E:{e:>6}  D:{d:>6}")

    _ping("Writing items workbook…", 75)

    # --- Write items upload (blank row1, headers row2, data row3+, blank col AP after AO)
    os.makedirs(output_dir, exist_ok=True)

    # --- Write items upload from CONFIG (row1 blank, row2 headers, data row3+, trailing blank col)
    items_headers = _headers_from_config(headers_cfg, "buz_inventory_item_file")

    items_wb = Workbook(); items_wb.remove(items_wb.active)
    for grp, df in inv_by_group.items():
        delta = df[df[INV_COL_OP].isin(["A","E","D"])].copy()
        if delta.empty:
            continue

        # make column names match config headers (Code*/Description* → Code/Description)
        delta = _normalize_items_columns_for_config(delta, items_headers)

        # ensure every required header exists as a column (fill blanks)
        for h in items_headers:
            if h not in delta.columns:
                delta[h] = ""

        # keep only configured columns, in order
        to_write = delta[items_headers]

        ws = items_wb.create_sheet(title=grp)
        ws.append([])                      # row 1 blank
        ws.append(items_headers + [""])    # row 2 headers + one trailing blank cell
        for _, r in to_write.iterrows():
            ws.append(list(r.values) + [""])  # trailing blank cell (Buz quirk)

    items_path = os.path.join(output_dir, "items_upload.xlsx")
    has_real_data = save_workbook_gracefully(items_wb, items_path)
    if has_real_data:
        _ping("Writing item workbook…", 85)
    else:
        _ping("No item data — exported a placeholder workbook.", 85)

    # --- Write pricing upload from CONFIG (preserve order per config)
    pricing_headers = _headers_from_config(headers_cfg, "buz_pricing_file")

    pwb = Workbook(); pwb.remove(pwb.active)
    for grp in curtain_tabs:
        ws = pwb.create_sheet(title=grp)
        ws.append(pricing_headers)

        rows = pricing_to_append.get(grp, [])
        for row_dict in rows:
            # Ensure all columns present in the expected header are provided
            row_out = []
            for h in pricing_headers:
                if h == "Operation":
                    row_out.append("A")     # enforced as per your logic
                elif h == "PkId":
                    row_out.append("")      # must be blank
                else:
                    row_out.append(row_dict.get(h, ""))  # use keys same as spreadsheet_column
            ws.append(row_out)

    pricing_path = os.path.join(output_dir, "pricing_upload.xlsx")
    has_real_data = save_workbook_gracefully(pwb, pricing_path)
    if has_real_data:
        _ping("Writing pricing workbook.", 87)
    else:
        _ping("No pricing data — exported a placeholder workbook.", 87)

    # --- (Optional) write change log CSV only if requested
    if write_change_log and change_log:
        _ping("Writing change log…", 90)
        pd.DataFrame(change_log).to_csv(os.path.join(output_dir, "change_log.csv"), index=False)
        reasons_counter = Counter()
        for r in change_log:
            for piece in r["Reason"].split(";"):
                reasons_counter[piece.strip()] += 1
        top = ", ".join(f"{k}={v}" for k,v in reasons_counter.most_common())
        log(f"change_log.csv written ({len(change_log)} rows) | Reasons: {top}")
    else:
        log("No CSV change log written (UI will display changes).")

    elapsed = time.perf_counter() - t0
    _ping("Curtain Sync complete ✅", 100)
    log(f"Files written: {items_path}, {pricing_path}")
    log(f"Elapsed: {elapsed:0.1f}s")
    print("✅ Done.")

    return {
        "items_path": items_path,
        "pricing_path": pricing_path,
        "elapsed_sec": elapsed,
        "summary": summary,
        "per_tab": per_tab,
        "change_log": change_log,   # <-- ALWAYS return the list
    }


# ---- Optional CLI (uses DatabaseManager + config headers) ----
if __name__ == "__main__":
    import argparse
    from services.database import create_db_manager
    from services.config_service import ConfigManager

    parser = argparse.ArgumentParser(description="Curtain sync (DB + Google XLSX) -> items/pricing uploads")
    parser.add_argument("google_sheet_xlsx", help="Path to the exported Google Sheet (XLSX)")
    parser.add_argument("--config", default="config.json", help="Path to config.json (default: config.json)")
    parser.add_argument("--out", default="uploads", help="Output directory for the two files (default: uploads)")
    parser.add_argument("--write-change-log", action="store_true", help="Also write change_log.csv to the output dir")
    args = parser.parse_args()

    # Load config (uses your ConfigManager)
    cm = ConfigManager(args.config)
    db_path = cm.get("database", default="buz_data.db")
    headers_cfg = cm.get("headers", default={})
    if not headers_cfg:
        raise SystemExit("Missing headers in config.json (need buz_inventory_item_file and buz_pricing_file).")

    # Open DB via your DatabaseManager
    db = create_db_manager(db_path)
    try:
        res = generate_uploads_from_db(
            args.google_sheet_xlsx,
            db=db,
            output_dir=args.out,
            write_change_log=args.write_change_log,
            headers_cfg=headers_cfg,
        )
        print("Items:", res["items_path"])
        print("Pricing:", res["pricing_path"])
    finally:
        db.close()
