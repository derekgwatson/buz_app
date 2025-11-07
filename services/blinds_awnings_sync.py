# services/blinds_awnings_sync.py
"""
Blinds & Awnings Fabric Sync Service

Syncs blinds and awnings fabrics from Google Sheets (Retail/Wholesale tabs) to Buz inventory.
Generates items upload and pricing upload Excel files.
"""

import os
import re
import logging
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple
from collections import defaultdict
from openpyxl import Workbook
import pandas as pd

from services.excel_safety import save_workbook_gracefully

logger = logging.getLogger(__name__)

# Constants
SUPPLIER_NAME = "Unleashed"
TAX_RATE = "GST"
DEPRECATED_WARNING = "Deprecated - DO NOT USE"


# ========== Utility Functions ==========

def _norm(s: Any) -> str:
    """Normalize string value."""
    return ("" if s is None else str(s)).strip()


def _build_desc_key(fd1: str, fd2: str, fd3: str) -> str:
    """Build a normalized key from 3 description parts (lowercase for matching)."""
    return f"{_norm(fd1).lower()}||{_norm(fd2).lower()}||{_norm(fd3).lower()}"


def _q2(x: Any) -> Decimal:
    """Convert to Decimal with 2dp, default to 0.00."""
    try:
        return Decimal(str(x).strip() or "0").quantize(Decimal("0.01"))
    except (InvalidOperation, AttributeError):
        return Decimal("0.00")


def _tomorrow_ddmmyyyy() -> str:
    """Return tomorrow's date in DD/MM/YYYY format."""
    return (datetime.today() + timedelta(days=1)).strftime("%d/%m/%Y")


def _is_wholesale_group(group_code: str) -> bool:
    """Check if a group is wholesale (starts with WS)."""
    return _norm(group_code).upper().startswith("WS")


def _normalize_colour_for_desc(colour: str) -> str:
    """
    Normalize colour for description field.
    If colour is 'To Be Confirmed', prefix with 'Colour'.
    """
    colour_norm = _norm(colour)
    if colour_norm.lower() == "to be confirmed":
        return "Colour To Be Confirmed"
    return colour_norm


def _build_description(prefix: str, fd1: str, fd2: str, fd3: str) -> str:
    """
    Build full description: [prefix] FD1 FD2 FD3
    Handle special case where FD3='To Be Confirmed' → 'Colour To Be Confirmed'
    """
    parts = [
        _norm(prefix),
        _norm(fd1),
        _norm(fd2),
        _normalize_colour_for_desc(fd3)
    ]
    return " ".join(p for p in parts if p)


def _next_code_for_group(existing_codes: List[str], group_prefix: str, start: int = 10000) -> str:
    """
    Generate next sequential code for a group.
    Finds all codes starting with group_prefix, extracts numeric suffix, returns next number.
    """
    pattern = re.compile(rf"^{re.escape(group_prefix)}(\d+)$", re.IGNORECASE)
    max_num = start - 1

    for code in existing_codes:
        match = pattern.match(_norm(code))
        if match:
            try:
                num = int(match.group(1))
                if num > max_num:
                    max_num = num
            except ValueError:
                pass

    return f"{group_prefix}{max_num + 1}"[:20]


def _check_material_restriction(group_code: str, fd2: str, material_restrictions: Dict[str, List[str]]) -> bool:
    """
    Check if FD2 (material) is allowed for this group.
    Returns True if allowed, False if restricted.
    """
    if group_code not in material_restrictions:
        return True  # No restriction = allowed

    allowed_materials = material_restrictions[group_code]
    fd2_norm = _norm(fd2)

    # Check if FD2 contains any of the allowed material keywords
    for material in allowed_materials:
        if material.lower() in fd2_norm.lower():
            return True

    return False


# ========== Google Sheets Data Loading ==========

def load_groups_config_from_sheet(
    sheets_service,
    spreadsheet_id: str,
    buz_template_tab: str,
    progress=None
) -> Dict[str, Dict[str, Any]]:
    """
    Load group configuration from 'Buz template' tab in Google Sheets.

    Expected columns:
        - Product (informational only)
        - Code (inventory group code)
        - Description (description prefix)
        - Price Grid Code
        - Cost Grid Code
        - Discount Group Code
        - Category
        - Markup (optional - pricing markup multiplier)
        - Wastage (optional - wastage percentage, e.g. 20 for 20%)
        - Price Type (optional - "SQM" or "LM", defaults to "SQM")

    Returns dict mapping group code to configuration.
    """
    def _p(msg: str, pct: Optional[int] = None):
        if callable(progress):
            try:
                progress(msg, pct)
            except Exception:
                pass

    _p("Loading Buz template configuration from Google Sheets...", 2)

    rows = sheets_service.fetch_sheet_data(spreadsheet_id, f"{buz_template_tab}!A:Z")

    if not rows or len(rows) < 2:
        raise RuntimeError(f"Buz template tab '{buz_template_tab}' is empty or has no data")

    # Parse header
    headers = [h.strip() for h in rows[0]]

    # Log available columns for debugging
    logger.info(f"Buz template columns: {headers}")

    # Expected columns (Markup and Wastage are optional)
    required = ["Code", "Description", "Price Grid Code", "Cost Grid Code", "Discount Group Code", "Category"]
    for col in required:
        if col not in headers:
            raise RuntimeError(f"Missing required column '{col}' in Buz template tab")

    # Build config dict
    groups_config = {}

    for row in rows[1:]:
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))

        row_dict = {headers[i]: _norm(row[i]) for i in range(len(headers))}

        code = row_dict.get("Code", "")
        if not code:
            continue  # Skip empty rows

        # Convert empty strings to None for grid codes
        price_grid_code = row_dict.get("Price Grid Code") or None
        cost_grid_code = row_dict.get("Cost Grid Code") or None

        # Parse optional markup
        markup_override = None
        markup_str = row_dict.get("Markup", "")
        if markup_str:
            try:
                markup_override = Decimal(markup_str)
            except (InvalidOperation, ValueError):
                pass  # Ignore invalid markup values

        # Parse optional wastage percentage
        wastage_pct = None
        wastage_str = row_dict.get("Wastage", "")
        if wastage_str:
            try:
                # Strip % sign if present, then convert to decimal (e.g., "20%" or "20" -> 0.20)
                wastage_cleaned = wastage_str.rstrip('%').strip()
                wastage_pct = Decimal(wastage_cleaned) / Decimal("100")
                logger.debug(f"Group {code}: Parsed wastage '{wastage_str}' -> {wastage_pct}")
            except (InvalidOperation, ValueError) as e:
                logger.warning(f"Group {code}: Failed to parse wastage value '{wastage_str}': {e}")

        # Parse optional price type (SQM or LM)
        price_type_raw = row_dict.get("Price Type", "")
        price_type = price_type_raw.strip().upper() if price_type_raw else ""
        if price_type not in ["SQM", "LM"]:
            price_type = "SQM"  # Default to SQM

        if price_type == "LM":
            logger.info(f"Group {code}: Price Type = LM (will convert from lineal metre to square metre)")

        groups_config[code] = {
            "description_prefix": row_dict.get("Description", ""),
            "price_grid_code": price_grid_code,
            "cost_grid_code": cost_grid_code,
            "discount_group_code": row_dict.get("Discount Group Code", ""),
            "category": row_dict.get("Category", ""),
            "markup_override": markup_override,
            "wastage_pct": wastage_pct,
            "price_type": price_type
        }

    _p(f"Loaded configuration for {len(groups_config)} groups", 3)

    # Log wastage values for debugging
    wastage_groups = {code: cfg.get("wastage_pct") for code, cfg in groups_config.items() if cfg.get("wastage_pct")}
    if wastage_groups:
        logger.info(f"Loaded wastage for {len(wastage_groups)} groups: {wastage_groups}")
    else:
        logger.warning("No wastage values found in Buz template configuration")

    return groups_config


def load_price_grids_lookup(
    sheets_service,
    spreadsheet_id: str,
    price_grids_tab: str,
    progress=None
) -> Dict[Tuple[str, str], str]:
    """
    Load price grid lookup from 'Price Grids' tab in Google Sheets.

    Expected columns:
        - Code (inventory group code)
        - Price Category (e.g., "89-1", "127-3")
        - Grid Code (e.g., "WG89-1", "WGR127-3")

    Returns dict mapping (group_code, price_category) → grid_code
    """
    def _p(msg: str, pct: Optional[int] = None):
        if callable(progress):
            try:
                progress(msg, pct)
            except Exception:
                pass

    _p("Loading Price Grids lookup from Google Sheets...", 3)

    rows = sheets_service.fetch_sheet_data(spreadsheet_id, f"{price_grids_tab}!A:Z")

    if not rows or len(rows) < 2:
        logger.warning(f"Price Grids tab '{price_grids_tab}' is empty or has no data - wholesale items will have no grid codes")
        return {}

    # Parse header
    headers = [h.strip() for h in rows[0]]

    # Expected columns
    required = ["Code", "Price Category", "Grid Code"]
    for col in required:
        if col not in headers:
            raise RuntimeError(f"Missing required column '{col}' in Price Grids tab")

    # Build lookup dict
    price_grids = {}

    for row in rows[1:]:
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))

        row_dict = {headers[i]: _norm(row[i]) for i in range(len(headers))}

        code = row_dict.get("Code", "")
        price_category = row_dict.get("Price Category", "")
        grid_code = row_dict.get("Grid Code", "")

        if not code or not price_category or not grid_code:
            continue  # Skip empty rows

        key = (code, price_category)
        price_grids[key] = grid_code

    _p(f"Loaded {len(price_grids)} price grid mappings", 4)
    logger.info(f"Loaded {len(price_grids)} price grid mappings for wholesale items")

    return price_grids


def load_fabric_data_from_sheets(
    sheets_service,
    spreadsheet_id: str,
    retail_tab: str,
    wholesale_tab: str,
    groups_config: Dict[str, Dict[str, Any]],
    material_restrictions: Dict[str, List[str]],
    progress=None
) -> Tuple[Dict[str, pd.DataFrame], Dict[str, set]]:
    """
    Load fabric data from Google Sheets Retail and Wholesale tabs.

    Returns:
        - fabrics_by_group: dict mapping group_code to DataFrame with columns: FD1, FD2, FD3, UnleashedCode, Category, Price, _key
        - filtered_by_material: dict mapping group_code to set of _keys that were filtered out due to material restrictions
    """
    def _p(msg: str, pct: Optional[int] = None):
        if callable(progress):
            try:
                progress(msg, pct)
            except Exception:
                pass

    _p("Loading Retail tab from Google Sheets...", 5)
    retail_rows = sheets_service.fetch_sheet_data(spreadsheet_id, f"{retail_tab}!A:Z")

    _p("Loading Wholesale tab from Google Sheets...", 10)
    wholesale_rows = sheets_service.fetch_sheet_data(spreadsheet_id, f"{wholesale_tab}!A:Z")

    def _parse_sheet(rows, is_wholesale=False):
        """Parse sheet rows into DataFrame."""
        if not rows or len(rows) < 2:
            return pd.DataFrame()

        # First row is header
        headers = [h.strip() for h in rows[0]]

        # Expected columns: FD1, FD2, FD3, Unleashed Code, Category, Price (or Price Category for wholesale)
        # Width is optional (needed for retail with LM price types)
        required = ["FD1", "FD2", "FD3", "Unleashed Code", "Category"]
        for col in required:
            if col not in headers:
                raise RuntimeError(f"Missing required column '{col}' in {'Wholesale' if is_wholesale else 'Retail'} tab")

        # Check for Price or Price Category column
        price_col = None
        if "Price" in headers:
            price_col = "Price"
        elif "Price Category" in headers:
            price_col = "Price Category"
        else:
            raise RuntimeError(f"Missing required column 'Price' or 'Price Category' in {'Wholesale' if is_wholesale else 'Retail'} tab")

        # Check for optional Width column (for vertical blind conversion)
        has_width = "Width" in headers

        # Build DataFrame
        data = []
        for row in rows[1:]:
            if len(row) < len(headers):
                row = row + [""] * (len(headers) - len(row))
            data.append(row[:len(headers)])

        df = pd.DataFrame(data, columns=headers, dtype=str).fillna("")

        # Normalize price column to "Price"
        if price_col == "Price Category":
            df["Price"] = df["Price Category"]

        # Filter out empty rows (where FD1, FD2, FD3 are all empty)
        df = df[
            (df["FD1"].str.strip() != "") |
            (df["FD2"].str.strip() != "") |
            (df["FD3"].str.strip() != "")
        ].copy()

        # Add key for matching
        df["_key"] = df.apply(lambda r: _build_desc_key(r["FD1"], r["FD2"], r["FD3"]), axis=1)

        # Log if Width column is available (for vertical blind conversion)
        if has_width and not is_wholesale:
            logger.debug(f"Width column available in {'Wholesale' if is_wholesale else 'Retail'} tab for LM→SQM conversion")

        return df

    _p("Parsing Retail data...", 15)
    retail_df = _parse_sheet(retail_rows, is_wholesale=False)

    _p("Parsing Wholesale data...", 20)
    wholesale_df = _parse_sheet(wholesale_rows, is_wholesale=True)

    _p("Mapping fabrics to inventory groups...", 25)

    # Build mapping: group_code → DataFrame of fabrics for that group
    fabrics_by_group = {}
    filtered_by_material = {}

    for group_code, group_cfg in groups_config.items():
        category = group_cfg.get("category", "")
        is_wholesale = _is_wholesale_group(group_code)

        # Select appropriate dataframe
        source_df = wholesale_df if is_wholesale else retail_df

        # Filter by category
        group_df = source_df[source_df["Category"].str.strip() == category].copy()

        # Apply material restrictions and track filtered items
        if group_code in material_restrictions:
            before_keys = set(group_df["_key"].tolist())
            group_df = group_df[
                group_df.apply(
                    lambda r: _check_material_restriction(group_code, r["FD2"], material_restrictions),
                    axis=1
                )
            ].copy()
            after_keys = set(group_df["_key"].tolist())
            filtered_by_material[group_code] = before_keys - after_keys
        else:
            filtered_by_material[group_code] = set()

        fabrics_by_group[group_code] = group_df

    _p(f"Loaded fabrics for {len(fabrics_by_group)} groups", 30)
    return fabrics_by_group, filtered_by_material


# ========== Database Loading ==========

def load_existing_buz_inventory(db, groups_config: Dict[str, Dict[str, Any]]) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Dict[str, str]]]:
    """
    Load existing Buz inventory for all blinds/awnings groups.

    Returns:
        - inv_by_group: dict[group_code] → DataFrame
        - existing_codes: dict[group_code] → set of existing codes
    """
    group_codes = list(groups_config.keys())

    if not group_codes:
        return {}, {}

    placeholders = ','.join('?' * len(group_codes))
    query = f"""
        SELECT
            Code,
            SupplierProductCode,
            DescnPart1,
            DescnPart2,
            DescnPart3,
            Description,
            Active,
            Warning,
            PriceGridCode,
            CostGridCode,
            DiscountGroupCode,
            inventory_group_code,
            PkId
        FROM inventory_items
        WHERE inventory_group_code IN ({placeholders})
    """

    rows = db.execute_query(query, group_codes).fetchall()

    # Convert to DataFrame
    if rows:
        df_all = pd.DataFrame([dict(r) for r in rows], dtype=str).fillna("")
    else:
        df_all = pd.DataFrame()

    # Add matching key
    if not df_all.empty:
        df_all["_key"] = df_all.apply(
            lambda r: _build_desc_key(r["DescnPart1"], r["DescnPart2"], r["DescnPart3"]),
            axis=1
        )

    # Split by group
    inv_by_group = {}
    existing_codes = {}

    for group_code in group_codes:
        if df_all.empty:
            inv_by_group[group_code] = pd.DataFrame()
            existing_codes[group_code] = set()
        else:
            group_df = df_all[df_all["inventory_group_code"] == group_code].copy()
            inv_by_group[group_code] = group_df
            existing_codes[group_code] = set(group_df["Code"].tolist())

    return inv_by_group, existing_codes


def load_existing_buz_pricing(db, groups_config: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    Load latest pricing for all inventory codes (retail groups only).

    Returns dict: inventory_code → {sell_price, cost_price, markup, date_from}
    """
    query = """
        SELECT
            InventoryCode,
            SellSQM,
            CostSQM,
            DateFrom,
            CustomerPriceGroupCode
        FROM pricing_data
    """

    rows = db.execute_query(query).fetchall()

    if not rows:
        return {}

    df = pd.DataFrame([dict(r) for r in rows], dtype=str).fillna("")

    # Filter out rows where CustomerPriceGroupCode is not empty
    # (these are price group overrides, not base pricing)
    df = df[df["CustomerPriceGroupCode"].str.strip() == ""].copy()

    # Parse dates and sort to get latest per code
    df["_date_dt"] = pd.to_datetime(df["DateFrom"], errors="coerce", dayfirst=True)
    df = df.sort_values(["InventoryCode", "_date_dt"], kind="mergesort")

    # Get last row per code
    latest = df.groupby("InventoryCode", as_index=False).tail(1)

    pricing_map = {}
    for _, row in latest.iterrows():
        code = _norm(row["InventoryCode"])
        sell = _q2(row["SellSQM"])
        cost = _q2(row["CostSQM"])

        # Calculate markup (sell / cost)
        markup = None
        if cost > 0:
            markup = sell / cost

        pricing_map[code] = {
            "sell_price": sell,
            "cost_price": cost,
            "markup": markup,
            "date_from": _norm(row["DateFrom"])
        }

    return pricing_map


# ========== Matching & Change Detection ==========

def compute_changes(
    fabrics_by_group: Dict[str, pd.DataFrame],
    inv_by_group: Dict[str, pd.DataFrame],
    existing_codes: Dict[str, set],
    groups_config: Dict[str, Dict[str, Any]],
    pricing_map: Dict[str, Dict[str, Any]],
    filtered_by_material: Dict[str, set],
    price_grids: Dict[Tuple[str, str], str],
    progress=None
) -> Tuple[Dict[str, List[Dict]], Dict[str, List[Dict]], List[Dict], Dict[str, Dict]]:
    """
    Compute ADD, EDIT, and DEPRECATE operations.

    Returns:
        - items_changes: dict[group_code] → list of item rows to write
        - pricing_changes: dict[group_code] → list of pricing rows to write (retail only)
        - change_log: list of change descriptions for UI
        - markup_info: dict[group_code] → {markup_used, markup_source, existing_avg_markup}
    """
    def _p(msg: str, pct: Optional[int] = None):
        if callable(progress):
            try:
                progress(msg, pct)
            except Exception:
                pass

    items_changes = defaultdict(list)
    pricing_changes = defaultdict(list)
    change_log = []
    markup_info = {}

    _p("Computing changes for each group...", 35)

    for group_code, group_cfg in groups_config.items():
        # Don't log per-group to avoid resetting progress bar
        fabrics_df = fabrics_by_group.get(group_code, pd.DataFrame())
        inv_df = inv_by_group.get(group_code, pd.DataFrame())
        codes_set = existing_codes.get(group_code, set())

        # Get config for this group
        prefix = group_cfg.get("description_prefix", "")
        price_grid_code = group_cfg.get("price_grid_code")
        cost_grid_code = group_cfg.get("cost_grid_code")
        discount_code = group_cfg.get("discount_group_code", "")
        markup_override = group_cfg.get("markup_override")
        wastage_pct = group_cfg.get("wastage_pct")  # e.g., 0.20 for 20%
        is_wholesale = _is_wholesale_group(group_code)

        # Debug logging for wastage
        if wastage_pct:
            logger.info(f"Group {group_code}: Using wastage {wastage_pct} ({float(wastage_pct)*100:.1f}%)")
        else:
            logger.debug(f"Group {group_code}: No wastage configured")

        # Calculate average markup from existing items in this group (retail only)
        avg_markup = None
        markup_used = None
        markup_source = None

        if not is_wholesale and not inv_df.empty:
            # Get markups for all existing items in this group
            # Only include active items with non-zero markup
            markups = []
            for _, row in inv_df.iterrows():
                code = _norm(row["Code"])
                is_active = _norm(row.get("Active", "")).upper() in ("TRUE", "YES", "1")

                if is_active and code in pricing_map:
                    item_markup = pricing_map[code].get("markup")
                    if item_markup and item_markup > 0:
                        markups.append(item_markup)

            if markups:
                avg_markup = sum(markups) / len(markups)

            # Determine which markup to use
            if markup_override:
                markup_used = markup_override
                markup_source = "override"
            elif avg_markup:
                markup_used = avg_markup
                markup_source = "calculated"
            else:
                markup_used = Decimal("2.0")  # Default 2x markup
                markup_source = "default"

            markup_info[group_code] = {
                "markup_used": markup_used,
                "markup_source": markup_source,
                "existing_avg_markup": avg_markup,
                "markup_override": markup_override
            }

        # Build sets of keys
        fabric_keys = set(fabrics_df["_key"].tolist()) if not fabrics_df.empty else set()
        inv_keys = set(inv_df["_key"].tolist()) if not inv_df.empty else set()

        # ADD: In fabrics, not in inventory
        add_keys = fabric_keys - inv_keys
        for key in add_keys:
            fabric_row = fabrics_df[fabrics_df["_key"] == key].iloc[0]

            fd1 = _norm(fabric_row["FD1"])
            fd2 = _norm(fabric_row["FD2"])
            fd3 = _norm(fabric_row["FD3"])
            unleashed_code = _norm(fabric_row["Unleashed Code"])
            price_value = _norm(fabric_row["Price"])

            # Generate new code
            new_code = _next_code_for_group(list(codes_set), group_code)
            codes_set.add(new_code)

            description = _build_description(prefix, fd1, fd2, fd3)

            # For wholesale groups, look up grid codes from Price Grids tab
            item_price_grid = price_grid_code
            item_cost_grid = cost_grid_code
            if is_wholesale:
                price_category = _norm(fabric_row.get("Price", ""))  # For wholesale, "Price" column contains Price Category
                lookup_key = (group_code, price_category)
                if lookup_key in price_grids:
                    grid_code = price_grids[lookup_key]
                    item_price_grid = grid_code
                    item_cost_grid = grid_code + "C"
                    logger.debug(f"Group {group_code} ADD {new_code}: Looked up grid codes for category '{price_category}' → Price: {item_price_grid}, Cost: {item_cost_grid}")
                else:
                    # No mapping found - leave grid codes blank (e.g., verishades don't use grids)
                    item_price_grid = ""
                    item_cost_grid = ""
                    logger.debug(f"Group {group_code} ADD {new_code}: No grid code mapping for category '{price_category}' - leaving blank")

            item_row = {
                "PkId": "",
                "Code": new_code,
                "Description": description,
                "DescnPart1 (Material)": fd1,
                "DescnPart2 (Material Types)": fd2,
                "DescnPart3 (Colour)": fd3,
                "Price Grid Code": item_price_grid or "",
                "Cost Grid Code": item_cost_grid or "",
                "Discount Group Code": discount_code,
                "Tax Rate": TAX_RATE,
                "Supplier": SUPPLIER_NAME,
                "Supplier Product Code": unleashed_code,
                "Active": "TRUE",
                "Warning": "",
                "Operation": "A"
            }

            items_changes[group_code].append(item_row)

            change_log.append({
                "Group": group_code,
                "Operation": "A",
                "Code": new_code,
                "Description": description,
                "Reason": "New fabric"
            })

            # Pricing for retail groups only
            if not is_wholesale and markup_used:
                cost_sqm = _q2(price_value)

                # Convert LM to SQM for vertical blinds
                price_type = group_cfg.get("price_type", "SQM")
                if price_type == "LM":
                    width_str = _norm(fabric_row.get("Width", ""))
                    price_before = cost_sqm
                    try:
                        width_mm = float(width_str) if width_str else 0
                        if width_mm > 0:
                            cost_sqm = cost_sqm / (Decimal(str(width_mm)) / Decimal("1000"))
                            logger.info(f"Group {group_code} ADD {new_code}: LM→SQM conversion: ${price_before}/LM ÷ ({width_mm}mm/1000) = ${cost_sqm:.2f}/SQM")
                        else:
                            logger.warning(f"Missing/invalid width for vertical blind {new_code}: '{width_str}'")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Failed to parse width for vertical blind {new_code}: '{width_str}' - {e}")

                # Apply wastage adjustment if configured
                if wastage_pct:
                    cost_sqm = cost_sqm * (Decimal("1") + wastage_pct)
                sell_sqm = cost_sqm * markup_used

                # Round to 2 decimals
                cost_sqm = _q2(cost_sqm)
                sell_sqm = _q2(sell_sqm)

                pricing_changes[group_code].append({
                    "PkId": "",
                    "Inventory Code": new_code,
                    "Description": description,
                    "Date From": "01/01/2020",  # New fabric date
                    "CostSQM": f"{cost_sqm:.2f}",
                    "SellSQM": f"{sell_sqm:.2f}",
                    "Operation": "A"
                })

        # EDIT: In both, check for changes
        shared_keys = fabric_keys & inv_keys
        for key in shared_keys:
            fabric_row = fabrics_df[fabrics_df["_key"] == key].iloc[0]
            inv_row = inv_df[inv_df["_key"] == key].iloc[0]

            fd1 = _norm(fabric_row["FD1"])
            fd2 = _norm(fabric_row["FD2"])
            fd3 = _norm(fabric_row["FD3"])
            unleashed_code = _norm(fabric_row["Unleashed Code"])
            price_value = _norm(fabric_row["Price"])

            existing_code = _norm(inv_row["Code"])
            existing_supp_code = _norm(inv_row["SupplierProductCode"])
            existing_active = _norm(inv_row["Active"]).upper() in ("TRUE", "YES", "1")
            existing_warning = _norm(inv_row["Warning"])

            reasons = []
            needs_edit = False

            # Check if supplier product code changed (case-insensitive comparison)
            if unleashed_code.lower() != existing_supp_code.lower():
                needs_edit = True
                reasons.append(f"Supplier code changed: {existing_supp_code} → {unleashed_code}")

            # Check if inactive (and not already deprecated)
            if not existing_active and existing_warning != DEPRECATED_WARNING:
                needs_edit = True
                reasons.append("Reactivated")

            if needs_edit:
                description = _build_description(prefix, fd1, fd2, fd3)

                # For wholesale groups, look up grid codes from Price Grids tab
                item_price_grid = price_grid_code or _norm(inv_row["PriceGridCode"])
                item_cost_grid = cost_grid_code or _norm(inv_row["CostGridCode"])
                if is_wholesale:
                    price_category = _norm(fabric_row.get("Price", ""))  # For wholesale, "Price" column contains Price Category
                    lookup_key = (group_code, price_category)
                    if lookup_key in price_grids:
                        grid_code = price_grids[lookup_key]
                        item_price_grid = grid_code
                        item_cost_grid = grid_code + "C"
                        logger.debug(f"Group {group_code} EDIT {existing_code}: Looked up grid codes for category '{price_category}' → Price: {item_price_grid}, Cost: {item_cost_grid}")
                    else:
                        # No mapping found - blank out grid codes (e.g., verishades, or clearing old grid codes during reactivation)
                        item_price_grid = ""
                        item_cost_grid = ""
                        logger.debug(f"Group {group_code} EDIT {existing_code}: No grid code mapping for category '{price_category}' - blanking out")

                item_row = {
                    "PkId": _norm(inv_row["PkId"]),
                    "Code": existing_code,
                    "Description": description,
                    "DescnPart1 (Material)": fd1,
                    "DescnPart2 (Material Types)": fd2,
                    "DescnPart3 (Colour)": fd3,
                    "Price Grid Code": item_price_grid or "",
                    "Cost Grid Code": item_cost_grid or "",
                    "Discount Group Code": discount_code or _norm(inv_row["DiscountGroupCode"]),
                    "Tax Rate": TAX_RATE,
                    "Supplier": SUPPLIER_NAME,
                    "Supplier Product Code": unleashed_code,
                    "Active": "TRUE",
                    "Warning": "",
                    "Operation": "E"
                }

                items_changes[group_code].append(item_row)

                change_log.append({
                    "Group": group_code,
                    "Operation": "E",
                    "Code": existing_code,
                    "Description": description,
                    "Reason": "; ".join(reasons)
                })

            # Check pricing (retail only)
            if not is_wholesale and markup_used:
                new_cost = _q2(price_value)

                # Convert LM to SQM for vertical blinds
                price_type = group_cfg.get("price_type", "SQM")
                if price_type == "LM":
                    width_str = _norm(fabric_row.get("Width", ""))
                    price_before = new_cost
                    try:
                        width_mm = float(width_str) if width_str else 0
                        if width_mm > 0:
                            new_cost = new_cost / (Decimal(str(width_mm)) / Decimal("1000"))
                            logger.info(f"Group {group_code} EDIT {existing_code}: LM→SQM conversion: ${price_before}/LM ÷ ({width_mm}mm/1000) = ${new_cost:.2f}/SQM")
                        else:
                            logger.warning(f"Missing/invalid width for vertical blind {existing_code}: '{width_str}'")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Failed to parse width for vertical blind {existing_code}: '{width_str}' - {e}")

                # Apply wastage adjustment if configured
                if wastage_pct:
                    new_cost = new_cost * (Decimal("1") + wastage_pct)
                new_sell = new_cost * markup_used

                # Round to 2 decimals for comparison
                new_cost = _q2(new_cost)
                new_sell = _q2(new_sell)

                existing_pricing = pricing_map.get(existing_code, {})
                existing_cost = _q2(existing_pricing.get("cost_price", Decimal("0.00")))
                existing_sell = _q2(existing_pricing.get("sell_price", Decimal("0.00")))

                # Check if cost or sell price changed by more than 1 cent (tolerance for rounding)
                cost_diff = abs(new_cost - existing_cost)
                sell_diff = abs(new_sell - existing_sell)
                if cost_diff > Decimal("0.01") or sell_diff > Decimal("0.01"):
                    description = _build_description(prefix, fd1, fd2, fd3)
                    pricing_changes[group_code].append({
                        "PkId": "",
                        "Inventory Code": existing_code,
                        "Description": description,
                        "Date From": _tomorrow_ddmmyyyy(),
                        "CostSQM": f"{new_cost:.2f}",
                        "SellSQM": f"{new_sell:.2f}",
                        "Operation": "A"
                    })

                    change_log.append({
                        "Group": group_code,
                        "Operation": "P",
                        "Code": existing_code,
                        "Description": description,
                        "Reason": f"Price changed: Cost {existing_cost:.2f} → {new_cost:.2f}, Sell {existing_sell:.2f} → {new_sell:.2f}"
                    })

        # DEPRECATE: In inventory (active, not already deprecated), not in fabrics
        deprecate_keys = inv_keys - fabric_keys
        for key in deprecate_keys:
            inv_row = inv_df[inv_df["_key"] == key].iloc[0]

            existing_code = _norm(inv_row["Code"])
            existing_active = _norm(inv_row["Active"]).upper() in ("TRUE", "YES", "1")
            existing_warning = _norm(inv_row["Warning"])

            # Only deprecate if active and not already deprecated
            if existing_active and existing_warning != DEPRECATED_WARNING:
                fd1 = _norm(inv_row["DescnPart1"])
                fd2 = _norm(inv_row["DescnPart2"])
                fd3 = _norm(inv_row["DescnPart3"])
                description = _build_description(prefix, fd1, fd2, fd3)

                # Special case: ROMNBQ - don't deprecate items where description starts with "1 "
                if group_code == "ROMNBQ" and description.startswith("1 "):
                    continue

                item_row = {
                    "PkId": _norm(inv_row["PkId"]),
                    "Code": existing_code,
                    "Description": description,
                    "DescnPart1 (Material)": fd1,
                    "DescnPart2 (Material Types)": fd2,
                    "DescnPart3 (Colour)": fd3,
                    "Price Grid Code": _norm(inv_row["PriceGridCode"]),
                    "Cost Grid Code": _norm(inv_row["CostGridCode"]),
                    "Discount Group Code": _norm(inv_row["DiscountGroupCode"]),
                    "Tax Rate": TAX_RATE,
                    "Supplier": SUPPLIER_NAME,
                    "Supplier Product Code": _norm(inv_row["SupplierProductCode"]),
                    "Active": "TRUE",  # Keep active
                    "Warning": DEPRECATED_WARNING,
                    "Operation": "E"
                }

                items_changes[group_code].append(item_row)

                # Determine deprecation reason
                filtered_keys = filtered_by_material.get(group_code, set())
                if key in filtered_keys:
                    reason = f"Material type '{fd2}' not allowed for this product group"
                else:
                    reason = "Not in Google Sheet"

                change_log.append({
                    "Group": group_code,
                    "Operation": "D",
                    "Code": existing_code,
                    "Description": description,
                    "Reason": reason
                })

    _p("Change computation complete", 60)
    return dict(items_changes), dict(pricing_changes), change_log, markup_info


# ========== Excel Generation ==========

def _format_worksheet(ws, headers: List[str]):
    """
    Format worksheet by hiding empty columns and autofitting columns with data.

    Args:
        ws: openpyxl worksheet
        headers: List of header column names
    """
    from openpyxl.utils import get_column_letter

    # Determine which columns have data (skip first row which is blank for items file)
    start_row = 3 if ws.max_row > 2 and not any(ws[1]) else 2

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)

        # Check if column has any non-empty data
        has_data = False
        for row_idx in range(start_row, ws.max_row + 1):
            cell_value = ws[f"{col_letter}{row_idx}"].value
            if cell_value is not None and str(cell_value).strip() != "":
                has_data = True
                break

        if has_data:
            # Autofit column width based on content
            max_length = 0
            for row_idx in range(start_row - 1, ws.max_row + 1):  # Include header
                cell = ws[f"{col_letter}{row_idx}"]
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Set column width (add a bit of padding, max 50)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        else:
            # Hide empty columns
            ws.column_dimensions[col_letter].hidden = True


def generate_workbooks_in_memory(
    items_changes: Dict[str, List[Dict]],
    pricing_changes: Dict[str, List[Dict]],
    headers_cfg: Dict[str, List[Dict]]
):
    """
    Generate items and pricing upload workbooks in memory.

    Returns (items_stream, pricing_stream) as BytesIO objects
    """
    from io import BytesIO
    from openpyxl.utils import get_column_letter

    # Get headers from config
    items_headers = [h["spreadsheet_column"] for h in headers_cfg["buz_inventory_item_file"]]
    pricing_headers = [h["spreadsheet_column"] for h in headers_cfg["buz_pricing_file"]]

    # ===== Items Workbook =====
    items_wb = Workbook()
    items_wb.remove(items_wb.active)

    for group_code, rows in items_changes.items():
        if not rows:
            continue

        ws = items_wb.create_sheet(title=group_code)
        ws.append([])  # Row 1 blank
        ws.append(items_headers + [""])  # Row 2 headers + trailing blank

        for row_dict in rows:
            row_values = []
            for header in items_headers:
                row_values.append(row_dict.get(header, ""))
            ws.append(row_values + [""])  # Trailing blank cell

        # Format sheet: autofit columns and hide empty ones
        _format_worksheet(ws, items_headers)

    # Ensure at least one sheet exists (openpyxl requires this)
    if len(items_wb.sheetnames) == 0:
        ws = items_wb.create_sheet(title="No Changes")
        ws.append(["No changes to upload"])

    # Save to BytesIO
    items_stream = BytesIO()
    items_wb.save(items_stream)
    items_stream.seek(0)
    items_wb.close()

    # ===== Pricing Workbook =====
    pricing_wb = Workbook()
    pricing_wb.remove(pricing_wb.active)

    for group_code, rows in pricing_changes.items():
        if not rows:
            continue

        ws = pricing_wb.create_sheet(title=group_code)
        ws.append(pricing_headers)  # Row 1 headers (pricing file format)

        for row_dict in rows:
            row_values = []
            for header in pricing_headers:
                if header == "Operation":
                    row_values.append("A")
                elif header == "PkId":
                    row_values.append("")
                else:
                    row_values.append(row_dict.get(header, ""))
            ws.append(row_values)

        # Format sheet: autofit columns and hide empty ones
        _format_worksheet(ws, pricing_headers)

    # Ensure at least one sheet exists (openpyxl requires this)
    if len(pricing_wb.sheetnames) == 0:
        ws = pricing_wb.create_sheet(title="No Changes")
        ws.append(["No changes to upload"])

    # Save to BytesIO
    pricing_stream = BytesIO()
    pricing_wb.save(pricing_stream)
    pricing_stream.seek(0)
    pricing_wb.close()

    return items_stream, pricing_stream


def generate_workbooks(
    items_changes: Dict[str, List[Dict]],
    pricing_changes: Dict[str, List[Dict]],
    headers_cfg: Dict[str, List[Dict]],
    output_dir: str,
    progress=None
) -> Tuple[str, str]:
    """
    DEPRECATED: Generate items and pricing upload workbooks to disk.
    Use generate_workbooks_in_memory() instead.

    Returns (items_path, pricing_path)
    """
    def _p(msg: str, pct: Optional[int] = None):
        if callable(progress):
            try:
                progress(msg, pct)
            except Exception:
                pass

    os.makedirs(output_dir, exist_ok=True)

    # Get headers from config
    items_headers = [h["spreadsheet_column"] for h in headers_cfg["buz_inventory_item_file"]]
    pricing_headers = [h["spreadsheet_column"] for h in headers_cfg["buz_pricing_file"]]

    # ===== Items Workbook =====
    _p("Generating items workbook...", 65)
    items_wb = Workbook()
    items_wb.remove(items_wb.active)

    for group_code, rows in items_changes.items():
        if not rows:
            continue

        ws = items_wb.create_sheet(title=group_code)
        ws.append([])  # Row 1 blank
        ws.append(items_headers + [""])  # Row 2 headers + trailing blank

        for row_dict in rows:
            row_values = []
            for header in items_headers:
                row_values.append(row_dict.get(header, ""))
            ws.append(row_values + [""])  # Trailing blank cell

        # Format sheet: autofit columns and hide empty ones
        _format_worksheet(ws, items_headers)

    items_path = os.path.join(output_dir, "blinds_awnings_items_upload.xlsx")
    has_items = save_workbook_gracefully(items_wb, items_path)

    if not has_items:
        _p("No item changes", 75)
    else:
        _p("Items workbook generated", 75)

    # ===== Pricing Workbook =====
    _p("Generating pricing workbook...", 80)
    pricing_wb = Workbook()
    pricing_wb.remove(pricing_wb.active)

    for group_code, rows in pricing_changes.items():
        if not rows:
            continue

        ws = pricing_wb.create_sheet(title=group_code)
        ws.append(pricing_headers)  # Row 1 headers (pricing file format)

        for row_dict in rows:
            row_values = []
            for header in pricing_headers:
                if header == "Operation":
                    row_values.append("A")
                elif header == "PkId":
                    row_values.append("")
                else:
                    row_values.append(row_dict.get(header, ""))
            ws.append(row_values)

        # Format sheet: autofit columns and hide empty ones
        _format_worksheet(ws, pricing_headers)

    pricing_path = os.path.join(output_dir, "blinds_awnings_pricing_upload.xlsx")
    has_pricing = save_workbook_gracefully(pricing_wb, pricing_path)

    if not has_pricing:
        _p("No pricing changes", 90)
    else:
        _p("Pricing workbook generated", 90)

    return items_path, pricing_path


# ========== Main Orchestrator ==========

def sync_blinds_awnings_fabrics(
    db,
    config: Dict[str, Any],
    sheets_service,
    output_dir: str = "uploads",
    progress=None
) -> Dict[str, Any]:
    """
    Main orchestrator for blinds/awnings fabric sync.

    Args:
        db: DatabaseManager instance
        config: Full config dict (from config.json)
        sheets_service: GoogleSheetsService instance
        output_dir: Output directory for generated files
        progress: Optional progress callback(msg, pct)

    Returns:
        Dict with:
            - items_file: path to items workbook
            - pricing_file: path to pricing workbook
            - summary: dict with counts (adds, edits, deprecates, pricing)
            - change_log: list of change descriptions
    """
    def _p(msg: str, pct: Optional[int] = None):
        if callable(progress):
            try:
                progress(msg, pct)
            except Exception:
                pass

    _p("Starting blinds/awnings fabric sync...", 1)

    # Extract config
    material_restrictions = config.get("material_restrictions_by_group", {})
    headers_cfg = config.get("headers", {})
    sheets_cfg = config["spreadsheets"]["blinds_awnings_sync"]

    spreadsheet_id = sheets_cfg["id"]
    retail_tab = sheets_cfg["retail_tab"]
    wholesale_tab = sheets_cfg["wholesale_tab"]
    buz_template_tab = sheets_cfg["buz_template_tab"]

    # Load groups configuration from Google Sheets
    groups_config = load_groups_config_from_sheet(
        sheets_service,
        spreadsheet_id,
        buz_template_tab,
        progress=_p
    )

    # Load price grids lookup for wholesale items
    price_grids_tab = sheets_cfg.get("price_grids_tab", "Price Grids")
    price_grids = load_price_grids_lookup(
        sheets_service,
        spreadsheet_id,
        price_grids_tab,
        progress=_p
    )

    # Load fabric data from Google Sheets
    fabrics_by_group, filtered_by_material = load_fabric_data_from_sheets(
        sheets_service,
        spreadsheet_id,
        retail_tab,
        wholesale_tab,
        groups_config,
        material_restrictions,
        progress=_p
    )

    # Load existing Buz data
    _p("Loading existing Buz inventory...", 32)
    inv_by_group, existing_codes = load_existing_buz_inventory(db, groups_config)

    _p("Loading existing Buz pricing...", 34)
    pricing_map = load_existing_buz_pricing(db, groups_config)

    # Compute changes
    items_changes, pricing_changes, change_log, markup_info = compute_changes(
        fabrics_by_group,
        inv_by_group,
        existing_codes,
        groups_config,
        pricing_map,
        filtered_by_material,
        price_grids,
        progress=_p
    )

    # Note: We don't generate workbooks here anymore - they'll be generated on-demand when downloaded
    _p("Changes computed, ready for download", 90)

    # Compute summary - overall and per-group
    total_adds = sum(len([r for r in rows if r.get("Operation") == "A"]) for rows in items_changes.values())
    total_edits = sum(len([r for r in rows if r.get("Operation") == "E"]) for rows in items_changes.values())
    total_deprecates = len([c for c in change_log if c.get("Operation") == "D"])
    total_pricing = sum(len(rows) for rows in pricing_changes.values())

    # Per-group breakdown
    groups_summary = {}
    for group_code in groups_config.keys():
        group_items = items_changes.get(group_code, [])
        adds = len([r for r in group_items if r.get("Operation") == "A"])
        edits = len([r for r in group_items if r.get("Operation") == "E"])
        deprecates = len([c for c in change_log if c.get("Group") == group_code and c.get("Operation") == "D"])
        pricing = len(pricing_changes.get(group_code, []))

        # Only include groups with changes or markup info
        if adds > 0 or edits > 0 or deprecates > 0 or pricing > 0 or group_code in markup_info:
            groups_summary[group_code] = {
                "A": adds,
                "E": edits,
                "D": deprecates,
                "P": pricing
            }

            # Add markup info if available
            if group_code in markup_info:
                markup_data = markup_info[group_code].copy()
                # Convert Decimal objects to float for JSON serialization
                if markup_data.get("markup_used"):
                    markup_data["markup_used"] = float(markup_data["markup_used"])
                if markup_data.get("existing_avg_markup"):
                    markup_data["existing_avg_markup"] = float(markup_data["existing_avg_markup"])
                if markup_data.get("markup_override"):
                    markup_data["markup_override"] = float(markup_data["markup_override"])
                groups_summary[group_code]["markup"] = markup_data

    summary = {
        "A": total_adds,
        "E": total_edits,
        "D": total_deprecates,
        "P": total_pricing,
        "by_group": groups_summary
    }

    _p("Sync complete!", 100)

    logger.info(f"Blinds/Awnings sync complete: A={total_adds}, E={total_edits}, D={total_deprecates}, P={total_pricing}")

    return {
        "summary": summary,
        "change_log": change_log,
        "items_changes": items_changes,
        "pricing_changes": pricing_changes,
        "headers_cfg": headers_cfg  # Include headers for on-demand generation
    }


def apply_changes_to_database(items_changes: Dict[str, List[Dict]], pricing_changes: Dict[str, List[Dict]], db, progress=None) -> Dict[str, int]:
    """
    Apply the computed changes to the database.

    Args:
        items_changes: Dict of group_code -> list of item change dicts
        pricing_changes: Dict of group_code -> list of pricing change dicts
        db: DatabaseManager instance
        progress: Optional progress callback function(message, pct)

    Returns:
        Dict with counts: {"items_added": N, "items_updated": N, "pricing_added": N}
    """
    def _p(msg, pct=None):
        if progress:
            progress(msg, pct)

    _p("Applying changes to database...", 0)

    items_added = 0
    items_updated = 0
    pricing_added = 0

    # Apply item changes
    total_items = sum(len(rows) for rows in items_changes.values())
    processed = 0

    for group_code, rows in items_changes.items():
        for row in rows:
            operation = row.get("Operation")
            code = row.get("Code")

            if operation == "A":
                # ADD: Insert new item
                db.execute_query("""
                    INSERT INTO inventory_items (
                        inventory_group_code, Code, Description,
                        DescnPart1, DescnPart2, DescnPart3,
                        PriceGridCode, CostGridCode, DiscountGroupCode,
                        SupplierProductCode, Active, Warning,
                        TaxRate, Supplier
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    group_code,
                    code,
                    row.get("Description", ""),
                    row.get("DescnPart1 (Material)", ""),
                    row.get("DescnPart2 (Material Types)", ""),
                    row.get("DescnPart3 (Colour)", ""),
                    row.get("Price Grid Code", ""),
                    row.get("Cost Grid Code", ""),
                    row.get("Discount Group Code", ""),
                    row.get("Supplier Product Code", ""),
                    row.get("Active", "TRUE"),
                    row.get("Warning", ""),
                    row.get("Tax Rate", TAX_RATE),
                    row.get("Supplier", SUPPLIER_NAME)
                ))
                items_added += 1

            elif operation == "E":
                # EDIT: Update existing item
                pk_id = row.get("PkId", "")

                if pk_id:
                    # Update by PkId if available
                    db.execute_query("""
                        UPDATE inventory_items
                        SET Description = ?,
                            DescnPart1 = ?,
                            DescnPart2 = ?,
                            DescnPart3 = ?,
                            PriceGridCode = ?,
                            CostGridCode = ?,
                            DiscountGroupCode = ?,
                            SupplierProductCode = ?,
                            Active = ?,
                            Warning = ?,
                            TaxRate = ?,
                            Supplier = ?
                        WHERE PkId = ?
                    """, (
                        row.get("Description", ""),
                        row.get("DescnPart1 (Material)", ""),
                        row.get("DescnPart2 (Material Types)", ""),
                        row.get("DescnPart3 (Colour)", ""),
                        row.get("Price Grid Code", ""),
                        row.get("Cost Grid Code", ""),
                        row.get("Discount Group Code", ""),
                        row.get("Supplier Product Code", ""),
                        row.get("Active", "TRUE"),
                        row.get("Warning", ""),
                        row.get("Tax Rate", TAX_RATE),
                        row.get("Supplier", SUPPLIER_NAME),
                        pk_id
                    ))
                else:
                    # Update by Code if no PkId
                    db.execute_query("""
                        UPDATE inventory_items
                        SET Description = ?,
                            DescnPart1 = ?,
                            DescnPart2 = ?,
                            DescnPart3 = ?,
                            PriceGridCode = ?,
                            CostGridCode = ?,
                            DiscountGroupCode = ?,
                            SupplierProductCode = ?,
                            Active = ?,
                            Warning = ?,
                            TaxRate = ?,
                            Supplier = ?
                        WHERE Code = ?
                    """, (
                        row.get("Description", ""),
                        row.get("DescnPart1 (Material)", ""),
                        row.get("DescnPart2 (Material Types)", ""),
                        row.get("DescnPart3 (Colour)", ""),
                        row.get("Price Grid Code", ""),
                        row.get("Cost Grid Code", ""),
                        row.get("Discount Group Code", ""),
                        row.get("Supplier Product Code", ""),
                        row.get("Active", "TRUE"),
                        row.get("Warning", ""),
                        row.get("Tax Rate", TAX_RATE),
                        row.get("Supplier", SUPPLIER_NAME),
                        code
                    ))
                items_updated += 1

            processed += 1
            if total_items > 0:
                _p(f"Processing items: {processed}/{total_items}", int(50 * processed / total_items))

    # Apply pricing changes
    total_pricing = sum(len(rows) for rows in pricing_changes.values())
    processed = 0

    for group_code, rows in pricing_changes.items():
        for row in rows:
            inventory_code = row.get("Inventory Code")

            # Delete old pricing records for this inventory code
            db.execute_query("""
                DELETE FROM pricing_data
                WHERE InventoryCode = ?
            """, (inventory_code,))

            # Insert new pricing record
            db.execute_query("""
                INSERT INTO pricing_data (
                    inventory_group_code, InventoryCode, Description,
                    DateFrom, SellSQM, CostSQM
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                group_code,
                inventory_code,
                row.get("Description", ""),
                row.get("Date From", ""),
                _q2(row.get("SellSQM", "0.00")),
                _q2(row.get("CostSQM", "0.00"))
            ))
            pricing_added += 1

            processed += 1
            if total_pricing > 0:
                _p(f"Processing pricing: {processed}/{total_pricing}", 50 + int(50 * processed / total_pricing))

    _p("Database update complete!", 100)

    logger.info(f"Applied changes to database: items_added={items_added}, items_updated={items_updated}, pricing_added={pricing_added}")

    return {
        "items_added": items_added,
        "items_updated": items_updated,
        "pricing_added": pricing_added
    }
