from datetime import datetime
from openpyxl.utils.exceptions import InvalidFileException
from io import BytesIO


def _sheet_has_content(ws) -> bool:
    """Heuristic: any non-A1 dimension, a non-empty A1, or any objects."""
    try:
        dim = ws.calculate_dimension()
    except Exception:
        dim = None

    if dim and dim != "A1":
        return True

    if getattr(ws, "A1", None) is not None and ws["A1"].value not in (None, ""):
        return True

    if getattr(ws, "_tables", None) and ws._tables:
        return True
    if getattr(ws, "_charts", None) and ws._charts:
        return True
    if getattr(ws, "_images", None) and ws._images:
        return True
    if getattr(ws, "_comments", None) and ws._comments:
        return True
    if getattr(ws, "_pivots", None) and ws._pivots:
        return True
    if getattr(ws, "_drawing", None) and ws._drawing and (ws._drawing.charts or ws._drawing.images):
        return True

    return False


def _unique_title(wb, base: str) -> str:
    """Ensure sheet title is unique in this workbook."""
    title = base
    i = 2
    existing = {ws.title for ws in wb.worksheets}
    while title in existing:
        title = f"{base} ({i})"
        i += 1
    return title


def prune_empty_sheets_with_placeholder(
    wb,
    placeholder_title: str = "No Data",
    placeholder_message: str | None = None,
) -> bool:
    """
    Remove truly empty worksheets. If all are empty, keep exactly one placeholder
    sheet so saving never fails. Returns True if any non-empty sheet remains.
    """
    empty = [ws for ws in wb.worksheets if not _sheet_has_content(ws)]
    non_empty = [ws for ws in wb.worksheets if ws not in empty]

    # Remove empties only if at least one non-empty remains.
    if non_empty:
        for ws in empty:
            wb.remove(ws)
        return True  # we still have real content

    # All sheets were empty -> keep one friendly placeholder
    if not wb.worksheets:
        # Extremely rare, but guard anyway.
        ws = wb.create_sheet(_unique_title(wb, placeholder_title))
    else:
        # Keep the first empty sheet, rename it.
        ws = wb.worksheets[0]
        ws.title = _unique_title(wb, placeholder_title)

    if placeholder_message is None:
        placeholder_message = (
            "No data to export.\n"
            f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
    ws["A1"].value = placeholder_message
    return False  # no real data, just the placeholder


def save_workbook_gracefully(wb, filename: str) -> bool:
    """
    Prune empties; if nothing left, leave a single placeholder sheet so the
    file is still valid. Returns True if the saved file contains real data.
    """
    has_real_data = prune_empty_sheets_with_placeholder(wb)
    try:
        wb.save(filename)
    except InvalidFileException as exc:
        # Shouldn't happen, but surface a clearer message if it does.
        raise InvalidFileException(f"Failed to save workbook '{filename}': {exc}") from exc
    return has_real_data


def build_excel_stream(wb):
    """Return (has_real_data, BytesIO) ready for send_file()."""
    has_real_data = prune_empty_sheets_with_placeholder(wb)
    out = BytesIO()
    wb.save(out)
    wb.close()
    out.seek(0)
    return has_real_data, out
