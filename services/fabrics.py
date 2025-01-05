from services.data_processing import (get_all_fabrics, get_all_fabric_group_mappings, get_inventory_groups)
from services.database import DatabaseManager
from openpyxl import Workbook


def get_fabric_grid_data(db_manager: DatabaseManager):
    # Get all fabrics with concatenated descriptions
    fabrics = get_all_fabrics(db_manager=db_manager)

    # Get all inventory groups
    inventory_groups = get_inventory_groups(db_manager=db_manager)
    print("Inventory Groups:", inventory_groups)
    for group in inventory_groups:
        print("Available keys in group:", group.keys())  # This shows the actual column names

    # Get all mappings
    mappings = get_all_fabric_group_mappings(db_manager=db_manager)

    # Convert data to dictionaries for easy processing
    fabric_list = {fabric["id"]: fabric for fabric in fabrics}
    group_list = {group["group_code"]: group["group_description"] for group in inventory_groups}
    mapping_set = {(mapping["fabric_id"], mapping["inventory_group_code"]) for mapping in mappings}

    return fabric_list, group_list, mapping_set


def abbreviate(group_name):
    """
    Generate an abbreviation for a group name.

    Args:
        group_name (str): Full group name.

    Returns:
        str: Abbreviated group name.
    """
    words = group_name.split()
    if len(words) == 1:
        return group_name[:3].upper()  # Use the first 3 letters if it's a single word
    else:
        return ''.join(word[0].upper() for word in words)  # Use initials for multi-word names


def prepare_fabric_grid_data(fabric_list, group_list, mapping_set):
    """
    Prepares the data for the fabric grid template.

    Args:
        fabric_list (dict): Dictionary of fabrics with fabric_id as key.
        group_list (dict): Dictionary of groups with group_code as key.
        mapping_set (set): Set of tuples (fabric_id, group_code) representing mappings.

    Returns:
        dict: Prepared data including the grid and groups with abbreviations.
    """
    # Add abbreviations to the group list
    groups_with_abbrev = {
        group_code: {
            "name": group_name,
            "abbrev": abbreviate(group_name)  # Generate abbreviation
        }
        for group_code, group_name in group_list.items()
    }

    grid = []
    for fabric_id, fabric in fabric_list.items():
        # Concatenate descriptions
        concatenated_description = " ".join(
            filter(None, [fabric["description_1"], fabric["description_2"], fabric["description_3"]])
        )
        grid.append({
            "fabric_id": fabric_id,
            "fabric_description": concatenated_description,
            "description_1": fabric["description_1"],
            "description_2": fabric["description_2"],
            "description_3": fabric["description_3"],
            "fabric_code": fabric["supplier_code"],
            "groups": {
                group_code: (fabric_id, group_code) in mapping_set
                for group_code in group_list.keys()
            }
        })

    return {"grid": grid, "groups": groups_with_abbrev}


def process_fabric_mappings(mappings, db_manager):
    """
    Update the database based on the submitted mappings.
    """
    # Example: Convert mappings into usable format and update database
    for fabric_id, group_mappings in mappings.items():
        for group_code, is_checked in group_mappings.items():
            if is_checked == "on":
                # Add mapping to the database
                add_fabric_to_group(fabric_id, group_code, db_manager)
            else:
                # Remove mapping from the database
                remove_fabric_from_group(fabric_id, group_code, db_manager)


def add_fabric_to_group(db_manager: DatabaseManager, fabric_id, group_code):
    """
    Add a fabric-to-group mapping to the database with error handling.
    """
    try:
        db_manager.execute_query(
            "INSERT INTO fabric_group_mappings (fabric_id, inventory_group_code) VALUES (?, ?)",
            (fabric_id, group_code), True
        )
        print(f"Mapping added successfully: fabric_id={fabric_id}, group_code={group_code}")
    except Exception as e:
        print(f"Error adding mapping: fabric_id={fabric_id}, group_code={group_code}, error={e}")
        raise


def remove_fabric_from_group(db_manager, fabric_id, group_code):
    """
    Remove a fabric-to-group mapping from the database with error handling.
    """
    try:
        db_manager.execute_query(
            "DELETE FROM fabric_group_mappings WHERE fabric_id = ? AND inventory_group_code = ?",
            (fabric_id, group_code), True
        )
        print(f"Mapping removed successfully: fabric_id={fabric_id}, group_code={group_code}")
    except Exception as e:
        print(f"Error removing mapping: fabric_id={fabric_id}, group_code={group_code}, error={e}")
        raise


def get_fabric_by_id(fabric_id, db):
    query = "SELECT * FROM fabrics WHERE id = ?"
    return db.execute_query(query, (fabric_id,)).fetchone()


def add_new_fabric(fabric_data, db):
    query = """
        INSERT INTO fabrics (description_1, description_2, description_3, supplier_code)
        VALUES (?, ?, ?, ?)
    """
    cursor = db.execute_query(query, (
        fabric_data["description_1"],
        fabric_data["description_2"],
        fabric_data["description_3"],
        fabric_data["supplier_code"],
    ))
    return cursor.lastrowid


def get_fabric_mappings(fabric_id, db_manager):
    query = "SELECT * FROM fabric_group_mappings WHERE fabric_id = ?"
    return db_manager.execute_query(query, (fabric_id,)).fetchall()


def add_mapping(fabric_id, group_code, db):
    query = "INSERT INTO fabric_group_mappings (fabric_id, inventory_group_code) VALUES (?, ?)"
    db.execute_query(query, (fabric_id, group_code))


def update_fabric_in_db(fabric_id, fabric_data, db_manager):
    query = """
        UPDATE fabrics
        SET supplier_code = ?, description_1 = ?, description_2 = ?, description_3 = ?
        WHERE id = ?
    """
    db_manager.execute_query(query, (
        fabric_data["supplier_code"],
        fabric_data["description_1"],
        fabric_data["description_2"],
        fabric_data["description_3"],
        fabric_id,
    ), True)


def get_fabrics_and_mappings(db_manager):
    query = """
        SELECT 
            f.id AS fabric_id,
            f.description_1,
            f.description_2,
            f.description_3,
            f.supplier_code,
            fm.inventory_group_code
        FROM fabrics f
        INNER JOIN fabric_group_mappings fm ON f.id = fm.fabric_id
    """
    return db_manager.execute_query(query).fetchall()


def get_inventory_items(db_manager):
    """
    Get just the inventory group code and descriptions to compare changes with fabric master list
    """
    query = """
        SELECT 
            inventory_group_code, 
            DescnPart1, 
            DescnPart2, 
            DescnPart3 
        FROM inventory_items
    """
    return db_manager.execute_query(query).fetchall()


def compare_fabrics_to_inventory(fabrics, inventory_items):
    to_add = []
    to_delete = []

    inventory_set = {
        (item['inventory_group_code'], item['DescnPart1'], item['DescnPart2'], item['DescnPart3'])
        for item in inventory_items
    }

    fabric_set = {
        (fabric['inventory_group_code'], fabric['description_1'], fabric['description_2'], fabric['description_3'])
        for fabric in fabrics
    }

    # Determine additions and deletions
    for fabric in fabric_set - inventory_set:
        to_add.append(fabric)

    for item in inventory_set - fabric_set:
        to_delete.append(item)

    return to_add, to_delete


def create_workbook(headers_config, additions, deletions, output_path):
    from services.helper import parse_headers

    # Create workbook and parse headers
    wb = Workbook()
    inventory_file_excel_headers, inventory_file_db_fields = parse_headers(
        headers_config, "buz_inventory_item_file"
    )

    # Dictionary to track sheets by group
    sheets = {}

    def get_or_create_sheet(group_name):
        """Retrieve or create a sheet for the given group."""
        if group_name not in sheets:
            ws = wb.create_sheet(title=group_name)
            ws.append([])  # Row 1 blank
            ws.append(inventory_file_excel_headers)  # Row 2: Column headings
            sheets[group_name] = ws
        return sheets[group_name]

    # Add additions to their respective sheets
    for group, items in additions.items():
        ws = get_or_create_sheet(group)
        for item in items:
            row = [
                item.get(db_field, "") for db_field in inventory_file_db_fields
            ]
            ws.append(row)

    # Add deletions to their respective sheets
    for group, items in deletions.items():
        ws = get_or_create_sheet(group)
        for item in items:
            row = [
                item.get(db_field, "") for db_field in inventory_file_db_fields
            ]
            ws.append(row)

    # Save workbook
    wb.save(output_path)


def process_data(fabrics, inventory_items, inventory_groups):
    """
    Process data to determine additions and deletions.

    :param fabrics: List of fabrics with descriptions and supplier codes.
    :param inventory_items: List of inventory items with all data fields.
    :param inventory_groups: Mapping of group codes to group descriptions.
    :return: (additions, deletions)
    """
    additions = {}
    deletions = {}

    # Helper function to normalize strings for case-insensitivity
    def normalize(value):
        return value.lower() if isinstance(value, str) else value

    fabric_groups = set(normalize(fabric["inventory_group_code"]) for fabric in fabrics)

    # Convert inventory_items to a dictionary for quick lookup by group and description
    inventory_dict = {
        (normalize(item["inventory_group_code"]),
         normalize(item["DescnPart1"]),
         normalize(item["DescnPart2"]),
         normalize(item["DescnPart3"])): item
        for item in inventory_items
        if normalize(item["inventory_group_code"]) in fabric_groups
    }

    # Identify additions and deletions
    for fabric in fabrics:
        key = (
            normalize(fabric["inventory_group_code"]),
            normalize(fabric["description_1"]),
            normalize(fabric["description_2"]),
            normalize(fabric["description_3"]),
        )
        if key not in inventory_dict:
            # Addition
            group = fabric["inventory_group_code"]
            group_description = inventory_groups.get(group, "Unknown")  # Fallback to "Unknown" if group is missing
            # Use an existing item in the same group as a template
            template_item = next(
                (dict(item) for item in inventory_items if normalize(item["inventory_group_code"]) == group), {}
            )
            # Create a new record based on the template
            addition = template_item.copy() if template_item else {}
            addition.update({
                "inventory_group_code": group,
                "Description": f"{group_description} {fabric['description_1']} {fabric['description_2']} {fabric['description_3']}",
                "DescnPart1": fabric["description_1"],
                "DescnPart2": fabric["description_2"],
                "DescnPart3": fabric["description_3"],
                "SupplierProductCode": fabric["supplier_code"],
                "Operation": "A",  # Set Operation to 'A'
            })
            # Ensure all fields are present, fill with defaults if missing
            for column in template_item.keys():
                addition.setdefault(column, "")
            additions.setdefault(group, []).append(addition)
        else:
            # Remove from inventory_dict if it exists in the fabric list
            del inventory_dict[key]

        # Remaining items in inventory_dict are deletions
    for key, item in inventory_dict.items():
        group = item["inventory_group_code"]
        deletion = dict(item)  # Convert sqlite3.Row to a dictionary
        deletion["Operation"] = "D"  # Add Operation field
        deletions.setdefault(group, []).append(deletion)

    return additions, deletions
