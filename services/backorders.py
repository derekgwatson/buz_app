from datetime import datetime
from openpyxl import Workbook
from services.database import DatabaseManager
from services.google_sheets_service import GoogleSheetsService
from services.data_processing import get_all_inventory_items_by_group
import logging


# Configure logging
logger = logging.getLogger(__name__)


def backorder_datestr_to_date(backorder_date_str: str) -> datetime:
    """
    Format a backorder date string into a human-readable format.

    Args:
        backorder_date_str (str): The backorder date in "dd/mm/yyyy" format.

    Returns:
        str: The formatted date in "dd MMM yyyy" format.

    Raises:
        ValueError: If the input date string is not in the expected format.
    """
    try:
        return datetime.strptime(backorder_date_str, "%d/%m/%Y")
    except ValueError:
        raise ValueError(f"Invalid date format: {backorder_date_str}. Expected 'dd/mm/yyyy'.")


def generate_backorder_message(product: str, backorder_date_str: str) -> str:
    """
    Generate a backorder message for a product with a given backorder date.

    Args:
        product (str): The name or description of the product.
        backorder_date_str (str): The backorder date in "dd/mm/yyyy" format.

    Returns:
        str: A formatted backorder message.
    """
    formatted_date = backorder_datestr_to_date(backorder_date_str)
    return f"{product} on backorder until {formatted_date.strftime('%d %b %Y')}."


def process_inventory_backorder_with_services(
    _db_manager: DatabaseManager,
    _sheets_service: GoogleSheetsService,
    spreadsheet_id: str,
    range_name: str
):
    """
    Process inventory items for backorder messages using custom service classes and generate upload and original files.

    Args:
        _db_manager: Database Manager
        _sheets_service (GoogleSheetsService): Instance of GoogleSheetsService for interacting with Google Sheets.
        spreadsheet_id (str): Google Sheet ID containing supplier codes and backorder dates.
        range_name (str): Range in the Google Sheet to fetch data from.
    """
    # Step 1: Query inventory items from the database
    inventory_data = get_all_inventory_items_by_group(_db_manager)

    # Step 2: Read supplier codes and backorder dates from Google Sheets
    google_sheet_data = _sheets_service.fetch_sheet_data(spreadsheet_id, range_name)

    # Extract headers and map them to indices
    headers = google_sheet_data[0]
    header_map = {header: index for index, header in enumerate(headers)}

    # Ensure required columns exist
    required_columns = ['Unleashed Code', 'On backorder until']
    missing_columns = [col for col in required_columns if col not in header_map]
    if missing_columns:
        raise ValueError(f"Missing required columns in Google Sheet: {', '.join(missing_columns)}")

    # Convert Google Sheet data into a lookup dictionary
    backorder_lookup = {
        row[header_map['Unleashed Code']]: row[header_map['On backorder until']]
        for row in google_sheet_data[1:]  # Skip the header row
        if row[header_map['Unleashed Code']] and row[header_map['On backorder until']]  # Ensure non-empty values
    }

    # Step 3: Process each sheet in the inventory workbook
    upload_workbook = Workbook()
    original_workbook = Workbook()
    upload_workbook.remove(upload_workbook.active)  # Remove the default sheet
    original_workbook.remove(original_workbook.active)

    # Read all sheets at once
    processed_sheets = set()  # Track groups that are processed

    for group_code, rows in inventory_data.items():
        updated_rows = []
        original_rows = []

        for row in rows:
            original_row = dict(row)  # Preserve the original row before any changes (and convert to dict)
            new_row = dict(row)  # Preserve the original row before any changes (and convert to dict)
            old_message = row['Warning'] or ''
            if not row['SupplierProductCode']:  # skip when supplier product code is blank
                continue

            # Concatenate columns D, E, and F with spaces
            # Safely concatenate columns D, E, and F with spaces
            description_parts = [
                (row["DescnPart1"] or "").strip(),
                (row["DescnPart2"] or "").strip(),
                (row["DescnPart3"] or "").strip(),
            ]
            description = " ".join(part for part in description_parts if part)

            new_message = ''
            if row['SupplierProductCode'] in backorder_lookup:
                if backorder_datestr_to_date(backorder_lookup[row['SupplierProductCode']]) >= datetime.now():
                    new_message = generate_backorder_message(description, backorder_lookup[row['SupplierProductCode']])

            # Check if backorder message should be updated
            if new_message != old_message.strip():
                new_row['Warning'] = new_message
                new_row['Operation'] = 'E'
                updated_rows.append(new_row)
                original_rows.append(original_row)  # Keep the unmodified row in the original file

        # Exclude 'id' and 'inventory_group_code' from headers and rows
        exclude_columns = {'id', 'inventory_group_code'}

        # Only add updated rows to the upload workbook
        if updated_rows:
            processed_sheets.add(group_code)
            upload_sheet = upload_workbook.create_sheet(title=group_code)
            filtered_headers = [key for key in updated_rows[0].keys() if key not in exclude_columns]
            upload_sheet.append([])  # Write blank row
            upload_sheet.append(filtered_headers)  # Write filtered header
            for updated_row in updated_rows:
                filtered_row = [value for key, value in updated_row.items() if key not in exclude_columns]
                upload_sheet.append(filtered_row)

        # Add original rows to the original workbook
        if original_rows:
            original_sheet = original_workbook.create_sheet(title=group_code)
            filtered_headers = [key for key in original_rows[0].keys() if key not in exclude_columns]
            original_sheet.append([])  # Write blank row
            original_sheet.append(filtered_headers)  # Write filtered header
            for original_row in original_rows:
                filtered_row = [value for key, value in original_row.items() if key not in exclude_columns]
                original_sheet.append(filtered_row)

    # Step 4: return the workbooks (to save them, or inspect them if testing)
    return upload_workbook, original_workbook
