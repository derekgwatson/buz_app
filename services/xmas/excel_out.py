
from __future__ import annotations

import shutil
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from .parse import display_date_ddmmyy
from .model import LeadRow, CutoffRow


def _find_header_row_and_col(ws, header_text: str) -> Tuple[Optional[int], Optional[int]]:
    max_r = min(ws.max_row or 0, 100) or 100
    max_c = min(ws.max_column or 0, 30) or 30
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower() == header_text.lower():
                return r, c
    return None, None


def _first_false_row(ws, start_row: int, col_idx: int) -> Optional[int]:
    for r in range(start_row + 1, min((ws.max_row or 0), start_row + 200) + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v in (False, 0) or (isinstance(v, str) and str(v).strip().lower() in ("false", "no", "0")):
            return r
    return None


def _first_nonempty_row(ws, start_row: int, col_idx: int) -> Optional[int]:
    for r in range(start_row + 1, min((ws.max_row or 0), start_row + 200) + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v not in (None, ""):
            return r
    return None


def _append_message(value: str, lead_text: str, cutoff_display: Optional[str]) -> str:
    base = (value or "").rstrip()
    if base:
        base += ", "
    msg = lead_text
    if cutoff_display:
        msg += f" ***CHRISTMAS CUTOFF {cutoff_display}***"
    return (base + msg).strip()


def inject_and_prune(
    *,
    template_path: Path,
    out_path: Path,
    store_name: str,
    leads_by_code: Dict[str, LeadRow],
    cutoffs_by_code: Dict[str, CutoffRow],
    insertion_col_letter: str,
    header_text: str = "Do Not Show?",
    control_codes: Sequence[str],
    warnings: List[str],
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, out_path)

    wb = load_workbook(out_path, keep_vba=True, data_only=False, read_only=False)

    # Prune tabs
    if control_codes:
        wanted = set(c.upper() for c in control_codes)
        to_delete = [ws.title for ws in wb.worksheets if ws.title.upper() not in wanted]
        for name in to_delete:
            wb.remove(wb[name])

    for ws in list(wb.worksheets):
        code = ws.title.upper()
        if control_codes and code not in set(c.upper() for c in control_codes):
            continue
        header_row, header_col = _find_header_row_and_col(ws, header_text)
        if not header_row or not header_col:
            warnings.append(f"[{store_name}/{code}] Header '{header_text}' not found — skipped.")
            continue
        tgt_row = _first_false_row(ws, header_row, header_col)
        if not tgt_row:
            warnings.append(f"[{store_name}/{code}] No FALSE below '{header_text}' — skipped.")
            continue

        ins_col = ord(insertion_col_letter.upper()) - ord("A") + 1

        nn_row = _first_nonempty_row(ws, header_row, ins_col)
        if nn_row is not None and nn_row != tgt_row:
            warnings.append(f"[{store_name}/{code}] Row mismatch: first non-empty {insertion_col_letter}{nn_row} vs first FALSE at row {tgt_row}.")

        lr = leads_by_code.get(code)
        if not lr:
            warnings.append(f"[{store_name}/{code}] Lead-time not found — skipped.")
            continue
        cutoff = cutoffs_by_code.get(code)
        cutoff_display = display_date_ddmmyy(cutoff.cutoff_date) if cutoff else None

        current_val = ws.cell(row=tgt_row, column=ins_col).value or ""
        ws.cell(row=tgt_row, column=ins_col).value = _append_message(str(current_val), lr.lead_text, cutoff_display)

    wb.save(out_path)
