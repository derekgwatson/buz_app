from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


import logging
logger = logging.getLogger(__name__)


def get_current_buz_pricing(db):
    """
    Fetch current pricing records from the pricing_data table.

    Args:
        db: Database connection or DatabaseManager instance.

    Returns:
        dict: Mapping InventoryCode → row dictionary.
    """
    cursor = db.execute_query("SELECT * FROM pricing_data WHERE inventory_group_code IN ('CRTWT', 'CRTNT')")
    rows = cursor.fetchall()
    pricing = {}
    for row in rows:
        code = row['InventoryCode'].strip()
        pricing[code] = row
    return pricing


def prepare_pricing_changes(sheet_dict, buz_pricing):
    """
    Prepare pricing changes by comparing Google Sheet data to Buz pricing data.

    Args:
        sheet_dict (dict): Google Sheet data keyed by SupplierProductCode.
        buz_pricing (dict): Current Buz pricing data keyed by InventoryCode.

    Returns:
        dict: Grouped pricing changes ready for workbook export.
    """
    changes = {}
    tolerance = 0.005  # 0.5% tolerance

    for code, sheet_item in sheet_dict.items():
        price_value = sheet_item['raw_row'][8]
        if not price_value or not price_value.replace('.', '', 1).isdigit():
            logger.warning(f"Skipping non-numeric price for code {code}: {price_value}")
            continue

        sheet_price = float(sheet_item['raw_row'][8]) if sheet_item['raw_row'][8] else 0  # Example: column I (SellSQM)
        buz_item = buz_pricing.get(code)

        needs_update = False
        if not buz_item:
            needs_update = True
        else:
            buz_price = float(buz_item.get('SellSQM', 0))
            if buz_price == 0 or abs(sheet_price - buz_price) / buz_price > tolerance:
                needs_update = True

        if needs_update:
            row_data = {
                'PkId': '',
                'Operation': 'A',
                'InventoryCode': code,
                'SellSQM': sheet_price,
                'SupplierCode': sheet_item['code'],
                'SupplierDescn': sheet_item['fabric_name'],
                'DateFrom': (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d"),
                # Add more fields as needed if required for the upload
            }
            group = sheet_item.get('inventory_group_code', 'CRT')  # Derive group if available, default to 'CRT'
            changes.setdefault(group, []).append(row_data)

    return changes


class PricingWorkbookCreator:
    def __init__(self, headers_config, parse_headers_func):
        self.headers_config = headers_config
        self.parse_headers_func = parse_headers_func
        self.pricing_file_excel_headers, self.pricing_file_db_fields = self.parse_headers_func(
            self.headers_config, "buz_pricing_file"
        )
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.sheets = {}

    def _get_or_create_sheet(self, group_name):
        if group_name not in self.sheets:
            ws = self.workbook.create_sheet(title=group_name)
            ws.append([])
            ws.append(self.pricing_file_excel_headers)
            self.sheets[group_name] = ws
        return self.sheets[group_name]

    def _add_items_to_sheet(self, group_name, items):
        ws = self._get_or_create_sheet(group_name)
        for item in items:
            item_dict = dict(item)
            row = [item_dict.get(db_field, "") for db_field in self.pricing_file_db_fields]
            ws.append(row)

    def populate_workbook(self, changes=None):
        for group, items in changes.items():
            self._add_items_to_sheet(group, items)

    def save_workbook(self, output_path):
        if not self.sheets:
            logger.info("No pricing updates found. No workbook created.")
            return None  # Optional: return None to indicate no file was saved
        self.workbook.save(output_path)
        return self.workbook

    def auto_fit_columns(self):
        for sheet_name, sheet in self.sheets.items():
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 2


def create_pricing_workbook_creator(app):
    from services.helper import parse_headers

    return PricingWorkbookCreator(
        headers_config=app.config["headers"],
        parse_headers_func=parse_headers
    )
