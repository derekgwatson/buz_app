from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class InventoryWorkbookCreator:
    def __init__(self, headers_config, parse_headers_func):
        """
        Initialize the InventoryWorkbookCreator with inventory data and header configuration.

        Args:
            headers_config (dict): JSON configuration for headers.
            parse_headers_func (function): Function to parse headers based on configuration.
        """
        self.headers_config = headers_config
        self.parse_headers_func = parse_headers_func
        self.inventory_file_excel_headers, self.inventory_file_db_fields = self.parse_headers_func(
            self.headers_config, "buz_inventory_item_file"
        )
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.sheets = {}

    def _get_or_create_sheet(self, group_name):
        """
        Retrieve or create a sheet for the given group.

        Args:
            group_name (str): The name of the group.

        Returns:
            Worksheet: The worksheet for the group.
        """
        if group_name not in self.sheets:
            ws = self.workbook.create_sheet(title=group_name)
            ws.append([])  # Row 1 blank
            ws.append(self.inventory_file_excel_headers)  # Row 2: Column headings
            self.sheets[group_name] = ws
        return self.sheets[group_name]

    def _add_items_to_sheet(self, group_name, items):
        """
        Add items to a specific sheet.

        Args:
            group_name (str): The name of the group.
            items (list): List of item dictionaries to add to the sheet.
        """
        ws = self._get_or_create_sheet(group_name)
        for item in items:
            item_dict = dict(item)
            row = [item_dict.get(db_field, "") for db_field in self.inventory_file_db_fields]
            ws.append(row)

    def populate_workbook(self, changes=None):
        """
        Populate the workbook with additions and deletions.

        Args:
            changes (dict): A dictionary of items to add/edit/delete where keys are group codes and values are lists
                            of rows (dict).

        """

        for group, items in changes.items():
            self._add_items_to_sheet(group, items)

    def save_workbook(self, output_path):
        """
        Save the workbook to the specified path.

        Args:
            output_path (str): The file path to save the workbook.

        Returns:
            Workbook: The populated workbook.
        """
        self.workbook.save(output_path)
        return self.workbook

    def auto_fit_columns(self):
        """
        Adjust the width of all columns in all sheets to fit their content.
        """
        for sheet_name, sheet in self.sheets.items():
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)  # Get column letter
                for cell in column_cells:
                    try:
                        # Calculate the length of the cell's value
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        # Handle errors gracefully (e.g., if cell.value is None)
                        pass
                # Set column width (add a little extra space for padding)
                sheet.column_dimensions[column_letter].width = max_length + 2


def create_inventory_workbook_creator(app):
    """
    Factory function to create an instance of InventoryWorkbookCreator.

    Args:
        app: App context

    Returns:
        InventoryWorkbookCreator: An instance of the class configured with app headers.
    """
    from services.helper import parse_headers

    return InventoryWorkbookCreator(
        headers_config=app.config["headers"],
        parse_headers_func=parse_headers
    )
