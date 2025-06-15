import openpyxl
from io import BytesIO
import logging
from openpyxl.utils import column_index_from_string, get_column_letter
from services.database import DatabaseManager


# Configure logging
logger = logging.getLogger(__name__)


class OpenPyXLFileHandler:
    """
    A file handler class that abstracts operations for reading Excel files using openpyxl.
    """

    def __init__(self, workbook=None):
        """
        Initialize the file handler with an existing workbook.
        """
        self.workbook = workbook

    @classmethod
    def from_file(cls, file_path, data_only=True):
        """
        Initialize the file handler with an Excel file from disk.

        Args:
            file_path (str): Path to the Excel file.
            data_only (bool): Whether to read the values instead of formulas.

        Returns:
            OpenPyXLFileHandler: An initialized file handler.
        """
        logger.debug(f"File path we're loading the excel from is {file_path}")
        workbook = openpyxl.load_workbook(file_path, data_only=data_only)
        return cls(workbook=workbook)

    @classmethod
    def from_file_like(cls, file, data_only=True):
        """
        Initialize the file handler with a file-like object.

        Args:
            file: A file-like object (e.g., from `request.files`).
            data_only (bool): Whether to read the values instead of formulas.

        Returns:
            OpenPyXLFileHandler: An initialized file handler.
        """
        workbook = openpyxl.load_workbook(BytesIO(file.read()), data_only=data_only)
        return cls(workbook=workbook)

    @classmethod
    def from_sheets_data(cls, sheets_data, sheets_header_data):
        """
        Initialize the file handler with sheets data.

        Args:
            sheets_data (dict): Dictionary where keys are sheet names, and values are lists of rows.
            sheets_header_data (dict): Dictionary with:
                - `headers`: List of column headers (shared by all sheets).
                - `header_row`: Row index where headers should appear.

        Returns:
            OpenPyXLFileHandler: An initialized file handler.
        """
        handler = cls()
        handler._create_excel_file(sheets_data, sheets_header_data)
        return handler

    @classmethod
    def from_items(cls, items, headers_config, header_row=1):
        """
        Initialize the file handler with items and headers configuration.

        Args:
            items (list[dict]): Inventory items to process.
            headers_config (list[dict]): Configuration mapping database fields to headers.
            header_row (int): Row where headers are located in the Excel sheet.

        Returns:
            OpenPyXLFileHandler: An initialized file handler.
        """
        sheets_data = cls.transform_items_to_sheets_data(items, headers_config, header_row)
        return cls.from_sheets_data(sheets_data)

    @staticmethod
    def transform_items_to_sheets_data(items, headers_config, header_row=1):
        """
        Transform items into sheets data format for Excel creation.
        """
        headers = [header["spreadsheet_column"] for header in headers_config]
        database_fields = [header["database_field"] for header in headers_config]

        grouped_data = {}
        for item in items.values():
            group_code = item["inventory_group_code"]
            if group_code not in grouped_data:
                grouped_data[group_code] = []
            grouped_data[group_code].append({field: item[field] for field in database_fields})

        sheets_data = {}
        for group_code, group_items in grouped_data.items():
            sheet_data = [
                [item[field] for field in database_fields]
                for item in group_items
            ]
            sheets_data[group_code] = (sheet_data, headers, header_row)

        return sheets_data

    def get_sheet_names(self):
        """
        Get the names of all sheets in the workbook.

        :return: List of sheet names
        :rtype: list[str]
        """
        if self.workbook is None:
            raise ValueError("Workbook is not loaded.")
        return self.workbook.sheetnames

    def get_sheet(self, sheet_name):
        """
        Get a specific sheet by name.

        :param sheet_name: Name of the sheet
        :type sheet_name: str
        :return: The sheet object
        :rtype: openpyxl.worksheet.worksheet.Worksheet
        """
        if self.workbook is None:
            raise ValueError("Workbook is not loaded.")
        return self.workbook[sheet_name]

    def get_headers(self, sheet, header_row):
        """
        Get headers from a specific row in a sheet.

        :param sheet: The sheet object
        :type sheet: openpyxl.worksheet.worksheet.Worksheet
        :param header_row: The row number containing headers
        :type header_row: int
        :return: List of headers
        :rtype: list[str]
        """
        return [sheet.cell(row=header_row, column=col).value for col in range(1, sheet.max_column + 1)]

    def get_rows(self, sheet, start_row):
        """
        Get all rows starting from a specific row.

        :param sheet: The sheet object
        :type sheet: openpyxl.worksheet.worksheet.Worksheet
        :param start_row: The starting row number
        :type start_row: int
        :return: List of rows, where each row is a tuple of cell values
        :rtype: list[tuple]
        """
        return list(sheet.iter_rows(min_row=start_row, values_only=True))

    def read_sheet_to_dict(self, header_row: int = 1):
        """
        Read all sheets in the workbook into dictionaries, with configurable header rows.

        :param header_row: A dictionary mapping sheet names to header row numbers.
                            If None, defaults to the first row for all sheets.
        :type header_row: int
        :return: A dictionary where keys are sheet names, and values are lists of row dictionaries
        :rtype: dict[str, list[dict]]
        """
        if self.workbook is None:
            raise ValueError("Workbook is not loaded.")

        # Default to header row 1 for all sheets if no configuration is provided
        all_data = {}

        for sheet_name in self.get_sheet_names():
            sheet = self.get_sheet(sheet_name)
            headers = self.get_headers(sheet, header_row)
            rows = self.get_rows(sheet, header_row + 1)
            all_data[sheet_name] = [dict(zip(headers, row)) for row in rows if any(row)]

        return all_data

    def _create_excel_file(self, sheets_data, sheets_header_data):
        """
        Internal method to create a new Excel workbook with multiple sheets.

        Args:
            sheets_data (dict): Dictionary where keys are sheet names, and values are lists of rows.
            sheets_header_data (dict): Contains:
                - `headers`: List of column headers.
                - `header_row`: Row index for headers (default is 1).

        Modifies:
            self.workbook: Sets this attribute to the newly created workbook.
        """
        self.workbook = openpyxl.Workbook()

        headers = sheets_header_data["headers"]
        header_row = sheets_header_data.get("header_row", 1)

        for idx, (sheet_name, rows) in enumerate(sheets_data.items(), start=1):
            # Add a new sheet or use the default active sheet
            if idx == 1:
                sheet = self.workbook.active
                sheet.title = sheet_name
            else:
                sheet = self.workbook.create_sheet(title=sheet_name)

            # Write headers
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Write data starting below the header row
            data_start_row = header_row + 1
            for row_idx, row in enumerate(rows, start=data_start_row):
                for col_idx, value in enumerate(row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

    def save_workbook(self, save_path):
        """
        Save the current workbook to the specified file path.

        :param save_path: The file path where the workbook should be saved.
        :type save_path: str
        """
        if self.workbook is None:
            raise ValueError("No workbook is loaded or created to save.")

        if not self.workbook.sheets:
            logger.info("No pricing updates found. No workbook created.")
            return None  # Optional: return None to indicate no file was saved

        self.workbook.save(save_path)
        logger.info(f"Workbook saved to {save_path}")

    def get_column_by_header(self, sheet_name, header_name, header_row=1):
        """
        Get all values in a column by its header text.

        Args:
            sheet_name (str): The name of the sheet to search in.
            header_name (str): The header text of the column.
            header_row (int): The row number containing the headers. Defaults to 1.

        Returns:
            list: A list of values from the specified column (excluding the header).

        Raises:
            ValueError: If the header or sheet is not found.
        """
        # Get the sheet
        sheet = self.get_sheet(sheet_name)

        # Get the headers
        headers = self.get_headers(sheet, header_row)

        # Find the column index of the header
        if header_name not in headers:
            raise ValueError(f"Header '{header_name}' not found in sheet '{sheet_name}'.")
        column_index = headers.index(header_name) + 1  # 1-based index

        # Collect values from the column, excluding the header
        values = [
            sheet.cell(row=row, column=column_index).value
            for row in range(header_row + 1, sheet.max_row + 1)
        ]
        return values

    def set_value_by_header(self, sheet_name, header_name, row, value, header_row=1):
        """
        Set a single value in a specific row of a column identified by its header text.

        Args:
            sheet_name (str): The name of the sheet to search in.
            header_name (str): The header text of the column.
            row (int): The row number (1-based) where the value should be set.
            value: The value to set in the specified cell.
            header_row (int): The row number containing the headers. Defaults to 1.

        Raises:
            ValueError: If the header or sheet is not found.
            IndexError: If the specified row is outside the range of the sheet.
        """
        # Get the sheet
        sheet = self.get_sheet(sheet_name)

        # Get the headers
        headers = self.get_headers(sheet, header_row)

        # Find the column index of the header
        if header_name not in headers:
            raise ValueError(f"Header '{header_name}' not found in sheet '{sheet_name}'.")
        column_index = headers.index(header_name) + 1  # 1-based index

        # Validate the row number
        if row < header_row + 1 or row > sheet.max_row:
            raise IndexError(f"Row {row} is out of range for sheet '{sheet_name}'.")

        # Set the value in the specified cell
        sheet.cell(row=row, column=column_index, value=value)

    @classmethod
    def create_blank_pricing_upload_from_config(cls, config: list[dict], group_codes: list[str], header_row: int = 1):
        """
        Create a blank pricing upload workbook using the buz_pricing_file config and inventory group codes.

        :param config: The buz_pricing_file config section (list of dicts with spreadsheet_column keys).
        :param group_codes: List of inventory_group_codes to create sheets for.
        :param header_row: Row number to place headers (default is 1).
        :return: OpenPyXLFileHandler with empty sheets per group.
        """
        headers = [entry["spreadsheet_column"] for entry in config]
        sheets_data = {group: [] for group in group_codes}
        sheets_header_data = {
            "headers": headers,
            "header_row": header_row
        }
        return cls.from_sheets_data(sheets_data, sheets_header_data)

    def clean_for_upload(cls, db_manager: DatabaseManager, allowed_sheets: list[str], show_only_valid_unleashed=False):
        """
        Clean the workbook:
        - Remove sheets not in allowed_sheets
        - Strip trailing '*' from cell C2
        - Hide all columns except A-F, AA-AC, AO
        - Optionally filter out rows with invalid Unleashed codes (column AB)
        """

        keep_cols = (
                list(range(column_index_from_string("A"), column_index_from_string("F") + 1)) +
                list(range(column_index_from_string("AA"), column_index_from_string("AC") + 1)) +
                [column_index_from_string("AO")]
        )

        # Step 1: Remove unwanted sheets
        for sheet_name in cls.get_sheet_names()[:]:  # copy to avoid mutation issues
            if sheet_name not in allowed_sheets:
                del cls.workbook[sheet_name]

        # Step: add UL tab with ProductCode and ProductDescription
        ul_data = db_manager.execute_query("""
            SELECT ProductCode, ProductDescription
            FROM unleashed_products
            WHERE ProductCode IS NOT NULL
        """).fetchall()

        if "UL" in cls.get_sheet_names():
            del cls.workbook["UL"]

        ul_sheet = cls.workbook.create_sheet("UL")
        ul_sheet.cell(row=1, column=1, value="ProductCode")
        ul_sheet.cell(row=1, column=2, value="ProductDescription")

        for i, row in enumerate(ul_data, start=2):
            ul_sheet.cell(row=i, column=1, value=row["ProductCode"])
            ul_sheet.cell(row=i, column=2, value=row["ProductDescription"])

        # Create set of valid ProductCodes
        valid_codes = {str(row["ProductCode"]).strip().upper() for row in ul_data}

        # Step 2 & 3: Process remaining sheets
        for sheet in cls.workbook.worksheets:
            if sheet.title == "UL":
                continue

            # Clean C2
            val = sheet["C2"].value
            if isinstance(val, str) and val.endswith("*"):
                sheet["C2"].value = val.rstrip("*")

            # Hide unwanted columns
            for col_idx in range(1, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(col_idx)].hidden = col_idx not in keep_cols

            # Step 3: Extract data from row 3 down, optionally filtering invalid ABs
            data_rows = []
            for row in sheet.iter_rows(min_row=3, values_only=True):
                if not any(row):
                    continue  # skip empty rows

                if show_only_valid_unleashed:
                    ab_value = row[27] if len(row) > 27 else None  # AB is column 28, index 27
                    ab_str = str(ab_value).strip().upper() if ab_value is not None else ""

                    if not ab_str or ab_str in valid_codes:
                        continue  # remove row if blank or matched

                data_rows.append(row)

            # Sort by column C (index 2)
            data_rows.sort(key=lambda r: r[2])

            # Overwrite rows
            start_row = 3
            max_col = sheet.max_column
            for i, row_data in enumerate(data_rows, start=start_row):
                for j, val in enumerate(row_data, start=1):
                    sheet.cell(row=i, column=j, value=val)

            # Clear any leftover rows
            for i in range(len(data_rows) + start_row, sheet.max_row + 1):
                for j in range(1, max_col + 1):
                    sheet.cell(row=i, column=j).value = None

            # Set row heights
            for i in range(1, sheet.max_row + 1):
                sheet.row_dimensions[i].height = 18.75 if i == 1 else 15

            # Add formula to column AC
            for row_idx in range(3, sheet.max_row + 1):
                sheet.cell(row=row_idx, column=29).value = f'=VLOOKUP(AB{row_idx},UL!A:B,2,FALSE)'

        # Collect empty sheets based on A3
        empty_sheets = []
        for sheet in cls.workbook.worksheets:
            if sheet.title == "UL":
                continue
            if sheet["A3"].value is None or str(sheet["A3"].value).strip() == "":
                empty_sheets.append(sheet.title)

        # Delete empty sheets
        for sheet_name in empty_sheets:
            del cls.workbook[sheet_name]

    def extract_motorisation_data(self, db_manager: DatabaseManager) -> tuple[list[dict], list[str]]:
        """
        Extracts and groups motorisation-related data:
        - Skips rows with no code.
        - Groups by code.
        - Collects unique products and question headings.
        - Picks one description per code.

        Returns:
            list[dict]: Grouped data including code, description, products, questions, pricing, and
            supplier_product_code.
        """
        if not self.workbook:
            raise ValueError("Workbook not loaded.")

        raw_entries = []

        for sheet_name in self.get_sheet_names():
            sheet = self.get_sheet(sheet_name)
            max_col = sheet.max_column
            max_row = sheet.max_row

            for col_idx in range(1, max_col + 1):
                question_heading = sheet.cell(row=7, column=col_idx).value
                has_lookback = bool(sheet.cell(row=6, column=col_idx).value)

                col_has_motor = any(
                    (cell := sheet.cell(row=r, column=col_idx)).value
                    and isinstance(cell.value, str)
                    and 'motor' in cell.value.lower()
                    for r in range(1, max_row + 1)
                )

                if not col_has_motor:
                    continue

                for row_idx in range(17, max_row + 1):
                    raw_value = sheet.cell(row=row_idx, column=col_idx).value
                    if not raw_value or not isinstance(raw_value, str):
                        continue

                    parts = [p.strip() for p in raw_value.split('|')]

                    if has_lookback:
                        description = parts[1] if len(parts) > 1 else ''
                        code = parts[2] if len(parts) > 2 else ''
                    else:
                        description = parts[0] if len(parts) > 0 else ''
                        code = parts[1] if len(parts) > 1 else ''

                    if not code:
                        continue

                    raw_entries.append({
                        "product": sheet_name,
                        "question": question_heading,
                        "description": description,
                        "code": code
                    })

        # Group by code
        grouped = {}
        for entry in raw_entries:
            code = entry["code"]
            if code not in grouped:
                grouped[code] = {
                    "code": code,
                    "description": entry["description"],
                    "products": set(),
                    "questions": set()
                }
            grouped[code]["products"].add(entry["product"])
            if entry["question"]:
                grouped[code]["questions"].add(entry["question"])

        # Fetch pricing for codes found
        pricing_rows = db_manager.execute_query("""
            SELECT * FROM pricing_data
            WHERE inventorycode IN ({})
        """.format(','.join('?' for _ in grouped)), tuple(grouped.keys())).fetchall()

        # Fetch supplier product codes for relevant inventory codes
        supplier_codes = db_manager.execute_query("""
            SELECT code, supplierproductcode
            FROM inventory_items
            WHERE code IN ({})
        """.format(','.join('?' for _ in grouped)), tuple(grouped.keys())).fetchall()

        supplier_lookup = {row["code"]: row["supplierproductcode"] for row in supplier_codes}

        # Organize pricing per code, filtering out empty/zero values
        pricing_lookup = {}
        for row in pricing_rows:
            code = row["inventorycode"]
            pricing_fields = {
                key: row[key]
                for key in row.keys()
                if (key.lower().startswith("sell") or key.lower().startswith("cost"))
                   and row[key] not in (None, 0, 0.0)
            }
            if pricing_fields:
                pricing_lookup[code] = pricing_fields

        # Get all pricing keys actually used (across all items)
        all_pricing_keys = set()
        for fields in pricing_lookup.values():
            all_pricing_keys.update(fields.keys())

        pricing_fields = sorted(all_pricing_keys)  # Sort if you want consistent order

        # Final output
        result = []
        for code, data in grouped.items():
            if code not in pricing_lookup:
                continue  # Skip codes with no pricing

            result.append({
                "code": code,
                "description": data["description"],
                "products": ", ".join(sorted(data["products"])),
                "questions": ", ".join(sorted(data["questions"])),
                "pricing": pricing_lookup[code],
                "supplier_product_code": supplier_lookup.get(code, "")
            })

        return result, pricing_fields
