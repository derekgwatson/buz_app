from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Set, Optional
import os
import time

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

# ---- Lead-time rewrites that preserve all surrounding text exactly ----
import re

_LEAD_HEADER_RE = re.compile(r'(?i)lead\s*time\s*:\s*')
_READY_IN_RE = re.compile(r'(?i)\bready\s*in\b')


def _review(review_set: set[str], warnings: list[str], store: str, code: str, reason: str) -> None:
    review_set.add(code)
    warnings.append(f"[{store}/{code}] {reason} — needs review.")


def _tab_rgb(ws: Worksheet) -> tuple[int, int, int] | None:
    """Return (r,g,b) from the sheet tab colour if present, else None."""
    col = getattr(ws.sheet_properties, "tabColor", None)
    if not col or not getattr(col, "rgb", None):
        return None
    h = col.rgb.upper()
    # ARGB or RGB → keep last 6 as RRGGBB
    if len(h) >= 6:
        h = h[-6:]
    try:
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except Exception:
        return None


def _tab_is_reddish(ws: Worksheet) -> bool:
    """True if tab colour is red-dominant (robust to shades)."""
    rgb = _tab_rgb(ws)
    if not rgb:
        return False
    r, g, b = rgb
    return r >= 180 and (r - max(g, b)) >= 40  # clearly red-dominant


_READY_IN_CELL_RE = re.compile(r'(?i)\bready\s*in\b')
_LEAD_HEADER_CELL_RE = re.compile(r'(?i)\blead\s*time\s*:\s*')


def _find_summary_ready_row(ws: Worksheet, ins_idx: int, header_row: int, ready_phrase: str = "Ready in") -> Optional[int]:
    start = max(1, int(header_row)) + 1
    max_row = ws.max_row or start
    for r in range(start, max_row + 1):
        v = ws.cell(row=r, column=ins_idx).value
        if isinstance(v, str) and _READY_IN_CELL_RE.search(v):
            return r
    return None


def _find_detailed_lead_row(ws: Worksheet, ins_idx: int, header_row: int) -> Optional[int]:
    start = max(1, int(header_row)) + 1
    max_row = ws.max_row or start
    for r in range(start, max_row + 1):
        v = ws.cell(row=r, column=ins_idx).value
        if isinstance(v, str) and _LEAD_HEADER_CELL_RE.search(v):
            return r
    return None


def _is_nonblank(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def _is_false_checkbox(value) -> bool:
    if value is False:
        return True
    if isinstance(value, str) and value.strip().upper() == "FALSE":
        return True
    if value == 0:
        return True
    return False


def _is_effectively_empty_sheet(ws: Worksheet, header_row: int | None) -> bool:
    """
    Treat as a heading-only tab when there's no meaningful data below the header.
    We scan A:Z; ignore a lone value in G{header_row+1}.
    """
    start_row = (int(header_row) if header_row and header_row >= 1 else 1) + 1
    max_row = ws.max_row or start_row

    for r in range(start_row, max_row + 1):
        for c in range(1, 27):  # A..Z
            if r == start_row and c == 7:  # ignore G on the first data row
                continue
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            return False
    return True


def _norm_name(s: str) -> str:
    return (s or "").strip().upper()


def _prune_tabs_safe(wb, keep_names: set[str], warnings: list[str]) -> None:
    """
    Keep only tabs whose names match keep_names (case-insensitive, trimmed).
    Never remove all sheets. Emit a concise diff warning.
    """
    keep_norm = {_norm_name(s) for s in keep_names if s}
    tab_raw = list(wb.sheetnames)
    tab_norm_map = {_norm_name(n): n for n in tab_raw}

    if not keep_norm:
        warnings.append("[PRUNE] Control list is empty — skipping prune.")
        return

    tabs_norm = set(tab_norm_map.keys())
    missing_tabs_for_codes = sorted(keep_norm - tabs_norm)
    extra_tabs_not_codes   = sorted(tabs_norm - keep_norm)

    def _pv(items, k=12):
        return "—" if not items else (", ".join(items[:k]) + (f", +{len(items)-k} more" if len(items)>k else ""))

    warnings.append(
        "[PRUNE] control_codes=%d, workbook_tabs=%d; codes-without-tabs: %s; tabs-not-in-codes: %s"
        % (len(keep_norm), len(tab_raw), _pv(missing_tabs_for_codes), _pv(extra_tabs_not_codes))
    )

    to_keep_raw = [tab_norm_map[k] for k in (keep_norm & tabs_norm)]
    if not to_keep_raw:
        warnings.append("[PRUNE] No tabs match control list — skipping prune (will NOT delete anything).")
        return

    for name in tab_raw:
        if name not in to_keep_raw:
            wb.remove(wb[name])

    first_keep = to_keep_raw[0]
    ws0 = wb[first_keep]
    if getattr(ws0, "sheet_state", "visible") != "visible":
        ws0.sheet_state = "visible"
    wb.active = wb.sheetnames.index(first_keep)


def _col_to_idx(letter: str) -> int:
    letter = (letter or "").strip().upper()
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - 64)
    return max(1, n)


def _find_first_false(ws, col_idx: int, header_row: int) -> int | None:
    max_row = ws.max_row or header_row
    for r in range(header_row + 1, max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is False:
            return r
        if isinstance(v, str) and v.strip().upper() == "FALSE":
            return r
    return None


def _apply_template(template: str, lead: str) -> str:
    # Accept {LEAD}, {lead}, {lead_time}
    return (template or "{LEAD}").replace("{LEAD}", lead).replace("{lead}", lead).replace("{lead_time}", lead)


@dataclass
class InjectResult:
    saved_path: Path
    review_codes: Set[str]  # tabs that need manual review for this workbook


def _find_eol_index(s: str, start: int) -> int:
    """First EOL marker at/after start; supports literal '\\n' or real newlines; -1 if none."""
    i_lit = s.find("\\n", start)
    i_n = s.find("\n", start)
    i_r = s.find("\r", start)
    candidates = [i for i in (i_lit, i_n, i_r) if i != -1]
    return min(candidates) if candidates else -1


def rewrite_detailed_preserving_context(cell_text: str, new_lead: str) -> str:
    """
    Detailed: replace ONLY the value after 'Lead Time:' up to the next EOL/end,
    preserving any whitespace right before that EOL.
    """
    if not cell_text:
        return cell_text
    m = _LEAD_HEADER_RE.search(cell_text)
    if not m:
        return cell_text

    s = cell_text
    a = m.end()                      # start of current lead value
    b = _find_eol_index(s, a)        # index of first EOL token; -1 if none
    if b == -1:
        # No EOL — replace to end (no boundary whitespace to preserve)
        return s[:a] + new_lead

    # Preserve any spaces/tabs immediately before the EOL token
    p = b - 1
    while p >= a and s[p].isspace():
        p -= 1
    pre_eol_ws = s[p + 1:b]          # could be '' or ' ' or multiple spaces

    return s[:a] + new_lead + pre_eol_ws + s[b:]


def _scan_paren_span(s: str, open_idx: int) -> int:
    """Given s[open_idx] == '(', return index just after its matching ')'; nesting OK; len(s) if unclosed."""
    depth = 0
    i = open_idx
    n = len(s)
    while i < n:
        ch = s[i]
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth == 0:
                return i + 1
        i += 1
    return n


def rewrite_summary_preserving_context(cell_text: str, new_lead: str) -> str:
    """
    Summary: replace ONLY the lead value after 'Ready in', preserving:
      - the exact whitespace after 'Ready in'
      - the exact whitespace before the first top-level '(' or ',' (or EOL)
      - the entire bracket group (if present) and the suffix unchanged
    """
    if not cell_text:
        return cell_text
    m = _READY_IN_RE.search(cell_text)
    if not m:
        return cell_text

    s = cell_text
    n = len(s)
    phrase_end = m.end()

    # Keep exactly the original spaces after 'Ready in'
    j = phrase_end
    while j < n and s[j].isspace():
        j += 1
    ws_after_phrase = s[phrase_end:j]
    lead_start = j

    # Scan to the first top-level delimiter
    k = lead_start
    delim_idx = n
    delim_kind = "end"
    while k < n:
        if k + 1 < n and s[k] == "\\" and s[k + 1] == "n":
            delim_idx = k
            delim_kind = "eol_lit"
            break
        ch = s[k]
        if ch in ("\n", "\r"):
            delim_idx = k
            delim_kind = "eol"
            break
        if ch == "(":
            delim_idx = k
            delim_kind = "paren"
            break
        if ch == ",":
            delim_idx = k
            delim_kind = "comma"
            break
        k += 1

    # Preserve any whitespace immediately before the delimiter
    p = delim_idx - 1
    while p >= lead_start and s[p].isspace():
        p -= 1
    pre_delim_ws = s[p + 1:delim_idx]  # could be '' or ' '

    # Build bracket chunk and after
    if delim_kind == "paren":
        bracket_end = _scan_paren_span(s, delim_idx)
        bracket_chunk = s[delim_idx:bracket_end]
        after = s[bracket_end:]
    else:
        bracket_chunk = ""
        after = s[delim_idx:]

    prefix = s[:phrase_end] + ws_after_phrase
    return prefix + new_lead + pre_delim_ws + bracket_chunk + after


def inject_and_prune(
    *,
    template_path: Path,
    out_path: Path,
    store_name: str,
    leads_by_code: Dict[str, Dict],
    insertion_col_letter: str,
    anchor_col_letter: str,
    anchor_header_row: int,
    control_codes: Set[str],
    warnings: list[str],
    workbook_kind: str | None = None,
    # DETAILED: template to prefix when NO 'Lead Time:' exists in col B
    detailed_prefix_template: Optional[str] = None,
) -> InjectResult:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(filename=str(template_path), keep_vba=False, data_only=True)

    _prune_tabs_safe(wb, keep_names=set(control_codes), warnings=warnings)

    ins_idx = _col_to_idx(insertion_col_letter)
    anc_idx = _col_to_idx(anchor_col_letter)
    is_summary = (workbook_kind or "").lower() == "summary"

    review_codes: Set[str] = set()

    for code in list(wb.sheetnames):
        ws = wb[code]

        # Silently drop heading-only tabs
        if _is_effectively_empty_sheet(ws, anchor_header_row):
            wb.remove(ws)
            continue

        # Review by tab colour
        if _tab_is_reddish(ws):
            _review(review_codes, warnings, store_name, code, f"Tab coloured red (RGB={_tab_rgb(ws)})")
            continue  # do not write into coloured tabs

        # Lead record
        lead_rec = leads_by_code.get(code)
        if not lead_rec:
            warnings.append(f"[{store_name}/{code}] No lead-time text found — skipped.")
            _review(review_codes, warnings, store_name, code, "No lead-time text")
            continue
        lt_text = (lead_rec.get("lead_time_text") or "").strip()
        if not lt_text:
            warnings.append(f"[{store_name}/{code}] Empty lead-time text — skipped.")
            continue

        if is_summary:
            # --- SUMMARY ---
            ready_row = _find_summary_ready_row(ws, ins_idx, anchor_header_row, ready_phrase="Ready in")
            if ready_row is None:
                warnings.append(
                    f"[{store_name}/{code}] Summary: no 'Ready in' in {insertion_col_letter} — needs review."
                )
                _review(review_codes, warnings, store_name, code, "Summary: missing 'Ready in' in column C")
                continue

            cell = ws.cell(row=ready_row, column=ins_idx)
            original = cell.value if isinstance(cell.value, str) else ""

            # Update the lead that follows "Ready in"
            new_text = rewrite_summary_preserving_context(original, lt_text)
            if new_text != original:
                cell.value = new_text

        else:
            # === DETAILED ===
            lead_row = _find_detailed_lead_row(ws, ins_idx, anchor_header_row)

            if lead_row is None:
                # No 'Lead Time:' anywhere in col B → prefix at first FALSE in anchor column
                anchor_row = _find_first_false(ws, anc_idx, header_row=anchor_header_row)
                if anchor_row is None:
                    warnings.append(
                        f"[{store_name}/{code}] Detailed: no 'Lead Time:' and no FALSE in {anchor_col_letter} — skipped."
                    )
                    review_codes.add(code)
                    continue

                # --- DETAILED, no 'Lead Time:' present ---
                cell = ws.cell(row=anchor_row, column=ins_idx)
                existing = cell.value if isinstance(cell.value, str) else (
                    "" if cell.value is None else str(cell.value))
                template = detailed_prefix_template or "\n       -       Lead Time: {LEAD} \n       -       "

                proposed = f"{_apply_template(template, lt_text)}{existing}"

                if proposed == (
                cell.value if isinstance(cell.value, str) else ("" if cell.value is None else str(cell.value))):
                    _review(review_codes, warnings, store_name, code,
                            "Detailed: no 'Lead Time:'; proposed equals existing")
                else:
                    cell.value = proposed
                    _review(review_codes, warnings, store_name, code,
                            "Detailed: no 'Lead Time:'; prefixed at first FALSE")

            else:
                # Replace in-place on the found 'Lead Time:' line
                cell = ws.cell(row=lead_row, column=ins_idx)
                original = cell.value if isinstance(cell.value, str) else ""

                # Replace only the 'Lead Time:' value, then normalize banners for the whole cell
                new_text = rewrite_detailed_preserving_context(original, lt_text)
                if new_text != original:
                    cell.value = new_text

    # Keep at least one visible sheet
    if not any(s.sheet_state == "visible" for s in wb.worksheets):
        stub = wb.create_sheet("NO CHANGES")
        stub["A1"] = f"{store_name} — All tabs unchanged (lead times match current)."
        stub.sheet_state = "visible"

    final_path = out_path.with_suffix(".xlsx")
    wb.save(str(final_path))
    wb.close()

    try:
        ts = time.strftime("%H:%M:%S", time.localtime(os.path.getmtime(final_path)))
        warnings.append(f"[DEBUG] Saved {Path(final_path).name} at {ts} -> {final_path}")
    except Exception:
        pass

    if review_codes:
        warnings.append(f"[DEBUG/{store_name}] Review tabs: " + ", ".join(sorted(review_codes))[:400])

    return InjectResult(saved_path=final_path, review_codes=review_codes)


def save_review_only_workbook(
    *,
    template_path: Path,
    out_path: Path,
    review_codes: Set[str],
    warnings: list[str],
) -> Path:
    """
    Create a review-only workbook: no edits, only the tabs in `review_codes`.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(filename=str(template_path), keep_vba=False, data_only=True)
    _prune_tabs_safe(wb, keep_names=set(review_codes), warnings=warnings)

    final = out_path.with_suffix(".xlsx")
    wb.save(str(final))
    wb.close()
    return final
