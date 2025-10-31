# services/lead_times/api.py
from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

from .excel_out import (
    inject_and_prune,
    save_review_only_workbook,
    InjectResult,
)
from .sheets import import_and_merge, codes_from_rows
from services.excel_safety import save_workbook_gracefully

import re
from typing import Dict, Tuple, List, Iterable

# replace the existing _LEAD_LINE_RE with this
# If you also parse/insert cutoffs, re-use your cutoff detector here.
# Kept simple: looks for "cutoff" word in the entire cell (case-insensitive).
_CUTOFF_WORD_RE = re.compile(r"cut\s*off", re.I)


# --- Banner + row/column helpers ---
# Exactly: optional leading \n or real newline, then ***CHRISTMAS CUTOFF d/m/yy ***
_CUTOFF_BANNER_RE = re.compile(
    r'(?:\\n|[\r\n])?\*{3}CHRISTMAS CUTOFF\s+\d{1,2}/\d{1,2}/\d{2}\s*\*{3}'
)


_DURATION_FINDER_RE = re.compile(
    r'(?i)'
    r'(\d+(?:\.\d+)?)\s*(?:-|–|to)\s*(\d+(?:\.\d+)?)\s*(weeks?|days?)'
    r'|\b(\d+(?:\.\d+)?)\s*(weeks?|days?)\b'
)


# Minimal header finder: literal "\n" or real newline, then bullet, then "Lead Time:"
_LEAD_HEADER_RE = re.compile(r'(?i)(?:\\n|[\r\n])\s*-\s*Lead\s*Time\s*:\s*')


# Detailed "Lead Time:" line — capture one bracket block, eat any extras
_DETAILED_LEAD_LINE_RE = re.compile(
    r'''(?imx)
        (?P<header> (?:\\n|[\r\n])\s*-\s*Lead\s*Time\s*:\s* )
        (?P<val>
            (?:\d+(?:\.\d+)?\s*(?:-|–|to)\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?)
            \s*(?:weeks?|days?)
        )
        (?P<brackets> \s*\([^)]*\) )?   # keep at most one
        (?: \s*\([^)]*\) )*             # eat any extra bracket blocks from old text
        (?P<trail>\s*)
        (?P<eol> (?:\\n|[\r\n]) | $ )
    '''
)


def _deep_strip_banners(obj):
    """
    Recursively strip banners from any str inside dict/list/tuple structures.
    Safe to run on leads_by_code before inject_and_prune.
    """
    if isinstance(obj, dict):
        return {k: _deep_strip_banners(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_deep_strip_banners(v) for v in obj]
    if isinstance(obj, tuple):
        return tuple(_deep_strip_banners(v) for v in obj)
    if isinstance(obj, str):
        return _strip_only_banner(obj)
    return obj


def _lit(s: str) -> str:
    """Use literal \\n tokens, keep spaces exactly."""
    return (s or "").replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\\n")


def _find_lead_line_span(s: str) -> tuple[int, int, int, str, str] | None:
    if not s:
        return None
    m = _DETAILED_LEAD_LINE_RE.search(s)
    if m:
        start = m.start('header')
        header_end = m.end('header')
        eol_end = m.end('eol')
        header_text = m.group('header')
        # Keep exactly one bracket block if present
        value_text = (m.group('val') or '') + (m.group('brackets') or '')
        return start, header_end, eol_end, header_text, value_text

    # Fallback (rare) – old behavior
    m = _LEAD_HEADER_RE.search(s)
    if not m:
        return None
    start = m.start()
    header_end = m.end()
    m_eol = re.search(r'(?:\\n|[\r\n])', s[header_end:])
    if m_eol:
        v_end = header_end + m_eol.start()
        eol_end = header_end + m_eol.end()
    else:
        v_end = len(s)
        eol_end = len(s)
    header_text = s[start:header_end]
    value_text = s[header_end:v_end]
    return start, header_end, eol_end, header_text, value_text


def _to_weeks(value: float, unit: str) -> float:
    return value / 7.0 if unit.lower().startswith('day') else value


def _normalize_lead_line_spacing_from_template(
    *,
    template_path: Path,
    output_path: Path,
    do_not_show_col_letter: str,
    header_row_1based: int,
    target_col_letter: str,  # "B" for Detailed
    warnings: list[str],
    label: str,
) -> int:
    if not output_path.exists():
        return 0

    wb_out = load_workbook(filename=str(output_path), keep_vba=False, data_only=True)
    wb_tpl = load_workbook(filename=str(template_path), keep_vba=False, data_only=True)

    changed = 0
    tgt_col = _col1(target_col_letter)
    common = set(wb_out.sheetnames) & set(wb_tpl.sheetnames)

    for code in list(common):
        ws_out = wb_out[code]
        ws_tpl = wb_tpl[code]

        r_out = _first_row_do_not_show_false(ws_out, do_not_show_col_letter, header_row_1based)
        r_tpl = _first_row_do_not_show_false(ws_tpl, do_not_show_col_letter, header_row_1based)
        if r_out == -1 or r_tpl == -1:
            continue

        s_out = "" if ws_out.cell(row=r_out, column=tgt_col).value is None else str(ws_out.cell(row=r_out, column=tgt_col).value)
        s_tpl = "" if ws_tpl.cell(row=r_tpl, column=tgt_col).value is None else str(ws_tpl.cell(row=r_tpl, column=tgt_col).value)

        # find spans
        span_out = _find_lead_line_span(s_out)
        if not span_out:
            continue
        start, _header_end, eol_end, header_out, value_out = span_out

        span_tpl = _find_lead_line_span(s_tpl)

        # --- header: keep output header
        header_use = header_out

        # prefix: take it from template up to its Lead Time, but ensure it's banner-free
        pre_tpl = s_tpl[:span_tpl[0]] if span_tpl else ""  # input’s prefix up to its Lead Time:
        pre_tpl = _strip_only_banner(pre_tpl)
        prefix_use = _lit(pre_tpl)

        # --- tail: ALWAYS from the uploaded Detailed input (it can vary per tab);
        # if the input has no Lead Time line, fall back to the output tail
        raw_tail = s_tpl[span_tpl[2]:] if span_tpl else s_out[eol_end:]
        tail_use = _strip_only_banner(raw_tail)

        # rebuild ONLY the lead-time line (value_out stays as written by injector)
        new_segment = _lit(header_use) + value_out + "\\n"
        fixed = prefix_use + new_segment + _lit(tail_use)

        if fixed != s_out:
            ws_out.cell(row=r_out, column=tgt_col).value = fixed
            changed += 1

    if changed:
        has_real_data = save_workbook_gracefully(wb_out, str(output_path))
        if not has_real_data:
            warnings.append("No data matched your filters — exported a placeholder workbook.")

        warnings.append(f"[FORMAT] {label}: normalized Lead Time line spacing on {changed} tab(s).")
    return changed


def _strip_only_banner(s: str | None) -> str:
    """Remove only the exact CHRISTMAS CUTOFF banner instances; leave everything else intact."""
    return _CUTOFF_BANNER_RE.sub("", "" if s is None else str(s))


def _banner(date_str: str) -> str:
    return f"***CHRISTMAS CUTOFF {date_str}***"


def _is_trueish(v) -> bool:
    """Interpret TRUE-ish values in the Do Not Show? column."""
    if v is True or v == 1:
        return True
    s = ("" if v is None else str(v)).strip().lower()
    return s in {"true", "t", "yes", "y", "1"}


def _col1(a1: str) -> int:
    """1-based column index from a letter like 'F'."""
    return _col_idx(a1) + 1


def _first_row_do_not_show_false(ws, do_not_show_col_letter: str, header_row_1based: int) -> int:
    """
    Find the first row > header_row_1based where Do Not Show? is FALSE (or empty/blank).
    Returns -1 if none found.
    """
    c = _col1(do_not_show_col_letter)
    start = (header_row_1based or 1) + 1
    for r in range(start, ws.max_row + 1):
        val = ws.cell(row=r, column=c).value
        if not _is_trueish(val):
            return r
    return -1


def _apply_banner_detailed_text(old: str, cutoff: str | None) -> str:
    """
    Detailed policy:
      1) Remove ONLY our exact banner wherever it sits.
      2) If cutoff present, prepend '\n***CHRISTMAS CUTOFF d/m/yy ***' to the start.
      3) Do not touch any other spacing/newlines in the body.
    """
    body = _strip_only_banner(old or "")
    cutoff = (cutoff or "").strip()
    if not cutoff:
        return body
    return f"\\n***CHRISTMAS CUTOFF {cutoff} ***{body}"


def _apply_banner_summary_text(old: str, cutoff: str | None) -> str:
    base = _strip_only_banner(old or "")
    cutoff = (cutoff or "").strip()
    if not cutoff:
        return base
    sep = "" if (len(base) == 0 or base[-1].isspace()) else " "
    return f"{base}{sep}***CHRISTMAS CUTOFF {cutoff} ***"


def _canon(s: str) -> str:
    """
    Canonicalize for comparison: normalize whitespace and case.
    """
    if s is None:
        return ""
    s = s.replace("\xa0", " ")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"[ \t]+\n", "\n", s)
    return s.strip().lower()


def _apply_banners_to_workbook(
    *,
    output_path: Path,
    cutoffs_by_code: dict[str, dict],
    warnings: list[str],
    label: str,
    kind: str,  # "detailed" or "summary"
    do_not_show_col_letter: str,
    header_row_1based: int,
    target_col_letter: str,  # insertion col for that kind (e.g., "B" or "C")
) -> int:
    """
    For each sheet:
      - find the first row where Do Not Show? is FALSE
      - edit target cell (Detailed: Before Answer, Summary: After Answer)
      - remove old banner; insert new in canonical place if cutoff exists
    Returns: number of tabs modified (where cell content actually changed).
    """
    if not output_path.exists():
        return 0

    wb = load_workbook(filename=str(output_path), keep_vba=False, data_only=True)
    changed = 0
    target_col = _col1(target_col_letter)

    for code in list(wb.sheetnames):
        ws = wb[code]
        r = _first_row_do_not_show_false(ws, do_not_show_col_letter, header_row_1based)
        if r == -1:
            continue

        cell = ws.cell(row=r, column=target_col)
        old = "" if cell.value is None else str(cell.value)

        cutoff = ""
        if cutoffs_by_code and code in cutoffs_by_code:
            cutoff = str(cutoffs_by_code[code].get("cutoff", "")).strip()

        new = _apply_banner_detailed_text(old, cutoff) if kind == "detailed" else _apply_banner_summary_text(old, cutoff)

        if new != old:
            cell.value = new
            changed += 1

    if changed:
        has_real_data = save_workbook_gracefully(wb, str(output_path))
        if not has_real_data:
            warnings.append("No data matched your filters — exported a placeholder workbook.")

        warnings.append(f"[CUTOFF] {label}: normalized banner on {changed} tab(s).")

    return changed


def _prune_unchanged_tabs_cell_based(
    *,
    template_path: Path,
    output_path: Path,
    warnings: list[str],
    label: str,
    do_not_show_col_letter: str,
    header_row_1based: int,
    target_col_letter: str,  # "B" for detailed, "C" for summary (from config)
) -> list[str]:
    """
    Remove a tab if the target cell text (first row where Do Not Show? is FALSE, target column)
    is EXACTLY the same as in the template workbook. Any character difference => keep.
    """
    pruned: list[str] = []
    if not output_path.exists():
        return pruned

    wb_out = load_workbook(filename=str(output_path), keep_vba=False, data_only=True)
    wb_tpl = load_workbook(filename=str(template_path), keep_vba=False, data_only=True)

    target_col = _col1(target_col_letter)
    common = set(wb_out.sheetnames) & set(wb_tpl.sheetnames)

    for code in list(common):
        ws_out = wb_out[code]
        ws_tpl = wb_tpl[code]

        r_out = _first_row_do_not_show_false(ws_out, do_not_show_col_letter, header_row_1based)
        r_tpl = _first_row_do_not_show_false(ws_tpl, do_not_show_col_letter, header_row_1based)
        if r_out == -1 or r_tpl == -1:
            # If either workbook can't find a display row, don't prune on this heuristic
            continue

        out_text = "" if ws_out.cell(row=r_out, column=target_col).value is None else str(ws_out.cell(row=r_out, column=target_col).value)
        tpl_text = "" if ws_tpl.cell(row=r_tpl, column=target_col).value is None else str(ws_tpl.cell(row=r_tpl, column=target_col).value)

        if out_text == tpl_text:
            wb_out.remove(ws_out)
            pruned.append(code)

    if pruned:
        has_real_data = save_workbook_gracefully(wb_out, str(output_path))
        if not has_real_data:
            warnings.append("No data matched your filters — exported a placeholder workbook.")

        samp = ", ".join(sorted(pruned)[:5]) + ("…" if len(pruned) > 5 else "")
        warnings.append(f"[PRUNE] {label}: removed {len(pruned)} unchanged tab(s) (e.g., {samp}).")

    return pruned


def _need(d: dict, *path):
    cur = d
    trail = []
    for k in path:
        trail.append(k)
        if not isinstance(cur, dict) or k not in cur:
            raise ValueError(f"Config missing `{'/'.join(trail)}`")
        cur = cur[k]
    return cur


def _opt(d: dict, default, *path):
    try:
        return _need(d, *path)
    except ValueError:
        return default


def _valid_codes_from_templates(*paths: str) -> set[str]:
    """Union of sheet names from the uploaded templates (exact, case-sensitive)."""
    codes: set[str] = set()
    for p in paths:
        wb = load_workbook(filename=p, keep_vba=False, read_only=True, data_only=True)
        codes.update(wb.sheetnames)
    return codes


def _col_idx(a1: str) -> int:
    """Convert a column letter (e.g. 'C') to a 0-based index."""
    a1 = a1.strip().upper()
    n = 0
    for ch in a1:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _cutoff_attach_note(
    warnings: list[str],
    store_name: str,
    by_product_html: list[tuple[str, str]],
    product_to_codes: dict[str, set[str]],
    cutoff_rows: dict[str, dict],
) -> None:
    """
    Report how many products actually received a CHRISTMAS CUTOFF banner.
    A product is 'attached' if ANY of its mapped buz codes has a cutoff date.
    """
    cutoff_codes_with_date = {
        code
        for code, rec in cutoff_rows.items()
        if str(rec.get("cutoff", "")).strip()
    }
    products = [str(p).strip() for p, _ in by_product_html if str(p).strip()]
    attached = [
        p for p in products
        if product_to_codes.get(p, set()) & cutoff_codes_with_date
    ]
    if attached:
        sample = ", ".join(sorted(attached)[:5]) + ("…" if len(attached) > 5 else "")
        warnings.append(
            f"[{store_name}] CHRISTMAS CUTOFF appended to {len(attached)} product(s) (e.g., {sample})."
        )
    else:
        if cutoff_codes_with_date:
            cutoff_products = {
                str(rec.get("product", "")).strip()
                for code, rec in cutoff_rows.items()
                if code in cutoff_codes_with_date and str(rec.get("product", "")).strip()
            }
            missing_in_leads = sorted(set(cutoff_products) - set(products))
            if missing_in_leads:
                sample = ", ".join(missing_in_leads[:3]) + (
                    "…" if len(missing_in_leads) > 3 else ""
                )
                warnings.append(
                    f"[{store_name}] No banners appended. "
                    f"{len(missing_in_leads)} cutoff product(s) aren’t present in this store’s "
                    f"Lead Times mapping (e.g., {sample})."
                )


def run_publish(
    *,
    gsheets_service,
    lead_times_cfg: dict,
    detailed_template_path: str,
    summary_template_path: str,
    save_dir: str,
    scope: str,
) -> dict:
    """Read Sheets, validate/merge, generate HTML + Excel files (or review-only)."""

    if not isinstance(lead_times_cfg, dict) or "lead_times_ss" not in lead_times_cfg:
        raise ValueError("Missing or invalid config: get_cfg('lead_times') did not return expected structure")

    # 1) Read Sheets
    lt_block = lead_times_cfg["lead_times_ss"]
    co_block = lead_times_cfg["cutoffs"]
    lt_id = lt_block["sheet_id"]
    co_id = co_block["sheet_id"]
    lt_hdr = int(lt_block.get("header_row", 1))  # 1-based
    co_hdr = int(co_block.get("header_row", 1))  # 1-based

    t_lead = lt_block["tabs"][scope]
    t_cut = co_block["tabs"][scope]

    def a1(tab: str, cols: str = "A:Z") -> str:
        return f"{tab}!{cols}"

    lead_rows = gsheets_service.fetch_sheet_data(lt_id, a1(t_lead))
    cut_rows = gsheets_service.fetch_sheet_data(co_id, a1(t_cut))

    sa_email = getattr(gsheets_service, "service_account_email", lambda: "<unknown>")()
    if not lead_rows:
        raise ValueError(
            f"Could not read Lead Times sheet(s). Share {lt_id} with {sa_email} and re-run."
        )

    # Drop configured header rows (1-based)
    def strip_headers(rows: list[list[str]], header_row_1based: int) -> list[list[str]]:
        if not rows:
            return rows
        if header_row_1based is None or header_row_1based < 1:
            return rows
        return rows[header_row_1based:]

    lead_rows = strip_headers(lead_rows, lt_hdr)
    cut_rows = strip_headers(cut_rows, co_hdr)

    # Columns
    lead_cols = lead_times_cfg["columns"]["lead_times_ss"]
    cut_cols = lead_times_cfg["columns"]["cutoff"]
    map_i = _col_idx(lead_cols["mapping"])
    cut_map_i = _col_idx(cut_cols["mapping"])

    # Strict: workbook tabs define the valid code list
    valid_tabs = _valid_codes_from_templates(
        detailed_template_path, summary_template_path
    )
    lead_codes = codes_from_rows(lead_rows, map_i)
    cut_codes = codes_from_rows(cut_rows, cut_map_i)

    unknown_lead = lead_codes - valid_tabs
    unknown_cut = cut_codes - valid_tabs

    def _pv(s: set[str]) -> str:
        return "—" if not s else (", ".join(sorted(s)[:12]) + (f", +{len(s) - 12} more" if len(s) > 12 else ""))

    scope_codes = set(lead_codes)

    # If any relevant unknowns exist, raise with scope-aware sections
    has_any = bool(unknown_lead) or bool(unknown_cut)
    if has_any:
        parts: list[str] = [
            "<strong>Unknown codes</strong> — present in Google Sheets but not found as tabs "
            "in the uploaded templates (tabs define the valid list)."
        ]
        if unknown_lead:
            parts.append(
                f"<div class='mt-1'><small><strong>Leads:</strong> <code>{_pv(unknown_lead)}</code></small></div>"
            )
        if unknown_cut:
            parts.append(
                f"<div class='mt-1'><small><strong>Cutoffs:</strong> <code>{_pv(unknown_cut)}</code></small></div>"
            )

        from markupsafe import Markup
        raise ValueError(Markup("".join(parts)))

    warnings: list[str] = []

    # 2) Merge + HTML
    ir = import_and_merge(
        lead_rows=lead_rows,
        cutoff_rows=cut_rows,
        lead_cols=lead_cols,
        cutoff_cols=cut_cols,
        scope=scope
    )

    from services.lead_times.html_out import build_pasteable_html_direct_cutoffs

    html_out = build_pasteable_html_direct_cutoffs(
        ir.by_product_html,
        product_to_cutoff=ir.product_to_cutoff,
    )

    _cutoff_attach_note(
        warnings,
        scope,
        ir.by_product_html,
        ir.product_to_codes,
        ir.cutoff_rows,
    )

    # 3) Excel generation
    outdir = Path(save_dir)
    outdir.mkdir(parents=True, exist_ok=True)

    warnings.append(f"[DEBUG] Output dir resolved to: {outdir}")

    patterns = lead_times_cfg.get("filename_patterns", {})
    ins_cols = lead_times_cfg["insertion_columns"]
    anchor_col_letter = lead_times_cfg["columns"]["do_not_show_header"]  # e.g. "F"
    anchor_header_row = lead_times_cfg["columns"].get("do_not_show_header_row", 2)

    # --- timestamps for filename tokens ---
    _now = datetime.now()
    _tag_yyyyymmdd = _now.strftime("%Y%m%d")
    _tag_yymmdd = _now.strftime("%y%m%d")
    _tag_hhmmss = _now.strftime("%H%M%S")

    def outname(key: str, default: str) -> Path:
        pat = patterns.get(key, default)
        return outdir / (
            pat
            .replace("{YYYYMMDD}", _tag_yyyyymmdd)
            .replace("{YYMMDD}", _tag_yymmdd)
            .replace("{HHMMSS}", _tag_hhmmss)
        )

    # Text templates/regex from config (with safe defaults)
    tmpl = lead_times_cfg.get("templates", {})

    # IMPORTANT: Avoid formatting drift. Let excel_out do minimal edits.
    # Keep this template empty by default (we'll prune unchanged tabs post-write).
    detailed_prefix_template = tmpl.get("detailed_prefix_template", "")

    files: Dict[str, str] = {}
    review_detailed: set[str] = set()
    review_summary: set[str] = set()

    # Detailed
    leads_clean = _deep_strip_banners(ir.lead_rows)

    res: InjectResult = inject_and_prune(
        template_path=Path(detailed_template_path),
        out_path=outname("detailed", "Quote_Detailed_{YYYYMMDD}.xlsx"),
        store_name=scope,
        leads_by_code=leads_clean,
        insertion_col_letter=ins_cols["detailed"],   # "B"
        anchor_col_letter=anchor_col_letter,         # "F"
        anchor_header_row=anchor_header_row,         # 2
        control_codes=ir.control_codes,
        warnings=warnings,
        workbook_kind="Detailed",
        detailed_prefix_template=detailed_prefix_template,
    )

    _normalize_lead_line_spacing_from_template(
        template_path=Path(detailed_template_path),
        output_path=Path(res.saved_path),
        do_not_show_col_letter=anchor_col_letter,
        header_row_1based=anchor_header_row,
        target_col_letter=ins_cols["detailed"],  # "B"
        warnings=warnings,
        label="Detailed",
    )

    _apply_banners_to_workbook(
        output_path=Path(res.saved_path),
        cutoffs_by_code=ir.cutoff_rows,
        warnings=warnings,
        label="Detailed",
        kind="detailed",
        do_not_show_col_letter=anchor_col_letter,
        header_row_1based=anchor_header_row,
        target_col_letter=ins_cols["detailed"],
    )

    _prune_unchanged_tabs_cell_based(
        template_path=Path(detailed_template_path),
        output_path=Path(res.saved_path),
        warnings=warnings,
        label="Detailed",
        do_not_show_col_letter=anchor_col_letter,
        header_row_1based=anchor_header_row,
        target_col_letter=ins_cols["detailed"],
    )

    review_detailed |= set(res.review_codes)
    files["detailed"] = os.path.basename(str(res.saved_path))

    # Summary
    leads_clean = _deep_strip_banners(ir.lead_rows)

    res: InjectResult = inject_and_prune(
        template_path=Path(summary_template_path),
        out_path=outname("summary", "Quote_Summary_{YYYYMMDD}.xlsx"),
        store_name=scope,
        leads_by_code=leads_clean,
        insertion_col_letter=ins_cols["summary"],    # "C"
        anchor_col_letter=anchor_col_letter,         # "F"
        anchor_header_row=anchor_header_row,         # 2
        control_codes=ir.control_codes,
        warnings=warnings,
        workbook_kind="Summary",
    )

    _apply_banners_to_workbook(
        output_path=Path(res.saved_path),
        cutoffs_by_code=ir.cutoff_rows,
        warnings=warnings,
        label="Summary",
        kind="summary",
        do_not_show_col_letter=anchor_col_letter,
        header_row_1based=anchor_header_row,
        target_col_letter=ins_cols["summary"],
    )

    _prune_unchanged_tabs_cell_based(
        template_path=Path(summary_template_path),
        output_path=Path(res.saved_path),
        warnings=warnings,
        label="Summary",
        do_not_show_col_letter=anchor_col_letter,
        header_row_1based=anchor_header_row,
        target_col_letter=ins_cols["summary"],
    )

    review_summary |= set(res.review_codes)
    files["summary"] = os.path.basename(str(res.saved_path))

    # If anything needs review, emit consolidated review-only workbooks and suppress the 4 normal links.
    review_files: Dict[str, str] = {}
    if review_detailed:
        p = save_review_only_workbook(
            template_path=Path(detailed_template_path),
            out_path=outname("detailed_review", "Quote_Detailed_REVIEW_{YYYYMMDD}.xlsx"),
            review_codes=review_detailed,
            warnings=warnings,
        )
        review_files["detailed_review"] = os.path.basename(str(p))
        warnings.append(
            f"[REVIEW] Detailed: {len(review_detailed)} tab(s) require attention — providing review-only workbook."
        )
    if review_summary:
        p = save_review_only_workbook(
            template_path=Path(summary_template_path),
            out_path=outname("summary_review", "Quote_Summary_REVIEW_{YYYYMMDD}.xlsx"),
            review_codes=review_summary,
            warnings=warnings,
        )
        review_files["summary_review"] = os.path.basename(str(p))
        warnings.append(
            f"[REVIEW] Summary: {len(review_summary)} tab(s) require attention — providing review-only workbook."
        )

    # Prefer review-only links if any exist
    final_files = review_files if review_files else files

    return {
        "warnings": warnings,
        "html": html_out,
        "files": final_files,  # basenames; your download route serves from save_dir
    }
