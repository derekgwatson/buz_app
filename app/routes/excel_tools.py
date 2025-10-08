# excel_tools.py
from __future__ import annotations

import os
import re
import time
import uuid
import tempfile
from datetime import datetime
from io import BytesIO
from typing import Iterable, List, Tuple, Dict, Any

from flask import (
    Blueprint,
    render_template,
    request,
    send_file,
    flash,
    redirect,
    url_for,
    after_this_request,
)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

excel_tools_bp = Blueprint("excel_tools", __name__, url_prefix="/excel")

# --------------------------
# Utilities / helpers
# --------------------------

STAR_TRAIL_RE = re.compile(r"\s*\*$")

# Simple temp-file cache for downloads
_DOWNLOAD_CACHE: Dict[str, Dict[str, Any]] = {}
_CACHE_TTL_SECONDS = 30 * 60  # 30 minutes


def _cleanup_cache() -> None:
    now = time.time()
    stale = [tok for tok, meta in _DOWNLOAD_CACHE.items() if now - meta["ts"] > _CACHE_TTL_SECONDS]
    for tok in stale:
        try:
            path = _DOWNLOAD_CACHE[tok]["path"]
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            pass
        _DOWNLOAD_CACHE.pop(tok, None)


def _cache_file(data: bytes, download_name: str, mimetype: str) -> str:
    _cleanup_cache()
    token = uuid.uuid4().hex
    tmp_dir = tempfile.gettempdir()
    tmp_path = os.path.join(tmp_dir, f"excel-filter-{token}.bin")
    with open(tmp_path, "wb") as f:
        f.write(data)
    _DOWNLOAD_CACHE[token] = {
        "path": tmp_path,
        "name": download_name,
        "mimetype": mimetype,
        "ts": time.time(),
    }
    return token


def _pop_cached(token: str) -> Dict[str, Any] | None:
    meta = _DOWNLOAD_CACHE.pop(token, None)
    return meta


def _normalize(s: object, case_sensitive: bool) -> str:
    if s is None:
        return ""
    s = str(s)
    return s if case_sensitive else s.casefold()


def _terms_from_textarea(raw: str, case_sensitive: bool) -> list[str]:
    terms = [ln.strip() for ln in (raw or "").splitlines() if ln.strip()]
    if not case_sensitive:
        terms = [t.casefold() for t in terms]
    seen = set()
    out: list[str] = []
    for t in terms:
        if t not in seen:
            out.append(t)
            seen.add(t)
    return out


def _excel_letters_inclusive(start_col: str, end_col: str) -> List[str]:
    s = column_index_from_string(start_col)
    e = column_index_from_string(end_col)
    return [get_column_letter(i) for i in range(s, e + 1)]


def _hide_inventoryitems_columns(ws) -> None:
    """For InventoryItems files, hide J-Z, AD-AG, and AL."""
    to_hide = set()
    to_hide.update(_excel_letters_inclusive("J", "Z"))
    to_hide.update(_excel_letters_inclusive("AD", "AG"))
    to_hide.add("AL")
    for col in to_hide:
        ws.column_dimensions[col].hidden = True


def _autofit_sheet_columns(ws, skip_hidden: bool = True) -> None:
    """Approximate Excel autofit via max text length per (visible) column."""
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    max_len_by_col: dict[int, int] = {}

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        if skip_hidden and ws.column_dimensions.get(col_letter) and ws.column_dimensions[col_letter].hidden:
            continue
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is None:
                continue
            for part in str(val).splitlines():
                if len(part) > max_len:
                    max_len = len(part)
        if max_len > 0:
            max_len_by_col[col_idx] = max_len

    for col_idx, max_len in max_len_by_col.items():
        col_letter = get_column_letter(col_idx)
        width = min(100, max(6, int(max_len * 1.1) + 2))
        ws.column_dimensions[col_letter].width = width


def _strip_trailing_asterisk_in_b2_c2(ws) -> None:
    for addr in ("B2", "C2"):
        cell = ws[addr]
        val = cell.value
        if isinstance(val, str):
            new_val = STAR_TRAIL_RE.sub("", val)
            if new_val != val:
                cell.value = new_val


def _compress_indices_to_descending_ranges(indices: List[int]) -> List[Tuple[int, int]]:
    """Convert row indices to bottom-up (start, count) ranges for delete_rows."""
    if not indices:
        return []
    dedup_sorted = sorted(set(indices))
    ranges: List[Tuple[int, int]] = []
    start = prev = dedup_sorted[0]
    for i in dedup_sorted[1:]:
        if i == prev + 1:
            prev = i
        else:
            ranges.append((start, prev))
            start = prev = i
    ranges.append((start, prev))
    return [(s, e - s + 1) for (s, e) in reversed(ranges)]


def _scan_sheet(
    ws,
    terms: Iterable[str],
    case_sensitive: bool,
    search_sheet_names: bool,
    header_rows: int,
) -> Dict[str, Any]:
    """
    Scan a worksheet and summarize matches.
    Returns:
      {
        'sheet_name_match': bool,
        'content_match': bool,
        'matched_terms': set[str],
        'matching_data_rows': List[int],   # 1-based indices >= header_rows+1
        'nonmatching_data_rows': List[int],
        'pre_rows': int
      }
    """
    matched_terms: set[str] = set()
    sheet_name_match = False
    if search_sheet_names:
        title_norm = _normalize(ws.title, case_sensitive)
        for t in terms:
            if t in title_norm:
                matched_terms.add(t)
                sheet_name_match = True

    pre_rows = ws.max_row or 0
    start_row = max(1, header_rows + 1)
    content_match = False
    matching_data_rows: List[int] = []
    nonmatching_data_rows: List[int] = []

    # Iterate all rows (to allow headers to trigger content_match)
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_has_match = False
        # test each cell
        for cell_val in row:
            val = _normalize(cell_val, case_sensitive)
            if not val:
                continue
            found_any = False
            for t in terms:
                if t in val:
                    matched_terms.add(t)
                    found_any = True
            if found_any:
                row_has_match = True
                content_match = True
        # record row-level membership for data rows only
        if r_idx >= start_row:
            if row_has_match:
                matching_data_rows.append(r_idx)
            else:
                nonmatching_data_rows.append(r_idx)

    return {
        "sheet_name_match": sheet_name_match,
        "content_match": content_match,
        "matched_terms": matched_terms,
        "matching_data_rows": matching_data_rows,
        "nonmatching_data_rows": nonmatching_data_rows,
        "pre_rows": pre_rows,
    }


# --------------------------
# Routes
# --------------------------

@excel_tools_bp.route("/filter-tabs", methods=["GET", "POST"])
def filter_tabs():
    if request.method == "GET":
        return render_template("filter_tabs.html")

    # POST
    uploaded = request.files.get("workbook")
    if not uploaded or uploaded.filename == "":
        flash("Please choose an .xlsx or .xlsm file.")
        return redirect(request.url)

    case_sensitive = bool(request.form.get("case_sensitive"))
    search_sheet_names = bool(request.form.get("search_sheet_names"))
    prune_rows = bool(request.form.get("prune_rows"))
    try:
        header_rows = int(request.form.get("header_rows", "1"))
        if header_rows < 0:
            raise ValueError
    except ValueError:
        flash("Header rows must be a non-negative integer.")
        return redirect(request.url)

    terms = _terms_from_textarea(request.form.get("terms", ""), case_sensitive)
    if not terms:
        flash("Please enter at least one search term (one per line).")
        return redirect(request.url)

    lower_name = uploaded.filename.lower()
    is_xlsm = lower_name.endswith(".xlsm")
    is_inventory_items = "inventoryitems" in lower_name

    if not (lower_name.endswith(".xlsx") or is_xlsm):
        flash("Unsupported file type. Please upload an .xlsx or .xlsm workbook.")
        return redirect(request.url)

    try:
        wb = load_workbook(uploaded, data_only=True, keep_vba=is_xlsm)
    except Exception as exc:
        flash(f"Could not open workbook: {exc}")
        return redirect(request.url)

    results: Dict[str, Any] = {
        "source_filename": uploaded.filename,
        "is_inventory_items": is_inventory_items,
        "prune_rows": prune_rows,
        "header_rows": header_rows,
        "sheets": [],
        "kept_count": 0,
    }

    # First pass: scan every sheet and decide removal / pruning sets
    scanned_by_title: Dict[str, Dict[str, Any]] = {}
    to_delete = []

    for ws in list(wb.worksheets):
        scan = _scan_sheet(
            ws,
            terms=terms,
            case_sensitive=case_sensitive,
            search_sheet_names=search_sheet_names,
            header_rows=header_rows,
        )
        scanned_by_title[ws.title] = scan

        any_match = scan["content_match"] or scan["sheet_name_match"]
        if not any_match:
            to_delete.append(ws)

    # Remove unmatched sheets entirely
    for ws in to_delete:
        wb.remove(ws)

    # Second pass: for remaining sheets, optionally prune rows + apply formatting
    for ws in wb.worksheets:
        if ws.title == "No matches":
            # Add to results and skip formatting/pruning
            results["sheets"].append({
                "name": ws.title,
                "status": "kept",
                "matched_by": "none",
                "matched_terms_preview": "",
                "rows_before": ws.max_row,
                "rows_after": ws.max_row,
                "rows_removed": 0,
                "matching_rows": 0,
            })
            results["kept_count"] += 1
            continue

        scan = scanned_by_title.get(ws.title)
        if scan is None:
            # Shouldn't happen; safe fallback
            scan = {
                "sheet_name_match": False,
                "content_match": False,
                "matched_terms": set(),
                "matching_data_rows": [],
                "nonmatching_data_rows": [],
                "pre_rows": ws.max_row,
            }

        rows_before = scan["pre_rows"]
        rows_removed = 0

        if prune_rows and scan["nonmatching_data_rows"]:
            for start, count in _compress_indices_to_descending_ranges(scan["nonmatching_data_rows"]):
                ws.delete_rows(start, amount=count)
                rows_removed += count

        # InventoryItems custom formatting
        if is_inventory_items:
            _strip_trailing_asterisk_in_b2_c2(ws)
            _hide_inventoryitems_columns(ws)
            _autofit_sheet_columns(ws, skip_hidden=True)

        matched_by = []
        if scan["content_match"]:
            matched_by.append("content")
        if scan["sheet_name_match"]:
            matched_by.append("sheet name")
        matched_by_str = " + ".join(matched_by) if matched_by else "none"

        matched_terms_preview = ", ".join(sorted(list(scan["matched_terms"])))[:200]
        rows_after = rows_before - rows_removed if prune_rows else rows_before
        matching_rows = len(scan["matching_data_rows"])

        results["sheets"].append({
            "name": ws.title,
            "status": "kept",
            "matched_by": matched_by_str,
            "matched_terms_preview": matched_terms_preview,
            "rows_before": rows_before,
            "rows_after": rows_after,
            "rows_removed": rows_removed,
            "matching_rows": matching_rows,
        })
        results["kept_count"] += 1

    if results["kept_count"] == 0:
        # Nothing kept => nothing found. Don't build or cache a file.
        return render_template(
            "filter_tabs.html",
            results=results,
            download_token=None,
            download_name=None,
        )

    # Build file in memory
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    base = uploaded.filename.rsplit(".", 1)[0]
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    if is_xlsm:
        download_name = f"{base}.filtered.{ts}.xlsm"
        mimetype = "application/vnd.ms-excel.sheet.macroEnabled.12"
    else:
        download_name = f"{base}.filtered.{ts}.xlsx"
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    token = _cache_file(out.getvalue(), download_name, mimetype)

    return render_template(
        "filter_tabs.html",
        results=results,
        download_token=token,
        download_name=download_name,
    )


@excel_tools_bp.get("/download/<token>")
def download_filtered(token: str):
    meta = _pop_cached(token)
    if not meta or not os.path.exists(meta["path"]):
        flash("That download link has expired. Please re-run the filter.")
        return redirect(url_for("excel_tools.filter_tabs"))

    path = meta["path"]
    download_name = meta["name"]
    mimetype = meta["mimetype"]

    @after_this_request
    def _cleanup(response):
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            pass
        return response

    return send_file(
        path,
        as_attachment=True,
        download_name=download_name,
        mimetype=mimetype,
    )
