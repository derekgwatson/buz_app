# app/routes/max_discount_review.py
from __future__ import annotations

import asyncio
import logging
import threading
import uuid
from pathlib import Path
from flask import Blueprint, current_app, render_template, request, jsonify, url_for
from copy import copy

from services.auth import auth
from services.job_service import create_job, update_job, get_job
from services.database import create_db_manager

logger = logging.getLogger(__name__)

max_discount_review_bp = Blueprint("max_discount_review", __name__, url_prefix="/tools/max-discount-review")


@max_discount_review_bp.route("/", methods=["GET"])
@auth.login_required
def index():
    """Max discount review entry point"""
    return render_template("max_discount_review.html")


@max_discount_review_bp.route("/start-review", methods=["POST"])
@auth.login_required
def start_review():
    """
    Start max discount review across selected orgs.

    Expects:
        orgs: (required) List of org keys to process
        headless: (optional) Run browser in headless mode (default: true)
    """
    headless = request.form.get("headless", "true").lower() in ("true", "1", "yes")

    # Get selected orgs
    selected_orgs = request.form.getlist("orgs")
    if not selected_orgs:
        return jsonify({"error": "At least one organization must be selected"}), 400

    # Validate org keys
    valid_orgs = ['canberra', 'tweed', 'bay', 'shoalhaven', 'wagga']
    invalid_orgs = [org for org in selected_orgs if org not in valid_orgs]
    if invalid_orgs:
        return jsonify({"error": f"Invalid organizations: {', '.join(invalid_orgs)}"}), 400

    # Create job
    job_id = uuid.uuid4().hex
    db_path = current_app.config["database"]

    db = create_db_manager(db_path)
    create_job(job_id, db=db)

    # Create output directory for this job
    export_root = Path(current_app.config.get("EXPORT_ROOT", "exports"))
    output_dir = export_root / "max_discount_review" / job_id
    output_dir.mkdir(parents=True, exist_ok=True)

    def run_review():
        """Background thread to run max discount review"""
        db = create_db_manager(db_path)
        try:
            org_names = ', '.join([org.title() for org in selected_orgs])
            update_job(job_id, 5, f"Starting max discount review for: {org_names}", db=db)

            # Import and run the review
            from services.buz_max_discount_review import review_max_discounts_all_orgs
            from services.max_discount_comparison import build_max_discount_comparison

            def job_callback(pct: int, message: str):
                """Update job progress"""
                update_job(job_id, pct, message, db=db)

            # Run async function in new event loop
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                result = loop.run_until_complete(
                    review_max_discounts_all_orgs(
                        output_dir=output_dir,
                        headless=headless,
                        selected_orgs=selected_orgs,
                        job_update_callback=job_callback
                    )
                )
            finally:
                loop.close()

            # Build comparison
            update_job(job_id, 90, "Building comparison table", db=db)
            comparison = build_max_discount_comparison(result)

            # Build summary
            summary_lines = [
                f"✓ Downloaded and parsed inventory groups from {len(result.orgs)} orgs",
                f"✓ Found {comparison.to_dict()['summary']['total_products']} unique products",
                f"  - {comparison.to_dict()['summary']['matched_by_code']} matched by code",
                f"  - {comparison.to_dict()['summary']['matched_by_description']} matched by description"
            ]

            for org in result.orgs:
                summary_lines.append(f"  - {org.org_name}: {len(org.inventory_groups)} groups")

            summary = "\n".join(summary_lines)

            update_job(
                job_id,
                pct=100,
                message="Review completed successfully",
                done=True,
                result={
                    "summary": summary,
                    "comparison": comparison.to_dict(),
                    "raw_result": result.to_dict(),
                    "output_dir": str(output_dir)
                },
                db=db
            )

        except Exception as e:
            logger.exception(f"Max discount review failed")
            update_job(
                job_id,
                pct=0,
                message=f"Error: {str(e)}",
                error=str(e),
                done=True,
                result={"error": str(e)},
                db=db
            )

    # Start background thread
    thread = threading.Thread(target=run_review, daemon=True)
    thread.start()

    return jsonify({"job_id": job_id})


@max_discount_review_bp.route("/job/<job_id>", methods=["GET"])
@auth.login_required
def job_status(job_id):
    """Get job status"""
    db_path = current_app.config["database"]
    db = create_db_manager(db_path)
    job = get_job(job_id, db=db)

    if not job:
        return jsonify({"error": "Job not found"}), 404

    return jsonify(job)


@max_discount_review_bp.route("/generate-upload", methods=["POST"])
@auth.login_required
def generate_upload():
    """
    Generate upload Excel files for changed discounts.

    Expects JSON:
        changes: Dict of org_name -> list of changes
        raw_result: Raw result data with file paths
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    data = request.get_json()
    changes_by_org = data.get("changes", {})
    raw_result = data.get("raw_result", {})

    if not changes_by_org:
        return jsonify({"error": "No changes provided"}), 400

    export_root = Path(current_app.config.get("EXPORT_ROOT", "exports"))
    upload_dir = export_root / "max_discount_uploads" / uuid.uuid4().hex
    upload_dir.mkdir(parents=True, exist_ok=True)

    generated_files = []

    try:
        # Find the original files from raw_result
        orgs_data = {org['org_name']: org for org in raw_result.get('orgs', [])}

        for org_name, changes_list in changes_by_org.items():
            if org_name not in orgs_data:
                logger.warning(f"Org {org_name} not found in raw result")
                continue

            org_data = orgs_data[org_name]
            source_file = org_data.get('file_path')

            if not source_file or not Path(source_file).exists():
                logger.error(f"Source file not found for {org_name}: {source_file}")
                continue

            # Load the original Excel file
            wb = openpyxl.load_workbook(source_file)
            ws = wb["Inventory Groups"]

            # Create a new workbook for upload
            upload_wb = openpyxl.Workbook()
            upload_ws = upload_wb.active
            upload_ws.title = "Inventory Groups"

            # Helper function to safely copy cell value (handle text starting with =)
            def safe_copy_value(value):
                """
                Safely copy a cell value, handling text that starts with =
                which Excel would interpret as a formula.
                """
                if value is None:
                    return None
                # If the value is a string starting with =, prefix with ' to make it literal text
                if isinstance(value, str) and value.startswith('='):
                    return "'" + value
                return value

            # Copy header row (row 1)
            # Only copy up to column AE (31) which is the Operation column
            for col_idx in range(1, 32):
                source_cell = ws.cell(row=1, column=col_idx)
                target_cell = upload_ws.cell(row=1, column=col_idx)

                # Copy cell value and basic properties
                target_cell.value = safe_copy_value(source_cell.value)
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)

            # Track which product codes have been changed
            changed_codes = {change['productCode'] for change in changes_list}

            # Find and copy rows for changed products
            upload_row = 2
            for row_idx in range(2, ws.max_row + 1):
                code = ws.cell(row=row_idx, column=3).value  # Column C = Code

                if code in changed_codes:
                    # Find the change for this product
                    change = next((c for c in changes_list if c['productCode'] == code), None)

                    if change:
                        # Copy the entire row (up to column AE = 31)
                        for col_idx in range(1, 32):
                            source_cell = ws.cell(row=row_idx, column=col_idx)
                            target_cell = upload_ws.cell(row=upload_row, column=col_idx)

                            # Copy cell value, handling text that starts with =
                            target_cell.value = safe_copy_value(source_cell.value)

                            # Copy cell styling if present
                            if source_cell.has_style:
                                target_cell.font = copy(source_cell.font)
                                target_cell.border = copy(source_cell.border)
                                target_cell.fill = copy(source_cell.fill)
                                target_cell.number_format = copy(source_cell.number_format)
                                target_cell.protection = copy(source_cell.protection)
                                target_cell.alignment = copy(source_cell.alignment)

                        # Update max discount (column G = 7) - store as number, not decimal
                        upload_ws.cell(row=upload_row, column=7, value=change['newValue'])

                        # Set Operation to 'E' (column AE = 31)
                        upload_ws.cell(row=upload_row, column=31, value='E')

                        upload_row += 1

            # Save the upload file
            # Normalize org name for filename (remove special chars)
            org_filename = org_name.replace(' ', '_').replace('(', '').replace(')', '')
            upload_filename = f"upload_{org_filename}.xlsx"
            upload_path = upload_dir / upload_filename
            upload_wb.save(upload_path)

            wb.close()
            upload_wb.close()

            generated_files.append({
                'org_name': org_name,
                'filename': upload_filename,
                'path': str(upload_path),
                'changes_count': len([c for c in changes_list if c['productCode'] in changed_codes])
            })

        # Generate download URLs
        files_with_urls = []
        for file_info in generated_files:
            # Create a relative path for download
            rel_path = Path(file_info['path']).relative_to(export_root)
            download_url = url_for('max_discount_review.download_upload_file',
                                   filepath=str(rel_path),
                                   _external=False)
            files_with_urls.append({
                'org_name': file_info['org_name'],
                'filename': file_info['filename'],
                'download_url': download_url,
                'path': file_info['path'],  # Full path for upload
                'changes_count': file_info['changes_count']
            })

        return jsonify({
            "success": True,
            "files": files_with_urls
        })

    except Exception as e:
        logger.exception("Error generating upload files")
        return jsonify({"error": str(e)}), 500


@max_discount_review_bp.route("/download/<path:filepath>", methods=["GET"])
@auth.login_required
def download_upload_file(filepath):
    """Download a generated upload file"""
    from flask import send_file

    export_root = Path(current_app.config.get("EXPORT_ROOT", "exports"))
    file_path = export_root / filepath

    if not file_path.exists():
        return jsonify({"error": "File not found"}), 404

    return send_file(file_path, as_attachment=True, download_name=file_path.name)


@max_discount_review_bp.route("/upload-to-buz", methods=["POST"])
@auth.login_required
def upload_to_buz():
    """Upload generated files to Buz"""
    try:
        logger.info("Upload to Buz route called")
        from services.buz_max_discount_review import upload_max_discount_files

        data = request.get_json()
        logger.info(f"Received data: {data}")
        if not data:
            return jsonify({"error": "No data provided"}), 400

        file_paths = data.get('file_paths', {})  # Dict: org_key -> file_path
        headless = data.get('headless', True)

        if not file_paths:
            return jsonify({"error": "No files to upload"}), 400

        # Create a new job
        job_id = str(uuid.uuid4())
        db_path = current_app.config["database"]
        db = create_db_manager(db_path)
        create_job(job_id, db)

        # Run upload in background thread
        def run_upload():
            try:
                # Convert file paths to Path objects
                upload_files = {org_key: Path(fp) for org_key, fp in file_paths.items()}

                def job_callback(pct, message):
                    update_job(job_id, pct=pct, message=message, db=db)

                # Run the async upload
                result = asyncio.run(upload_max_discount_files(
                    upload_files=upload_files,
                    headless=headless,
                    job_update_callback=job_callback
                ))

                # Mark job as done
                update_job(job_id, pct=100, done=True, result=result, db=db)

            except Exception as e:
                logger.exception("Upload to Buz failed")
                update_job(job_id, error=str(e), done=True, db=db)

        thread = threading.Thread(target=run_upload, daemon=True)
        thread.start()

        return jsonify({"job_id": job_id})

    except Exception as e:
        logger.exception("Error in upload_to_buz route")
        return jsonify({"error": f"Server error: {str(e)}"}), 500
