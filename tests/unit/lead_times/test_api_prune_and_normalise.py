# tests/unit/lead_times/test_api_prune_and_normalize.py
from pathlib import Path
import openpyxl

from services.excel_safety import save_workbook_gracefully
from services.lead_times import api


def _mk_book(path: Path, sheets: dict[str, str], *, col_letter="B", row=3):
    wb = openpyxl.Workbook()
    first = True
    for name, text in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        # write in insertion column and an empty anchor col F
        idx = (ord(col_letter.upper()) - 64)
        ws.cell(row=row, column=idx, value=text)
        ws.cell(row=row, column=6, value=False)  # F=FALSE ensures the row is "visible"
    save_workbook_gracefully(wb, str(path))
    wb.close()


def test_prune_unchanged_tabs_cell_based(tmp_path: Path):
    # SAME tab has identical text; DIFF differs in output.
    tpl = tmp_path / "tpl.xlsx"
    out = tmp_path / "out.xlsx"

    _mk_book(tpl, {"SAME": "X", "DIFF": "Old"})
    _mk_book(out, {"SAME": "X", "DIFF": "New"})

    pruned = api._prune_unchanged_tabs_cell_based(
        template_path=tpl,
        output_path=out,
        warnings=[],
        label="UT",
        do_not_show_col_letter="F",
        header_row_1based=2,
        target_col_letter="B",
    )
    # SAME should be removed; DIFF should remain
    assert pruned == ["SAME"]

    wb = openpyxl.load_workbook(str(out), read_only=True)
    try:
        names = set(wb.sheetnames)
    finally:
        wb.close()
    assert "DIFF" in names and "SAME" not in names


def test_normalize_lead_line_spacing_from_template(tmp_path: Path):
    """
    Ensure the normalizer converts real newlines into literal '\\n' and preserves
    the lead line value chunk intact (we only sanity-check the conversion here).
    """
    tpl = tmp_path / "det_tpl.xlsx"
    out = tmp_path / "det_out.xlsx"

    # template: the "desired" spacing style
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ABC"
    ws.cell(row=3, column=2, value="\n       -       Lead Time:  4 - 5 Weeks \n       -       Tail")
    save_workbook_gracefully(wb, str(tpl))
    wb.close()

    # out: same sheet, but any value; the normalizer will align to template spacing
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ABC"
    ws.cell(row=3, column=2, value="\n       -       Lead Time:  6–8 Weeks \n       -       Tail")
    save_workbook_gracefully(wb,str(out))
    wb.close()

    api._normalize_lead_line_spacing_from_template(
        template_path=tpl,
        output_path=out,
        do_not_show_col_letter="F",
        header_row_1based=2,
        target_col_letter="B",
        warnings=[],
        label="DET",
    )

    wb = openpyxl.load_workbook(str(out), read_only=True, data_only=True)
    try:
        ws = wb["ABC"]
        v = ws.cell(row=3, column=2).value
    finally:
        wb.close()
    # literal "\n" tokens present after normalization
    assert "\\n       -       Lead Time:" in v
