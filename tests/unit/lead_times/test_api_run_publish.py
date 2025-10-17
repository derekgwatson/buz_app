# tests/unit/lead_times/test_api_run_publish.py
from pathlib import Path
import types
import sys
import openpyxl
import pytest

from services.lead_times import api
from services.excel_safety import save_workbook_gracefully

# --- lightweight helpers ------------------------------------------------

def _new_book_with_sheet(path: Path, name: str, *, detailed_b=None, summary_c=None, header_row=2):
    """
    Make a tiny workbook with one sheet named `name`.
    If detailed_b provided, write to B{header_row+1}.
    If summary_c provided, write to C{header_row+1}.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name
    r = header_row + 1
    if detailed_b is not None:
        ws.cell(row=r, column=2, value=detailed_b)
    if summary_c is not None:
        ws.cell(row=r, column=3, value=summary_c)
    path.parent.mkdir(parents=True, exist_ok=True)
    save_workbook_gracefully(wb, str(path))
    wb.close()


class DummySheets:
    def service_account_email(self):
        return "svc@example.com"

    def fetch_sheet_data(self, sheet_id: str, a1_range: str):
        # Always return a header + one data row for both lead and cutoff sheets.
        # Column A will be our mapping code ("ABC").
        # We keep it super-generic; codes_from_rows is monkeypatched anyway.
        return [
            ["code", "misc", "more"],
            ["ABC", "x", "y"],
        ]

def _fake_codes_from_rows(rows, map_i):
    return {rows[1][map_i]} if rows and len(rows) > 1 else set()

def _fake_import_and_merge(canberra_rows, regional_rows, cutoff_rows, lead_cols, cutoff_cols):
    # Build the minimal structure run_publish expects for CANBERRA only.
    IR = types.SimpleNamespace
    can = IR(
        by_product_html=[("Prod A", "<p>ignore</p>")],
        product_to_codes={"Prod A": {"ABC"}},
        cutoff_rows={"ABC": {"cutoff": "6/11/25", "product": "Prod A"}},
        lead_rows={"ABC": {"lead_time_text": "6–8 Weeks"}},
        control_codes={"ABC"},
    )
    reg = IR(
        by_product_html=[("RProd", "<p>r</p>")],
        product_to_codes={"RProd": {"ABC"}},
        cutoff_rows={"ABC": {"cutoff": "6/11/25", "product": "RProd"}},
        lead_rows={"ABC": {"lead_time_text": "4–6 Weeks"}},
        control_codes={"ABC"},
    )
    return {"CANBERRA": can, "REGIONAL": reg}

@pytest.fixture(autouse=True)
def _patch_deps(monkeypatch):
    # Fake sheets helpers
    monkeypatch.setattr(api, "codes_from_rows", _fake_codes_from_rows)
    monkeypatch.setattr(api, "import_and_merge", _fake_import_and_merge)

    # Provide a dummy HTML builder module so `from services.lead_times.html_out import build_pasteable_html` works.
    mod_name = "services.lead_times.html_out"
    pkg_name = "services.lead_times"
    if pkg_name not in sys.modules:
        sys.modules[pkg_name] = types.ModuleType(pkg_name)
    html_mod = types.ModuleType(mod_name)
    html_mod.build_pasteable_html = lambda *args, **kwargs: "<p>ok</p>"
    sys.modules[mod_name] = html_mod

def test_run_publish_applies_banner_and_updates_values(tmp_path: Path):
    # Templates (one sheet "ABC" each). Use *real* newlines here; API normalizer will convert to literal '\n'.
    detailed_tmpl = tmp_path / "tmpl_detailed.xlsx"
    summary_tmpl  = tmp_path / "tmpl_summary.xlsx"
    _new_book_with_sheet(
        detailed_tmpl, "ABC",
        detailed_b="\n       -       Lead Time:  4 - 5 Weeks \n       -       Location:  Canberra"
    )
    _new_book_with_sheet(
        summary_tmpl, "ABC",
        summary_c="Ready in 2 - 3 Weeks, manufactured locally"
    )

    cfg = {
        "lead_times_ss": {
            "sheet_id": "LT1",
            "tabs": {"canberra": "CAN", "regional": "REG"},
            "header_row": 1,
        },
        "cutoffs": {
            "sheet_id": "CO1",
            "tab": "CUT",
            "header_row": 1,
        },
        "columns": {
            "lead_times_ss": {"mapping": "A"},
            "cutoff": {"mapping": "A"},
            "do_not_show_header": "F",
            "do_not_show_header_row": 2,
        },
        "insertion_columns": {"detailed": "B", "summary": "C"},
        "filename_patterns": {},  # use defaults
        "templates": {
            # leave detailed_prefix_template empty; we already have a 'Lead Time:' line
        },
    }

    gs = DummySheets()
    out_dir = tmp_path / "out"

    result = api.run_publish(
        gsheets_service=gs,
        lead_times_cfg=cfg,
        detailed_template_path=str(detailed_tmpl),
        summary_template_path=str(summary_tmpl),
        save_dir=str(out_dir),
        scope="CANBERRA",  # keep the integration tight
    )

    # Files were emitted
    assert "canberra_detailed" in result["files"]
    assert "canberra_summary" in result["files"]

    # Open the actual outputs and validate contents
    det_path = out_dir / result["files"]["canberra_detailed"]
    sum_path = out_dir / result["files"]["canberra_summary"]

    wb = openpyxl.load_workbook(str(det_path), data_only=True, read_only=True)
    try:
        ws = wb["ABC"]
        v = ws.cell(row=3, column=2).value  # B3
    finally:
        wb.close()

    # Detailed: banner is *prepended* with a literal '\n', and the lead-time value is updated.
    assert v.startswith("\\n***CHRISTMAS CUTOFF 6/11/25 ***")
    assert "Lead Time:" in v and "6–8 Weeks" in v
    # Normalizer writes literal '\n' tokens
    assert "\\n       -       Location:  Canberra" in v

    wb = openpyxl.load_workbook(str(sum_path), data_only=True, read_only=True)
    try:
        ws = wb["ABC"]
        v = ws.cell(row=3, column=3).value  # C3
    finally:
        wb.close()

    # Summary: lead-time value updated, banner *appended* with a single space
    assert v == "Ready in 6–8 Weeks, manufactured locally ***CHRISTMAS CUTOFF 6/11/25 ***"
