# tests/test_inject_and_prune.py
from pathlib import Path
import openpyxl
import pytest

from services.lead_times.excel_out import inject_and_prune, InjectResult

# helpers -------------------------------------------------------------

def _new_book_with_sheet(path: Path, name: str, header_row: int, *,
                         detailed_cell_b=None, summary_cell_c=None,
                         anchor_false_row: int | None = None):
    """
    Create a minimal template workbook with one sheet:
    - header_row defines the header line index (1-based)
    - if detailed_cell_b is provided, write it at row header_row+1, col B
    - if summary_cell_c is provided, write it at row header_row+1, col C
    - if anchor_false_row is provided, write FALSE into col F at that row
    """
    wb = openpyxl.Workbook()
    # openpyxl creates default sheet; rename or recreate cleanly
    ws = wb.active
    ws.title = name

    r = header_row + 1
    if detailed_cell_b is not None:
        ws.cell(row=r, column=2, value=detailed_cell_b)  # B
    if summary_cell_c is not None:
        ws.cell(row=r, column=3, value=summary_cell_c)   # C
    if anchor_false_row is not None:
        ws.cell(row=anchor_false_row, column=6, value=False)  # F = Do Not Show? FALSE

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()


# tests ---------------------------------------------------------------

def test_inject_detailed_existing_lead(tmp_path: Path):
    template = tmp_path / "tmpl_detailed.xlsx"
    out = tmp_path / "out_detailed.xlsx"

    # Put a Lead Time line in B at row 3 (header_row=2)
    original = "\n some other random stuff first\n       -       Lead Time:  4 - 5 Weeks \n       -       Location:  Canberra"
    _new_book_with_sheet(template, "ABC", header_row=2, detailed_cell_b=original)

    # Prepare leads map
    leads = {"ABC": {"lead_time_text": "6–8 Weeks"}}

    res: InjectResult = inject_and_prune(
        template_path=template,
        out_path=out,
        store_name="CANBERRA",
        leads_by_code=leads,
        insertion_col_letter="B",
        anchor_col_letter="F",
        anchor_header_row=2,
        control_codes={"ABC"},
        warnings=[],
        workbook_kind="Detailed",
        detailed_prefix_template=None,
    )

    wb = openpyxl.load_workbook(str(res.saved_path), read_only=True, data_only=True)
    try:
        ws = wb["ABC"]
        val = ws.cell(row=3, column=2).value
    finally:
        wb.close()

    assert val == "\n some other random stuff first\n       -       Lead Time:  6–8 Weeks \n       -       Location:  Canberra"


def test_inject_detailed_no_lead_prefixes_at_false(tmp_path: Path):
    template = tmp_path / "tmpl_detailed2.xlsx"
    out = tmp_path / "out_detailed2.xlsx"

    # No 'Lead Time:' line; FALSE at F3; existing text continues after template.
    existing_tail = "Some other detail text"
    _new_book_with_sheet(template, "DEF", header_row=2,
                         detailed_cell_b=existing_tail,
                         anchor_false_row=3)

    leads = {"DEF": {"lead_time_text": "3–4 Weeks"}}

    res: InjectResult = inject_and_prune(
        template_path=template,
        out_path=out,
        store_name="REGIONAL",
        leads_by_code=leads,
        insertion_col_letter="B",
        anchor_col_letter="F",
        anchor_header_row=2,
        control_codes={"DEF"},
        warnings=[],
        workbook_kind="Detailed",
        detailed_prefix_template="\n       -       Lead Time: {LEAD} \n       -       ",
    )

    wb = openpyxl.load_workbook(str(res.saved_path), read_only=True, data_only=True)
    try:
        ws = wb["DEF"]
        val = ws.cell(row=3, column=2).value
    finally:
        wb.close()

    assert val.startswith("\n       -       Lead Time: 3–4 Weeks \n       -       ")
    assert val.endswith(existing_tail)


def test_inject_summary_updates_value_only_keeps_brackets_and_suffix(tmp_path: Path):
    template = tmp_path / "tmpl_summary.xlsx"
    out = tmp_path / "out_summary.xlsx"

    # Summary row in C3 (header_row=2)
    original = "Ready in 2 - 3 Weeks (PC Hardware 4-5 Weeks, Others 6 weeks), manufactured locally"
    _new_book_with_sheet(template, "GHI", header_row=2, summary_cell_c=original)

    leads = {"GHI": {"lead_time_text": "6–8 Weeks"}}

    res: InjectResult = inject_and_prune(
        template_path=template,
        out_path=out,
        store_name="CANBERRA",
        leads_by_code=leads,
        insertion_col_letter="C",
        anchor_col_letter="F",
        anchor_header_row=2,
        control_codes={"GHI"},
        warnings=[],
        workbook_kind="Summary",
        detailed_prefix_template=None,
    )

    wb = openpyxl.load_workbook(str(res.saved_path), read_only=True, data_only=True)
    try:
        ws = wb["GHI"]
        val = ws.cell(row=3, column=3).value
    finally:
        wb.close()

    assert val == "Ready in 6–8 Weeks (PC Hardware 4-5 Weeks, Others 6 weeks), manufactured locally"


def test_inject_summary_basic_suffix_after_comma(tmp_path: Path):
    template = tmp_path / "tmpl_summary2.xlsx"
    out = tmp_path / "out_summary2.xlsx"

    original = "Ready in 4 - 5 Weeks, manufactured locally, lifetime warranty"
    _new_book_with_sheet(template, "JKL", header_row=2, summary_cell_c=original)

    leads = {"JKL": {"lead_time_text": "6–8 Weeks"}}

    res: InjectResult = inject_and_prune(
        template_path=template,
        out_path=out,
        store_name="REGIONAL",
        leads_by_code=leads,
        insertion_col_letter="C",
        anchor_col_letter="F",
        anchor_header_row=2,
        control_codes={"JKL"},
        warnings=[],
        workbook_kind="Summary",
        detailed_prefix_template=None,
    )

    wb = openpyxl.load_workbook(str(res.saved_path), read_only=True, data_only=True)
    try:
        ws = wb["JKL"]
        val = ws.cell(row=3, column=3).value
    finally:
        wb.close()

    assert val == "Ready in 6–8 Weeks, manufactured locally, lifetime warranty"
