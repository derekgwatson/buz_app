import pytest
from services.excel import OpenPyXLFileHandler
from services.group_options_check import extract_codes_from_excel_flat_dedup
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from services.excel_safety import save_workbook_gracefully


@pytest.fixture
def create_test_excel_file():
    """Fixture to create a test Excel file dynamically."""
    def _create_test_file(sheet_data):
        wb = Workbook()
        for sheet_name, data in sheet_data.items():
            ws = wb.create_sheet(sheet_name)
            for row_idx, row in enumerate(data, start=1):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
        save_workbook_gracefully(wb, temp_file.name)
        return temp_file.name

    return _create_test_file


class TestExtractCodes:
    """Class-based tests for extract_codes_from_excel_flat_dedup."""

    def test_extract_codes_basic(self, create_test_excel_file):
        """Tests the basic functionality of extracting codes."""
        sheet_data = {
            "Sheet1": [
                ["", "", "", ""],  # Row 1 (ignored)
                ["Inventory Code for Pricing", "", "", ""],  # Row 2
                *[["", "", "", ""] for _ in range(3)],  # Row 3-5 (ignored)
                ["", "", "Value in Row 6", ""],  # Row 6
                *[["", "", "", ""] for _ in range(10)],  # Row 7-16 (ignored)
                ["", "", "Code|First|Second", "Another|Value"],  # Row 17
            ]
        }

        file_handler = OpenPyXLFileHandler.from_file(file_path=create_test_excel_file(sheet_data))
        result = extract_codes_from_excel_flat_dedup(file_handler)

        assert result == [("Sheet1", "Second"), ("Sheet1", "Value")]

    def test_ignore_sheets(self, create_test_excel_file):
        """Tests that sheets without the required text in A2 are ignored."""
        sheet_data = {
            "Sheet1": [["Wrong Header", "", "", ""], ["", "", "", ""]],
            "Sheet2": [
                ["", "", "", ""],  # Row 1 (ignored)
                ["Inventory Code for Pricing", "", "", ""],  # Row 2
                *[["", "", "", ""] for _ in range(14)],  # Row 3-16 (ignored)
                ["", "", "Code|First", ""],  # Row 17
            ],
        }

        file_handler = OpenPyXLFileHandler.from_file(file_path=create_test_excel_file(sheet_data))
        result = extract_codes_from_excel_flat_dedup(file_handler)

        assert result == [("Sheet2", "First")]

    def test_deduplication(self, create_test_excel_file):
        """Tests that duplicate codes are removed."""
        sheet_data = {
            "Sheet1": [
                ["", "", "", ""],  # Row 1 (ignored)
                ["Inventory Code for Pricing", "", "", ""],  # Row 2
                *[["", "", "", ""] for _ in range(14)],  # Row 3-16 (ignored)
                ["", "", "Code|First", "Code|First"],  # Row 17
            ],
        }

        file_handler = OpenPyXLFileHandler.from_file(file_path=create_test_excel_file(sheet_data))
        result = extract_codes_from_excel_flat_dedup(file_handler)

        assert result == [("Sheet1", "First")]
